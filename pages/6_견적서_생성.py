# ìš•ì‹¤ ê²¬ì ì„œ ìƒì„±ê¸°
# session_state ì—°ë™ ë²„ì „ - ë°”ë‹¥/ë²½/ì²œì¥ ê³„ì‚° ê²°ê³¼ë¥¼ ìë™ìœ¼ë¡œ ê°€ì ¸ì˜µë‹ˆë‹¤.

from common_styles import apply_common_styles, set_page_config
from common_sidebar import render_chatbot_sidebar
import auth

import json
import io
from typing import Dict, Any, List, Optional
from datetime import datetime
import pandas as pd
import streamlit as st

# Session state keys
FLOOR_RESULT_KEY = "floor_result"
WALL_RESULT_KEY = "wall_result"
CEIL_RESULT_KEY = "ceil_result"
SAVED_QUOTATIONS_KEY = "saved_quotations"  # ì €ì¥ëœ ì„¸ëŒ€ íƒ€ì…ë³„ ê²¬ì  ëª©ë¡ (ìµœëŒ€ 10ê°œ)
PROD_MGMT_SETTINGS_KEY = "prod_mgmt_settings"  # ìƒì‚°ê´€ë¦¬ë¹„ ì„¤ì •

# ê³µìœ  ì—‘ì…€ íŒŒì¼ í‚¤ (ë°”ë‹¥íŒì—ì„œ ì—…ë¡œë“œ)
SHARED_EXCEL_KEY = "shared_excel_file"
SHARED_EXCEL_NAME_KEY = "shared_excel_filename"

# ìƒì‚°ê´€ë¦¬ë¹„ ê¸°ë³¸ ì¹´í…Œê³ ë¦¬ ì •ì˜ (í’ˆëª©+ì‚¬ì–‘ ë‹¨ìœ„ë¡œ ì„¸ë¶€ ì§€ì •)
# items: [(í’ˆëª©, ì‚¬ì–‘íŒ¨í„´), ...] - ì‚¬ì–‘íŒ¨í„´ì´ Noneì´ë©´ í•´ë‹¹ í’ˆëª© ì „ì²´, ë¬¸ìì—´ì´ë©´ contains ë§¤ì¹­
DEFAULT_PROD_MGMT_CATEGORIES = {
    "íšŒì‚¬ìƒì‚°í’ˆ(ë°”ë‹¥íŒ,ìš•ì¡°)": {
        "items": [
            ("ë°”ë‹¥íŒ", "GRP"),
            ("ë°”ë‹¥íŒ", "FRP"),  # FRP í¬í•¨ (SMC/FRP ë“±)
            ("ë°”ë‹¥íŒ", "SMC"),  # SMC í¬í•¨
            ("ë°”ë‹¥íŒ", "PP"),  # PP/PE í¬í•¨
            ("ìš•ì¡°", None),  # ìš•ì¡° ì „ì²´
        ],
        "rate": 20.0,  # ê¸°ë³¸ê°’ 20%
    },
    "íšŒì‚¬ìƒì‚°í’ˆ(ì²œì¥íŒ)": {
        "items": [
            ("ì²œì¥íŒ", None),  # ì²œì¥íŒ ì „ì²´
        ],
        "rate": 15.0,  # ê¸°ë³¸ê°’ 15%
    },
    "íšŒì‚¬-ëª…ì§„(ë²½,PVEë°”ë‹¥íŒ)": {
        "items": [
            ("ë²½íŒ", None),  # ë²½íŒ ì „ì²´
            ("ë°”ë‹¥íŒ", "PVE"),  # PVE ë°”ë‹¥íŒë§Œ
        ],
        "rate": 15.0,  # ê¸°ë³¸ê°’ 15%
    },
    "íƒ€ì‚¬(ì²œì¥,ë°”ë‹¥íŒ,íƒ€ì¼)": {
        "items": [
            ("íƒ€ì¼", None),  # íƒ€ì¼ ì „ì²´
        ],
        "rate": 5.0,  # ê¸°ë³¸ê°’ 5%
    },
    "íƒ€ì‚¬(ë„ê¸°,ìˆ˜ì „,ê¸°íƒ€)": {
        "items": [
            ("ë„ê¸°ë¥˜", None),
            ("ìˆ˜ì „", None),
            ("ì•¡ì„¸ì„œë¦¬", None),
            ("ë¬¸ì„¸íŠ¸", None),
            ("ìš•ì‹¤ë“±", None),
            ("ê³µí†µìì¬", None),
            ("ëƒ‰ì˜¨ìˆ˜ë°°ê´€", None),
            ("ë¬¸í‹€ê·œê²©", None),
            ("ì€ê²½", None),
            ("ìš•ì‹¤ì¥", None),
            ("ì¹¸ë§‰ì´", None),
            ("í™˜ê¸°ë¥˜", None),
        ],
        "rate": 5.0,  # ê¸°ë³¸ê°’ 5%
    },
}

# ì˜ì—…ê´€ë¦¬ë¹„ ì„¤ì • í‚¤
SALES_MGMT_SETTINGS_KEY = "sales_mgmt_settings"


def get_item_key(í’ˆëª©: str, ì‚¬ì–‘: str) -> str:
    """í’ˆëª©+ì‚¬ì–‘ì„ ê³ ìœ  í‚¤ë¡œ ë³€í™˜"""
    return f"{í’ˆëª©}::{ì‚¬ì–‘}"


def parse_item_key(key: str) -> tuple:
    """ê³ ìœ  í‚¤ë¥¼ í’ˆëª©, ì‚¬ì–‘ìœ¼ë¡œ ë¶„ë¦¬"""
    parts = key.split("::", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (parts[0], "")


def item_matches_pattern(
    í’ˆëª©: str, ì‚¬ì–‘: str, pattern_í’ˆëª©: str, pattern_ì‚¬ì–‘: Optional[str]
) -> bool:
    """í’ˆëª©+ì‚¬ì–‘ì´ íŒ¨í„´ê³¼ ë§¤ì¹­ë˜ëŠ”ì§€ í™•ì¸"""
    if í’ˆëª© != pattern_í’ˆëª©:
        return False
    if pattern_ì‚¬ì–‘ is None:
        return True  # í’ˆëª©ë§Œ ë§¤ì¹­í•˜ë©´ ì „ì²´ í¬í•¨
    return pattern_ì‚¬ì–‘.upper() in ì‚¬ì–‘.upper()


set_page_config(page_title="ìš•ì‹¤ ê²¬ì ì„œ ìƒì„±ê¸°", layout="wide")
apply_common_styles()

auth.require_auth()

# ì‚¬ì´ë“œë°”ì— ì‹œë°©ì„œ ë¶„ì„ ê²°ê³¼ í‘œì‹œ
render_chatbot_sidebar()

# ----------------------------
# Helper Functions
# ----------------------------
REQ_COLUMNS = ["í’ˆëª©", "ë¶„ë¥˜", "ì‚¬ì–‘ ë° ê·œê²©", "ë‹¨ê°€", "ìˆ˜ëŸ‰"]


@st.cache_data(show_spinner=False)
def load_pricebook_from_excel(
    file_bytes: bytes, sheet_name: str = "ìì¬ë‹¨ê°€ë‚´ì—­"
) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    # Normalize columns - ì •í™•íˆ "ëŒ€ë¶„ë¥˜", "ì¤‘ë¶„ë¥˜", "ì‚¬ì–‘ ë° ê·œê²©"ë§Œ ì¸ì‹
    colmap = {}
    for c in df.columns:
        c2 = str(c).strip()
        if c2 == "ëŒ€ë¶„ë¥˜":
            colmap[c] = "í’ˆëª©"
        elif c2 == "ì¤‘ë¶„ë¥˜":
            colmap[c] = "ë¶„ë¥˜"
        elif c2 == "ì‚¬ì–‘ ë° ê·œê²©":
            colmap[c] = "ì‚¬ì–‘ ë° ê·œê²©"
        elif c2 == "ë‹¨ê°€":
            colmap[c] = "ë‹¨ê°€"
        elif c2 == "ìˆ˜ëŸ‰":
            colmap[c] = "ìˆ˜ëŸ‰"
        elif c2 == "ê¸ˆì•¡":
            colmap[c] = "ê¸ˆì•¡"
    df = df.rename(columns=colmap)
    # Ensure required columns exist
    for c in ["í’ˆëª©", "ë¶„ë¥˜", "ì‚¬ì–‘ ë° ê·œê²©"]:
        if c not in df.columns:
            df[c] = ""
    for c in ["ë‹¨ê°€", "ìˆ˜ëŸ‰"]:
        if c not in df.columns:
            df[c] = 0
    # Clean values
    for c in ["í’ˆëª©", "ë¶„ë¥˜", "ì‚¬ì–‘ ë° ê·œê²©"]:
        df[c] = df[c].fillna("").astype(str).str.strip()
    for c in ["ë‹¨ê°€", "ìˆ˜ëŸ‰"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    if "ê¸ˆì•¡" not in df.columns:
        df["ê¸ˆì•¡"] = df["ë‹¨ê°€"].fillna(0) * df["ìˆ˜ëŸ‰"].fillna(0)
    return df


@st.cache_data(show_spinner=False)
def load_auto_items_structure(file_bytes: bytes) -> Dict[str, Any]:
    """
    Excelì˜ 'ìë™ì§€ì •í•­ëª©' ì‹œíŠ¸ì—ì„œ ìë™ì§€ì • í’ˆëª© êµ¬ì¡°ë¥¼ ë¡œë“œ

    ì‹œíŠ¸ êµ¬ì¡°:
    - ëŒ€ë¶„ë¥˜: í’ˆëª©ì˜ ëŒ€ë¶„ë¥˜
    - ì¤‘ë¶„ë¥˜: í’ˆëª©ì˜ ì¤‘ë¶„ë¥˜ (ì—†ìœ¼ë©´ NaN)
    - ì‚¬ì–‘ ë° ê·œê²©: í’ˆëª©ì˜ ê·œê²© (ì—†ìœ¼ë©´ NaN)

    Returns:
        {
            "ëŒ€ë¶„ë¥˜ëª…": {
                "subcategories": {ì¤‘ë¶„ë¥˜: [ê·œê²©1, ê·œê²©2]} ë˜ëŠ” None,
                "category_map": "ëŒ€ë¶„ë¥˜ëª…"
            }
        }
    """
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="ìë™ì§€ì •í•­ëª©")
    except Exception:
        # ì‹œíŠ¸ê°€ ì—†ìœ¼ë©´ ë¹ˆ ë”•ì…”ë„ˆë¦¬ ë°˜í™˜
        return {}

    # ì»¬ëŸ¼ëª… í™•ì¸
    if "ëŒ€ë¶„ë¥˜" not in df.columns:
        return {}

    # NaN ì²˜ë¦¬
    df = df.copy()
    df["ëŒ€ë¶„ë¥˜"] = df["ëŒ€ë¶„ë¥˜"].fillna("").astype(str).str.strip()
    df["ì¤‘ë¶„ë¥˜"] = df.get("ì¤‘ë¶„ë¥˜", pd.Series()).fillna("").astype(str).str.strip()
    df["ì‚¬ì–‘ ë° ê·œê²©"] = df.get("ì‚¬ì–‘ ë° ê·œê²©", pd.Series()).fillna("").astype(str).str.strip()

    # ë¹ˆ ëŒ€ë¶„ë¥˜ ì œê±°
    df = df[df["ëŒ€ë¶„ë¥˜"] != ""]

    # ëŒ€ë¶„ë¥˜ë³„ë¡œ êµ¬ì¡° ìƒì„±
    structure = {}

    for major_cat in df["ëŒ€ë¶„ë¥˜"].unique():
        rows = df[df["ëŒ€ë¶„ë¥˜"] == major_cat]

        # ì¤‘ë¶„ë¥˜ê°€ ëª¨ë‘ ë¹„ì–´ìˆëŠ”ì§€ í™•ì¸
        has_subcategories = (rows["ì¤‘ë¶„ë¥˜"] != "").any()

        if not has_subcategories:
            # ì¼€ì´ìŠ¤ 1: ì¤‘ë¶„ë¥˜ ì—†ìŒ (ì˜ˆ: GRPë°”ë‹¥íŒ)
            structure[major_cat] = {
                "subcategories": None,
                "category_map": major_cat,
            }
        else:
            # ì¼€ì´ìŠ¤ 2: ì¤‘ë¶„ë¥˜ ìˆìŒ
            subcategories = {}

            for sub_cat in rows["ì¤‘ë¶„ë¥˜"].unique():
                if sub_cat == "":
                    continue

                sub_rows = rows[rows["ì¤‘ë¶„ë¥˜"] == sub_cat]

                # ì‚¬ì–‘ ë° ê·œê²© ë¦¬ìŠ¤íŠ¸ ìƒì„± (ë¹ˆ ê°’ ì œì™¸)
                specs = [
                    str(spec).strip() for spec in sub_rows["ì‚¬ì–‘ ë° ê·œê²©"].tolist()
                    if spec and str(spec).strip() and str(spec).strip().lower() != 'nan'
                ]

                subcategories[sub_cat] = specs

            structure[major_cat] = {
                "subcategories": subcategories,
                "category_map": major_cat,
            }

    return structure


@st.cache_data(show_spinner=False)
def load_ceiling_drilling_prices(file_bytes: bytes) -> Dict[str, float]:
    """ì²œì¥íŒíƒ€ê³µ ì‹œíŠ¸ì—ì„œ ê°€ê³µë¹„ ë‹¨ê°€ë¥¼ ë¡œë“œ (ë°”ë””/ì‚¬ì´ë“œ ì ˆë‹¨ í•­ëª© ì œì™¸)"""
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="ì²œì¥íŒíƒ€ê³µ")
        prices = {}
        for _, row in df.iterrows():
            name = str(row.get("í’ˆëª©", "")).strip()
            # ë°”ë””/ì‚¬ì´ë“œëŠ” ì ˆë‹¨ ë¹„ìš©ì´ë¯€ë¡œ íƒ€ê³µ ê°€ê³µë¹„ì—ì„œ ì œì™¸
            if name in ("ë°”ë””", "ì‚¬ì´ë“œ"):
                continue
            price = pd.to_numeric(row.get("ë‹¨ê°€", 0), errors="coerce") or 0
            if name:
                prices[name] = float(price)
        return prices
    except Exception:
        # ì‹œíŠ¸ê°€ ì—†ê±°ë‚˜ ì˜¤ë¥˜ì‹œ ê¸°ë³¸ê°’ ë°˜í™˜
        return {
            "í™˜í’ê¸°í™€": 4000,
            "ì‚¬ê°ë§¤ë¦½ë“±": 4000,
            "ì›í˜•ë“±": 2000,
            "ì§ì„ 1íšŒ": 1500,
        }


def find_item(
    df: pd.DataFrame,
    ëŒ€ë¶„ë¥˜: str,
    ì‚¬ì–‘_ë°_ê·œê²©_ì „ì²´: str,
) -> Optional[pd.Series]:
    """
    ë‹¨ê°€í‘œì—ì„œ í’ˆëª© ì°¾ê¸° (ëŒ€ë¶„ë¥˜ + í•©ì³ì§„ ì‚¬ì–‘)

    Args:
        df: ë‹¨ê°€í‘œ DataFrame
        ëŒ€ë¶„ë¥˜: í’ˆëª© ê°’ (ëŒ€ë¶„ë¥˜)
        ì‚¬ì–‘_ë°_ê·œê²©_ì „ì²´: "ì¤‘ë¶„ë¥˜ ì‚¬ì–‘ë°ê·œê²©" í˜•ì‹ì˜ í•©ì³ì§„ ë¬¸ìì—´

    Returns:
        ë§¤ì¹­ëœ í–‰ ë˜ëŠ” None
    """
    ëŒ€ë¶„ë¥˜_term = str(ëŒ€ë¶„ë¥˜).strip()
    ì‚¬ì–‘_term = str(ì‚¬ì–‘_ë°_ê·œê²©_ì „ì²´).strip()

    # ëŒ€ë¶„ë¥˜ í•„í„°ë§ (ëŒ€ì†Œë¬¸ì êµ¬ë¶„ ì—†ìŒ)
    ëŒ€ë¶„ë¥˜_matches = df[df["í’ˆëª©"].fillna("").astype(str).str.strip().str.lower() == ëŒ€ë¶„ë¥˜_term.lower()]

    if len(ëŒ€ë¶„ë¥˜_matches) == 0:
        return None

    # ì‚¬ì–‘ì´ ì—†ìœ¼ë©´ ì²« ë²ˆì§¸ ë§¤ì¹­ ë°˜í™˜
    if not ì‚¬ì–‘_term:
        return ëŒ€ë¶„ë¥˜_matches.iloc[0]

    # ë‹¨ê°€í‘œì˜ ì¤‘ë¶„ë¥˜ + ì‚¬ì–‘ ë° ê·œê²©ì„ í•©ì³ì„œ ë¹„êµ
    df_ì¤‘ë¶„ë¥˜ = ëŒ€ë¶„ë¥˜_matches["ë¶„ë¥˜"].fillna("").astype(str).str.strip()
    df_ì‚¬ì–‘ = ëŒ€ë¶„ë¥˜_matches["ì‚¬ì–‘ ë° ê·œê²©"].fillna("").astype(str).str.strip()

    # ë§¤ì¹­ ì „ëµ 1: ì¤‘ë¶„ë¥˜ + " " + ì‚¬ì–‘ (ê³µë°± í¬í•¨, ëŒ€ì†Œë¬¸ì êµ¬ë¶„ ì—†ìŒ)
    df_combined_space = (df_ì¤‘ë¶„ë¥˜ + " " + df_ì‚¬ì–‘).str.strip()
    mask1 = (df_combined_space.str.lower() == ì‚¬ì–‘_term.lower())

    # ë§¤ì¹­ ì „ëµ 2: ì¤‘ë¶„ë¥˜ + ì‚¬ì–‘ (ê³µë°± ì—†ìŒ, ëŒ€ì†Œë¬¸ì êµ¬ë¶„ ì—†ìŒ)
    df_combined_no_space = (df_ì¤‘ë¶„ë¥˜ + df_ì‚¬ì–‘).str.strip()
    mask2 = (df_combined_no_space.str.lower() == ì‚¬ì–‘_term.replace(" ", "").lower())

    # ë§¤ì¹­ ì „ëµ 3: ì‚¬ì–‘ ë° ê·œê²©ë§Œ ë§¤ì¹­ (ì¤‘ë¶„ë¥˜ ë¬´ì‹œ, ëŒ€ì†Œë¬¸ì êµ¬ë¶„ ì—†ìŒ)
    mask3 = (df_ì‚¬ì–‘.str.lower() == ì‚¬ì–‘_term.lower())

    # ë§¤ì¹­ ì „ëµ 4: í¬í•¨ ê²€ìƒ‰ (ì‚¬ì–‘ ë° ê·œê²©ì´ ê²€ìƒ‰ì–´ë¥¼ í¬í•¨, ëŒ€ì†Œë¬¸ì êµ¬ë¶„ ì—†ìŒ)
    mask4 = df_combined_space.str.lower().str.contains(ì‚¬ì–‘_term.lower(), regex=False, na=False)

    # ìš°ì„ ìˆœìœ„ëŒ€ë¡œ ë§¤ì¹­ ì‹œë„
    for mask in [mask1, mask2, mask3, mask4]:
        candidates = ëŒ€ë¶„ë¥˜_matches[mask]
        if len(candidates) > 0:
            return candidates.iloc[0]

    return None


def add_row(
    rows: List[Dict[str, Any]],
    í’ˆëª©: str,
    spec: str,
    qty: float,
    unit_price: Optional[float],
) -> None:
    unit_price = unit_price if unit_price is not None else 0
    amount = (qty or 0) * (unit_price or 0)
    rows.append(
        {
            "í’ˆëª©": í’ˆëª©,
            "ì‚¬ì–‘ ë° ê·œê²©": spec,
            "ìˆ˜ëŸ‰": qty,
            "ë‹¨ê°€": unit_price,
            "ê¸ˆì•¡": amount,
        }
    )


def add_all_by_category(
    rows: List[Dict[str, Any]], df: pd.DataFrame, í’ˆëª©: str, ë¶„ë¥˜: str
):
    sub = df[(df["í’ˆëª©"] == í’ˆëª©) & (df["ë¶„ë¥˜"] == ë¶„ë¥˜)]
    for _, r in sub.iterrows():
        add_row(
            rows,
            í’ˆëª©,
            str(r["ì‚¬ì–‘ ë° ê·œê²©"]),
            r["ìˆ˜ëŸ‰"] if pd.notna(r["ìˆ˜ëŸ‰"]) else 1,
            r["ë‹¨ê°€"] if pd.notna(r["ë‹¨ê°€"]) else 0,
        )


# ----------------------------
# Convert session_state to quotation format
# ----------------------------
def convert_floor_data(floor_result: dict) -> dict:
    """Convert floor_result to quotation format"""
    if not floor_result:
        return {}

    # session_state êµ¬ì¡°: {"section", "inputs", "result", "decision_log"}
    inputs = floor_result.get("inputs", {})
    result = floor_result.get("result", {})

    # ì†Œì¬ ì •ë³´ ì¶”ì¶œ (resultì—ì„œ)
    material = result.get("ì†Œì¬", "")
    # "PP/PE ë°”ë‹¥íŒ" -> "PP/PE" ë³€í™˜
    material_clean = material.replace(" ë°”ë‹¥íŒ", "").replace("ë°”ë‹¥íŒ", "").strip()

    # ê°€ê²© ì •ë³´ ì¶”ì¶œ (resultì—ì„œ) - ì†Œê³„ ì‚¬ìš©
    ë‹¨ê°€ = result.get("ì†Œê³„", 0)

    # ì„¸ëŒ€ìˆ˜ ì •ë³´ (inputsì—ì„œ)
    units = inputs.get("units", 1)

    # ê·œê²© ë¬¸ìì—´ ìƒì„±
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)
    spec = f"{W}Ã—{L}" if W and L else ""

    return {
        "ì¬ì§ˆ": material_clean,
        "ê·œê²©": spec,
        "ìˆ˜ëŸ‰": 1,  # ì„¸ëŒ€ë‹¹ 1ê°œ (ì„¸ëŒ€ìˆ˜ëŠ” inputsì— ìœ ì§€)
        "ë‹¨ê°€": ë‹¨ê°€,
        "ì£¼ê±°ì•½ì": inputs.get("user_type", "") == "ì£¼ê±°ì•½ì",
        "inputs": inputs,  # inputs ì •ë³´ ìœ ì§€ (ì„¸ëŒ€ìˆ˜ ë“±)
    }


def convert_wall_data(wall_result: dict) -> dict:
    """Convert wall_result to quotation format"""
    # ë²½íŒ ì›ê°€ ê³„ì‚° ê²°ê³¼ í™•ì¸ (wall_panel_cost_final.pyì˜ ê²°ê³¼)
    wall_cost_result = st.session_state.get("shared_wall_cost_result", {})

    if wall_cost_result:
        # ë²½íŒ ì›ê°€ ê³„ì‚° ê²°ê³¼ê°€ ìˆìœ¼ë©´ í•´ë‹¹ ê°’ ì‚¬ìš©
        total_panels = int(wall_cost_result.get("ì´íŒë„¬ìˆ˜", 0))
        unit_price = wall_cost_result.get("íŒë„¬1ì¥ë‹¹_ìƒì‚°ì›ê°€ê³„(AD)", 0)
        production_cost = wall_cost_result.get("ìš•ì‹¤1ì„¸íŠ¸_ìƒì‚°ì›ê°€ê³„(AD)", 0)
        # ë²½íƒ€ì¼ ì •ë³´ëŠ” wall_resultì—ì„œ ê°€ì ¸ì˜¤ê±°ë‚˜ ê¸°ë³¸ê°’ ì‚¬ìš©
        tile = "300Ã—600"
        if wall_result:
            inputs = wall_result.get("inputs", {})
            tile = inputs.get("tile", "300Ã—600")
        return {
            "ì´ê°œìˆ˜": total_panels,
            "ë‹¨ê°€": unit_price,
            "ë²½íƒ€ì¼": tile,
            "production_cost": production_cost,
        }

    # wall_cost_resultê°€ ì—†ìœ¼ë©´ ê¸°ì¡´ ë°©ì‹: wall_panel_final.pyì˜ ê²°ê³¼ ì‚¬ìš©
    if not wall_result:
        return {}

    result = wall_result.get("result", {})
    inputs = wall_result.get("inputs", {})
    unit_price = result.get("ad_per_panel", 0)
    total_panels = int(result.get("total_panels", 0))
    production_cost = result.get("production_cost", 0)

    return {
        "ì´ê°œìˆ˜": total_panels,
        "ë‹¨ê°€": unit_price,
        "ë²½íƒ€ì¼": inputs.get("tile", "300Ã—600"),
        "production_cost": production_cost,
    }


def convert_ceiling_data(ceil_result: dict) -> dict:
    """Convert ceil_result to quotation format"""
    if not ceil_result:
        return {}

    # ceil_panel_final.pyì˜ session_state êµ¬ì¡°ì— ë§ì¶° íŒŒì‹±
    inputs = ceil_result.get("inputs", {})
    result = ceil_result.get("result", {})

    # ì¬ì§ˆ ì •ë³´ ì¶”ì¶œ (inputsì—ì„œ)
    material = inputs.get("material", "GRP")  # GRP/FRP/ê¸°íƒ€

    # ì†Œê³„ ì‚¬ìš© (ê´€ë¦¬ë¹„ ì œì™¸)
    subtotal = result.get("ì†Œê³„", 0)

    # JSON export ë°ì´í„° ì‚¬ìš© (ì´ë¯¸ ë³€í™˜ëœ í¬ë§·)
    json_export = result.get("json_export", {})
    if json_export:
        # ì ê²€êµ¬ê°€ ë”•ì…”ë„ˆë¦¬ í˜•íƒœì¸ ê²½ìš° ê°œìˆ˜ë§Œ ì¶”ì¶œ
        jgm = json_export.get("ì ê²€êµ¬", 1)
        hole_count = jgm.get("ê°œìˆ˜", 1) if isinstance(jgm, dict) else jgm

        return {
            "ì¬ì§ˆ": json_export.get("ì¬ì§ˆ", material),
            "ì´ê°œìˆ˜": json_export.get("ì´ê°œìˆ˜", 0),
            "ë°”ë””íŒë„¬": json_export.get("ë°”ë””íŒë„¬", {}),
            "ì‚¬ì´ë“œíŒë„¬": json_export.get("ì‚¬ì´ë“œíŒë„¬", {}),
            "ì²œê³µêµ¬": hole_count,
            "ì†Œê³„": subtotal or json_export.get("ì†Œê³„", 0),
        }

    # Fallback: summary ë°ì´í„°ì—ì„œ ì¶”ì¶œ
    summary = result.get("summary", {})
    elements = result.get("elements", [])

    # ë°”ë””/ì‚¬ì´ë“œ ê°œìˆ˜ ì¹´ìš´íŠ¸
    body_cnt = sum(1 for e in elements if e.get("kind") == "BODY")
    side_cnt = sum(1 for e in elements if e.get("kind") == "SIDE")

    # ëŒ€í‘œ ëª¨ë¸ëª… ì¶”ì¶œ
    body_models = [e.get("model", "") for e in elements if e.get("kind") == "BODY"]
    side_models = [e.get("model", "") for e in elements if e.get("kind") == "SIDE"]

    body_info = {}
    if body_models:
        # ê°€ì¥ ë§ì´ ë‚˜ì˜¨ ëª¨ë¸
        from collections import Counter

        body_top = Counter(body_models).most_common(1)
        if body_top:
            body_info = {"ì¢…ë¥˜": body_top[0][0].replace("(rot)", ""), "ê°œìˆ˜": body_cnt}

    side_info = {}
    if side_models:
        from collections import Counter

        side_top = Counter(side_models).most_common(1)
        if side_top:
            side_info = {"ì¢…ë¥˜": side_top[0][0].replace("(rot)", ""), "ê°œìˆ˜": side_cnt}

    total_cnt = summary.get("ì´íŒë„¬ìˆ˜", body_cnt + side_cnt)

    return {
        "ì¬ì§ˆ": material,
        "ì´ê°œìˆ˜": int(total_cnt),
        "ë°”ë””íŒë„¬": body_info,
        "ì‚¬ì´ë“œíŒë„¬": side_info,
        "ì²œê³µêµ¬": 1,  # ê¸°ë³¸ê°’, json_export ì—†ìœ¼ë©´ 1ë¡œ ê°€ì •
        "ì†Œê³„": int(subtotal),
    }


# ----------------------------
# UI
# ----------------------------
st.title("ğŸ› ìš•ì‹¤ ê²¬ì ì„œ ìƒì„±ê¸°")

# Check for calculation results
floor_result = st.session_state.get(FLOOR_RESULT_KEY)
wall_result = st.session_state.get(WALL_RESULT_KEY)
ceil_result = st.session_state.get(CEIL_RESULT_KEY)

# ë²½íŒ ì›ê°€ ê³„ì‚° ì™„ë£Œ ì—¬ë¶€ í™•ì¸ (wall_panel_cost_final.py)
wall_cost_done = st.session_state.get("wall_cost_done", False)
wall_cost_result = st.session_state.get("shared_wall_cost_result", {})

has_floor = bool(floor_result)
# ë²½íŒ: wall_resultê°€ ìˆê±°ë‚˜, ë²½íŒ ì›ê°€ ê³„ì‚°ì´ ì™„ë£Œë˜ì—ˆìœ¼ë©´ OK
has_wall = bool(wall_result) or (wall_cost_done and bool(wall_cost_result))
has_ceil = bool(ceil_result)

# Status display
st.markdown("### ê³„ì‚° ê²°ê³¼ ìƒíƒœ")
col1, col2, col3, col4 = st.columns(4)
with col1:
    status = "âœ… ì™„ë£Œ" if has_floor else "âŒ ë¯¸ì™„ë£Œ"
    st.metric("ë°”ë‹¥íŒ", status)
with col2:
    status = "âœ… ì™„ë£Œ" if has_wall else "âŒ ë¯¸ì™„ë£Œ"
    st.metric("ë²½íŒ", status)
with col3:
    status = "âœ… ì™„ë£Œ" if has_ceil else "âŒ ë¯¸ì™„ë£Œ"
    st.metric("ì²œì¥íŒ", status)
with col4:
    # ë°”ë‹¥íŒ ì„¸ëŒ€ìˆ˜ í‘œì‹œ
    units_display = 1
    if floor_result:
        inputs = floor_result.get("inputs", {})
        units_display = int(inputs.get("units", 1))
    st.metric("ê³µì‚¬ ì„¸ëŒ€ìˆ˜", f"{units_display}ì„¸ëŒ€")

# ========== ë°”ë‹¥íŒ, ë²½íŒ, ì²œì¥íŒ ê³„ì‚° ì˜ì¡´ì„± ì²´í¬ ==========
missing_steps = []
if not has_floor:
    missing_steps.append("ğŸŸ¦ ë°”ë‹¥íŒ ê³„ì‚°")
if not has_wall:
    missing_steps.append("ğŸŸ© ë²½íŒ ê³„ì‚°")
if not has_ceil:
    missing_steps.append("ğŸŸ¨ ì²œì¥íŒ ê³„ì‚°")

if missing_steps:
    st.warning(
        f"âš ï¸ ê²¬ì ì„œë¥¼ ìƒì„±í•˜ë ¤ë©´ ë¨¼ì € **{', '.join(missing_steps)}**ì„(ë¥¼) ì™„ë£Œí•´ì•¼ í•©ë‹ˆë‹¤."
    )

    # ì•ˆë‚´ ì¹´ë“œ
    st.markdown(
        """
    <div style="
        border: 1px solid #f59e0b;
        border-radius: 12px;
        padding: 20px;
        margin: 16px 0;
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    ">
        <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 12px;">
            <span style="font-size: 24px;">ğŸ“‹</span>
            <h3 style="margin: 0; color: #0f172a; font-weight: 700;">ê³„ì‚° ìˆœì„œ ì•ˆë‚´</h3>
        </div>
        <p style="margin: 0 0 12px 36px; color: #78350f; line-height: 1.6;">
            ê²¬ì ì„œ ìƒì„±ì€ ëª¨ë“  ê³„ì‚°ì´ ì™„ë£Œëœ í›„ ì§„í–‰í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤:
        </p>
        <div style="margin-left: 36px; padding: 12px; background: white; border-radius: 8px; border: 1px solid #f59e0b;">
            <p style="margin: 0; color: #92400e; font-size: 0.95rem; line-height: 1.6;">
                <strong>1ë‹¨ê³„:</strong> ğŸŸ¦ ë°”ë‹¥íŒ ê³„ì‚°"""
        + (" â† <em style='color:#dc2626;'>ë¯¸ì™„ë£Œ</em>" if not has_floor else " âœ…")
        + """<br>
                <strong>2ë‹¨ê³„:</strong> ğŸŸ© ë²½íŒ ê³„ì‚°"""
        + (" â† <em style='color:#dc2626;'>ë¯¸ì™„ë£Œ</em>" if not has_wall else " âœ…")
        + """<br>
                <strong>3ë‹¨ê³„:</strong> ğŸŸ¨ ì²œì¥íŒ ê³„ì‚°"""
        + (" â† <em style='color:#dc2626;'>ë¯¸ì™„ë£Œ</em>" if not has_ceil else " âœ…")
        + """<br>
                <strong>4ë‹¨ê³„:</strong> ğŸ“‹ ê²¬ì ì„œ ìƒì„± â† <em>í˜„ì¬ í˜ì´ì§€</em>
            </p>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # ë¯¸ì™„ë£Œ ë‹¨ê³„ë¡œ ì´ë™í•˜ëŠ” ë²„íŠ¼
    col_spacer, col_btn, col_spacer2 = st.columns([1, 2, 1])
    with col_btn:
        if not has_floor:
            st.page_link(
                "pages/1_ë°”ë‹¥íŒ_ê³„ì‚°.py", label="ğŸŸ¦ ë°”ë‹¥íŒ ê³„ì‚° ì‹œì‘í•˜ê¸°", icon=None
            )
        elif not has_wall:
            st.page_link(
                "pages/2_ë²½íŒ_ê·œê²©.py", label="ğŸŸ© ë²½íŒ ê³„ì‚° ì‹œì‘í•˜ê¸°", icon=None
            )
        elif not has_ceil:
            st.page_link(
                "pages/5_ì²œì¥íŒ_ê³„ì‚°.py", label="ğŸŸ¨ ì²œì¥íŒ ê³„ì‚° ì‹œì‘í•˜ê¸°", icon=None
            )

    st.stop()  # ì´ì „ ë‹¨ê³„ ë¯¸ì™„ë£Œ ì‹œ ì´í›„ UI ì°¨ë‹¨

# ëª¨ë“  ë‹¨ê³„ ì™„ë£Œ ì‹œ ì„±ê³µ ë©”ì‹œì§€
st.success("âœ… ëª¨ë“  ê³„ì‚°ì´ ì™„ë£Œë˜ì—ˆìŠµë‹ˆë‹¤. ê²¬ì ì„œë¥¼ ìƒì„±í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.")

# Convert session_state data
floor_data = convert_floor_data(floor_result)
wall_data = convert_wall_data(wall_result)
ceiling_data = convert_ceiling_data(ceil_result)

# Sidebar: Pricebook upload
with st.sidebar:
    st.markdown("### â‘  ë‹¨ê°€í‘œ")

    # ë°”ë‹¥íŒì—ì„œ ê³µìœ ëœ ì—‘ì…€ íŒŒì¼ í™•ì¸
    shared_excel = st.session_state.get(SHARED_EXCEL_KEY)
    shared_excel_name = st.session_state.get(SHARED_EXCEL_NAME_KEY, "")

    # íŒŒì¼ ì†ŒìŠ¤ ì„ íƒ
    use_shared = st.checkbox(
        (
            f"ë°”ë‹¥íŒ ê³µìœ  íŒŒì¼ ì‚¬ìš© ({shared_excel_name})"
            if shared_excel
            else "ë°”ë‹¥íŒ ê³µìœ  íŒŒì¼ ì‚¬ìš© (ì—†ìŒ)"
        ),
        value=shared_excel is not None,
        disabled=shared_excel is None,
    )

    pricebook_file = None
    if use_shared and shared_excel:
        pricebook_file = shared_excel
        st.success(f"ê³µìœ  íŒŒì¼: {shared_excel_name}")
    else:
        pricebook_file = st.file_uploader(
            "ë³„ë„ ì—…ë¡œë“œ (ì‹œíŠ¸: ìì¬ë‹¨ê°€ë‚´ì—­)", type=["xlsx"]
        )

    st.markdown("---")
    st.markdown("### â‘¡ ê³„ì‚° ê²°ê³¼ (ìë™ ì—°ë™)")
    st.success(f"âœ… ë°”ë‹¥íŒ: {floor_data.get('ì¬ì§ˆ', 'N/A')}")
    st.success(f"âœ… ë²½íŒ: {wall_data.get('ì´ê°œìˆ˜', 0)}ì¥")
    st.success(f"âœ… ì²œì¥íŒ: {ceiling_data.get('ì´ê°œìˆ˜', 0)}ì¥")

    st.markdown("---")
    st.markdown("### â‘¢ ì˜µì…˜ ì„ íƒ")

# ìƒì‚°ê´€ë¦¬ë¹„ ì„¤ì •ì€ ê²¬ì ì„œ ë°ì´í„°ê°€ ìƒì„±ëœ í›„ì— í‘œì‹œ (ì•„ë˜ë¡œ ì´ë™)
# ë¨¼ì € rows ë°ì´í„°ë¥¼ ìƒì„±í•œ í›„ UIë¥¼ í‘œì‹œ

# Load pricebook
price_df: Optional[pd.DataFrame] = None
ceiling_drilling_prices: Dict[str, float] = {}
NEW_AUTO_ITEMS_STRUCTURE: Dict[str, Any] = {}
if pricebook_file is not None:
    try:
        # íŒŒì¼ í¬ì¸í„°ë¥¼ ì²˜ìŒìœ¼ë¡œ ë¦¬ì…‹ í›„ ì½ê¸°
        pricebook_file.seek(0)
        file_bytes = pricebook_file.read()
        pricebook_file.seek(0)  # ë‹¤ë¥¸ ê³³ì—ì„œ ì¬ì‚¬ìš©í•  ìˆ˜ ìˆë„ë¡ ë‹¤ì‹œ ë¦¬ì…‹
        price_df = load_pricebook_from_excel(file_bytes)
        ceiling_drilling_prices = load_ceiling_drilling_prices(file_bytes)
        NEW_AUTO_ITEMS_STRUCTURE = load_auto_items_structure(file_bytes)
        st.sidebar.success(f"ë‹¨ê°€í‘œ ë¡œë“œ ì™„ë£Œ: {len(price_df)}í–‰ (ì‹œíŠ¸: ìì¬ë‹¨ê°€ë‚´ì—­)")
        st.sidebar.success(f"ìë™ì§€ì •í•­ëª© ë¡œë“œ ì™„ë£Œ: {len(NEW_AUTO_ITEMS_STRUCTURE)}ê°œ ëŒ€ë¶„ë¥˜")

        # ë””ë²„ê¹…: ë¡œë“œëœ êµ¬ì¡° í™•ì¸
        if NEW_AUTO_ITEMS_STRUCTURE:
            with st.sidebar.expander("ğŸ“‹ ë¡œë“œëœ ìë™ì§€ì •í•­ëª© êµ¬ì¡°", expanded=False):
                for cat, info in NEW_AUTO_ITEMS_STRUCTURE.items():
                    if info.get("subcategories") is None:
                        st.write(f"- **{cat}**: ì¤‘ë¶„ë¥˜ ì—†ìŒ")
                    else:
                        st.write(f"- **{cat}**:")
                        for sub, specs in info.get("subcategories", {}).items():
                            st.write(f"  - {sub}: {len(specs)}ê°œ ê·œê²© ({', '.join(specs[:3])}{'...' if len(specs) > 3 else ''})")
    except Exception as e:
        st.sidebar.error(f"ë‹¨ê°€í‘œ ë¡œë“œ ì‹¤íŒ¨: {e}")

# ----------------------------
# ê³µí†µ í’ˆëª© ìë™ì§€ì • ì •ì˜ (í†µí•©)
# ----------------------------

# Session State í‚¤
AUTO_ITEMS_KEY = "auto_assigned_items"
AUTO_FLOOR_TYPE_KEY = "auto_floor_type"
AUTO_SHAPE_TYPE_KEY = "auto_shape_type"
CUSTOM_ITEMS_KEY = "custom_items"  # ì‚¬ìš©ì ì •ì˜ í’ˆëª©

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ã€Aã€‘ ìë™ì§€ì • í’ˆëª© (ê¸°ë³¸ í¬í•¨, ìˆ˜ëŸ‰ í¸ì§‘ ê°€ëŠ¥)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

# ã€A-1ã€‘ ìë™ì§€ì • í’ˆëª© êµ¬ì¡°ëŠ” Excelì˜ 'ìë™ì§€ì •í•­ëª©' ì‹œíŠ¸ì—ì„œ ë™ì ìœ¼ë¡œ ë¡œë“œë©ë‹ˆë‹¤
# NEW_AUTO_ITEMS_STRUCTUREëŠ” íŒŒì¼ ì—…ë¡œë“œ ì‹œ load_auto_items_structure() í•¨ìˆ˜ë¡œ ìƒì„±ë©ë‹ˆë‹¤

# ì²œì¥íŒ íƒ€ê³µ í’ˆëª© (ë³„ë„ ìœ ì§€)
CEILING_DRILLING_ITEMS = {
    "í™˜í’ê¸°í™€": 1,
    "ì‚¬ê°ë§¤ë¦½ë“±": 0,
    "ì›í˜•ë“± íƒ€ê³µ": 0,
    "ì§ì„  1íšŒ": 0,
}

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ã€Bã€‘ ì„ íƒ ìœ ì§€ í’ˆëª© (ì¢…ë¥˜ ì„ íƒ í•„ìš”)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ìë™ì§€ì • í’ˆëª© ê³„ì‚° í•¨ìˆ˜
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•


def calculate_auto_items(floor_type: str, shape_type: str) -> Dict[str, Any]:
    """
    ë°”ë‹¥íŒ ì¢…ë¥˜ì™€ í˜•íƒœì— ë”°ë¼ ìë™ì§€ì • í’ˆëª© ì´ˆê¸°ê°’ ê³„ì‚°
    ë°˜í™˜ í˜•ì‹: {ëŒ€ë¶„ë¥˜: {ì¤‘ë¶„ë¥˜: {ê·œê²©: ìˆ˜ëŸ‰}}}
    """
    result = {}

    # íƒ€ì¼ ì†Œìš”ê³„ì‚° ê²°ê³¼ ê°€ì ¸ì˜¤ê¸°
    total_wall_tiles = st.session_state.get("shared_total_wall_tiles", 0)
    total_floor_tiles = st.session_state.get("shared_total_floor_tiles", 0)
    wall_data = st.session_state.get("wall", {})
    tile_str = str(wall_data.get("ë²½íƒ€ì¼", "")).replace("Ã—", "x").replace(" ", "")

    # ê° ëŒ€ë¶„ë¥˜ë³„ë¡œ ì´ˆê¸° ìˆ˜ëŸ‰ ì„¤ì •
    for major_cat, config in NEW_AUTO_ITEMS_STRUCTURE.items():
        result[major_cat] = {}

        if config["subcategories"] is None:
            # ì¤‘ë¶„ë¥˜ ì—†ìŒ (ì˜ˆ: GRPë°”ë‹¥íŒ, FRPë°”ë‹¥íŒ)
            # ë°”ë‹¥íŒ ì¢…ë¥˜ì— ë”°ë¼ í™œì„±í™”
            if major_cat == "GRPë°”ë‹¥íŒ" and floor_type == "GRP":
                result[major_cat]["_self"] = 1
            elif major_cat == "FRPë°”ë‹¥íŒ" and floor_type in ["FRP", "SMC"]:
                result[major_cat]["_self"] = 1
            else:
                result[major_cat]["_self"] = 0
        else:
            # ì¤‘ë¶„ë¥˜ ìˆìŒ
            for subcat, specs in config["subcategories"].items():
                if not specs:
                    # ê·œê²© ì—†ìŒ - ê¸°ë³¸ê°’ ì„¤ì •
                    default_qty = 0

                    # íŠ¹ì • í•­ëª©ì€ ê¸°ë³¸ 1ë¡œ ì„¤ì •
                    if subcat in ["PBì´ì¤‘ê´€(ì˜¤í”ˆìˆ˜ì „í•¨)", "ë„ì–´ë½", "ê²½ì²©(ìŠ¤í…í”¼ìŠ¤)", "ìŠ¤í† í¼",
                                  "ì–‘ë³€ê¸°", "ì„¸ë©´ê¸°", "ì„¸ë©´ê¸° ìˆ˜ì „", "ê²¸ìš© ìˆ˜ì „", "ìƒ¤ì›Œ ìˆ˜ì „"]:
                        default_qty = 1

                    # íƒ€ì¼ë¥˜ ìˆ˜ëŸ‰ ë°˜ì˜ (íƒ€ì¼ì€ ê·œê²© ì—†ìŒ)
                    if major_cat == "íƒ€ì¼ë¥˜":
                        if tile_str in ["250x400", "250*400"]:
                            if subcat == "ë²½ì²´ìš© íƒ€ì¼ 250*400":
                                default_qty = total_wall_tiles
                            elif subcat == "ë°”ë‹¥ìš© íƒ€ì¼ 200*200":
                                default_qty = total_floor_tiles
                        else:
                            if subcat == "ë²½ì²´ìš© íƒ€ì¼ 300*600":
                                default_qty = total_wall_tiles
                            elif subcat == "ë°”ë‹¥ìš© íƒ€ì¼ 300*300":
                                default_qty = total_floor_tiles

                    result[major_cat][subcat] = {"_self": default_qty}
                else:
                    # ê·œê²© ìˆìŒ - ê° ê·œê²©ë³„ ì´ˆê¸°ê°’ 0
                    result[major_cat][subcat] = {spec: 0 for spec in specs}

    # í˜•íƒœë³„ ì¡°ì • (ì‚¬ê°í˜•/ì½”ë„ˆí˜•)
    # ê³µí†µ ë° ë¶€ì†ìì¬ > ì½”ë„ˆë§ˆê°ì¬, ì½”ë„ˆë¹„ë“œ
    if "ê³µí†µ ë° ë¶€ì†ìì¬" in result and "ì½”ë„ˆë§ˆê°ì¬(ë²½ì²´ ë’¤ìª½)" in result["ê³µí†µ ë° ë¶€ì†ìì¬"]:
        if shape_type == "ì‚¬ê°í˜•":
            result["ê³µí†µ ë° ë¶€ì†ìì¬"]["ì½”ë„ˆë§ˆê°ì¬(ë²½ì²´ ë’¤ìª½)"]["_self"] = 3
        elif shape_type == "ì½”ë„ˆí˜•":
            result["ê³µí†µ ë° ë¶€ì†ìì¬"]["ì½”ë„ˆë§ˆê°ì¬(ë²½ì²´ ë’¤ìª½)"]["_self"] = 5

    if "ê³µí†µ ë° ë¶€ì†ìì¬" in result and "ì½”ë„ˆë¹„ë“œ(ë²½ì²´ ì•ˆìª½)" in result["ê³µí†µ ë° ë¶€ì†ìì¬"]:
        if shape_type == "ì‚¬ê°í˜•":
            result["ê³µí†µ ë° ë¶€ì†ìì¬"]["ì½”ë„ˆë¹„ë“œ(ë²½ì²´ ì•ˆìª½)"]["17*17*2180"] = 0
        elif shape_type == "ì½”ë„ˆí˜•":
            result["ê³µí†µ ë° ë¶€ì†ìì¬"]["ì½”ë„ˆë¹„ë“œ(ë²½ì²´ ì•ˆìª½)"]["17*17*2180"] = 1

    # ì²œì¥íŒ íƒ€ê³µ í•­ëª© ì¶”ê°€
    result["_drilling"] = CEILING_DRILLING_ITEMS.copy()

    return result


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# AI í’ˆëª© íƒì§€ ê²°ê³¼ (AI ì‹œë°©ì„œ ë¶„ì„ í˜ì´ì§€ì—ì„œ íƒì§€ëœ ê²°ê³¼ í‘œì‹œ)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
AI_COMPARISON_RESULT_KEY = "ai_comparison_result"
AI_PENDING_ITEMS_KEY = "ai_pending_items"

st.markdown("---")
st.subheader("AI í’ˆëª© ìë™ íƒì§€")

comparison = st.session_state.get(AI_COMPARISON_RESULT_KEY)
pending_items = st.session_state.get(AI_PENDING_ITEMS_KEY, [])

if comparison or pending_items:
    if comparison:
        st.markdown(f"**{comparison.get('summary', '')}**")

    # ì¶”ê°€ ëŒ€ê¸° í’ˆëª© í‘œì‹œ
    if pending_items:
        st.markdown("#### ğŸ“‹ ì¶”ê°€ ëŒ€ê¸° í’ˆëª©")
        st.caption("ê° í•­ëª©ì˜ ëŒ€ë¶„ë¥˜/ì¤‘ë¶„ë¥˜/ì‚¬ì–‘ì„ ì§€ì •í•˜ì—¬ ê²¬ì ì„œì— ì¶”ê°€í•˜ì„¸ìš”.")

        items_to_remove = []
        for idx, item in enumerate(pending_items):
            with st.expander(f"ğŸ“¦ {item.get('name', '')} (ì¶œì²˜: {item.get('source', '')[:30]}...)", expanded=False):
                col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 1, 1])

                with col1:
                    item_major = st.text_input(
                        "ëŒ€ë¶„ë¥˜",
                        value=item.get("major", ""),
                        key=f"ai_major_{idx}",
                        placeholder="ì˜ˆ: ì˜¤ë°°ìˆ˜ë°°ê´€"
                    )

                with col2:
                    item_sub = st.text_input(
                        "ì¤‘ë¶„ë¥˜",
                        value=item.get("sub", ""),
                        key=f"ai_sub_{idx}",
                        placeholder="ì˜ˆ: PVCë³¸ë“œ"
                    )

                with col3:
                    item_spec = st.text_input(
                        "ì‚¬ì–‘ ë° ê·œê²©",
                        value=item.get("spec", ""),
                        key=f"ai_spec_{idx}",
                        placeholder="ì˜ˆ: 1kg"
                    )

                with col4:
                    item_qty = st.number_input(
                        "ìˆ˜ëŸ‰",
                        min_value=0.0,
                        value=float(item.get("qty") or 1),
                        step=0.5,
                        key=f"ai_qty_{idx}"
                    )

                with col5:
                    st.write("")  # ê³µë°±
                    st.write("")  # ë ˆì´ë¸” ë†’ì´ ë§ì¶”ê¸°
                    if st.button("â• ì¶”ê°€", key=f"ai_add_{idx}", use_container_width=True, type="primary"):
                        if item_major.strip():
                            custom_items = st.session_state.get(CUSTOM_ITEMS_KEY, [])
                            custom_items.append({
                                "major": item_major.strip(),
                                "sub": item_sub.strip(),
                                "spec": item_spec.strip(),
                                "qty": item_qty,
                                "source": "AI_DETECTED",
                            })
                            st.session_state[CUSTOM_ITEMS_KEY] = custom_items
                            items_to_remove.append(idx)
                            st.success(f"âœ… '{item_major}' ì¶”ê°€ë¨!")
                        else:
                            st.warning("âš ï¸ ëŒ€ë¶„ë¥˜ëŠ” í•„ìˆ˜ì…ë‹ˆë‹¤.")

                col_del, _ = st.columns([1, 4])
                with col_del:
                    if st.button("ğŸ—‘ ì‚­ì œ", key=f"ai_del_{idx}", use_container_width=True):
                        items_to_remove.append(idx)

        # ì‚­ì œí•  í•­ëª© ì²˜ë¦¬
        if items_to_remove:
            for idx in sorted(items_to_remove, reverse=True):
                pending_items.pop(idx)
            st.session_state[AI_PENDING_ITEMS_KEY] = pending_items
            st.rerun()

        # ëŒ€ê¸° ëª©ë¡ ë¹„ìš°ê¸° ë²„íŠ¼
        if st.button("ğŸ—‘ ëŒ€ê¸° ëª©ë¡ ëª¨ë‘ ë¹„ìš°ê¸°", use_container_width=False):
            st.session_state[AI_PENDING_ITEMS_KEY] = []
            st.rerun()

    # ì¶”ê°€ ê²€í†  í•„ìš” í’ˆëª© (ì•„ì§ ëŒ€ê¸° ëª©ë¡ì— ì—†ëŠ” ê²ƒë“¤)
    if comparison:
        to_add = comparison.get("to_add", [])
        pending_names = {p.get("name", "").lower() for p in pending_items}
        remaining = [
            item for item in to_add if item.get("name", "").lower() not in pending_names
        ]

        if remaining:
            with st.expander(
                f"ğŸ“ ì¶”ê°€ ê²€í†  í•„ìš” í’ˆëª© ({len(remaining)}ê°œ)", expanded=False
            ):
                for idx, item in enumerate(remaining):
                    col1, col2, col3 = st.columns([3.5, 1.5, 1])
                    with col1:
                        priority_icon = "ğŸ”´" if item.get("priority") == "high" else "ğŸŸ¡"
                        st.write(
                            f"{priority_icon} {item.get('name', '')} - {item.get('source', '')[:30] if item.get('source') else ''}"
                        )
                    with col2:
                        # ìˆ˜ëŸ‰ ì…ë ¥ í•„ë“œ (ê¸°ë³¸ê°’ 1)
                        review_qty = st.number_input(
                            "ìˆ˜ëŸ‰",
                            min_value=1,
                            value=item.get("qty") or 1,
                            key=f"qty_review_{idx}_{item.get('name', '')}",
                            label_visibility="collapsed",
                        )
                    with col3:
                        if st.button(
                            "ì¶”ê°€", key=f"est_review_add_{idx}_{item.get('name', '')}"
                        ):
                            item_to_add = item.copy()
                            item_to_add["qty"] = review_qty
                            pending_items.append(item_to_add)
                            st.session_state[AI_PENDING_ITEMS_KEY] = pending_items
                            st.rerun()
else:
    st.info(
        "ğŸ“‹ AI ì‹œë°©ì„œ ë¶„ì„ í˜ì´ì§€ì—ì„œ ì‹œë°©ì„œ PDFë¥¼ ì—…ë¡œë“œí•˜ê³  ì¸ë±ìŠ¤ë¥¼ ìƒì„±í•˜ë©´ í’ˆëª©ì´ ìë™ íƒì§€ë©ë‹ˆë‹¤."
    )

# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# UI: ë°”ë‹¥íŒ ì¢…ë¥˜ ë° í˜•íƒœ ì„ íƒ
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
st.markdown("---")
st.subheader("í’ˆëª© ì„¤ì •")

col_floor_type, col_shape_type = st.columns(2)

with col_floor_type:
    # ë°”ë‹¥íŒ ì¬ì§ˆì—ì„œ ì¢…ë¥˜ ì¶”ì¶œ (ë°”ë‹¥íŒ í˜ì´ì§€ì—ì„œ ê²°ì •ëœ ê°’ ì‚¬ìš©)
    floor_material = floor_data.get("ì¬ì§ˆ", "").upper() if floor_data else ""

    # PP/PE, PVEëŠ” PPë¡œ, GRP/SMC/FRPëŠ” GRPë¡œ ë§¤í•‘
    if "PP" in floor_material or "PE" in floor_material or "PVE" in floor_material:
        floor_type = "PP"
    else:
        floor_type = "GRP"

    # ë°”ë‹¥íŒì—ì„œ ë„˜ì–´ì˜¨ ê°’ í‘œì‹œ (ìˆ˜ì • ë¶ˆê°€)
    st.radio(
        "ë°”ë‹¥íŒ ì¢…ë¥˜",
        options=["PP", "GRP"],
        index=0 if floor_type == "PP" else 1,
        horizontal=True,
        help="ë°”ë‹¥íŒ í˜ì´ì§€ì—ì„œ ì§€ì •ëœ ê°’ (ìˆ˜ì • ë¶ˆê°€)",
        key="floor_type_radio",
        disabled=True,  # ìˆ˜ì • ë¶ˆê°€
    )

with col_shape_type:
    # ë°”ë‹¥íŒ í˜ì´ì§€ì—ì„œ ì €ì¥ëœ í˜•íƒœ ê°’ ì‚¬ìš© (shared_bath_shape)
    shape_type = st.session_state.get("shared_bath_shape", "ì‚¬ê°í˜•")

    # ë°”ë‹¥íŒì—ì„œ ë„˜ì–´ì˜¨ ê°’ í‘œì‹œ (ìˆ˜ì • ë¶ˆê°€)
    st.radio(
        "ìš•ì‹¤ í˜•íƒœ",
        options=["ì‚¬ê°í˜•", "ì½”ë„ˆí˜•"],
        index=0 if shape_type == "ì‚¬ê°í˜•" else 1,
        horizontal=True,
        help="ë°”ë‹¥íŒ í˜ì´ì§€ì—ì„œ ì§€ì •ëœ ê°’ (ìˆ˜ì • ë¶ˆê°€)",
        key="shape_type_radio",
        disabled=True,  # ìˆ˜ì • ë¶ˆê°€
    )

# ë³€ê²½ ê°ì§€ (ë°”ë‹¥íŒ í˜ì´ì§€ì—ì„œ ê°’ì´ ë³€ê²½ëœ ê²½ìš°)
floor_type_changed = st.session_state.get(AUTO_FLOOR_TYPE_KEY) != floor_type
shape_type_changed = st.session_state.get(AUTO_SHAPE_TYPE_KEY) != shape_type

st.session_state[AUTO_FLOOR_TYPE_KEY] = floor_type
st.session_state[AUTO_SHAPE_TYPE_KEY] = shape_type

current_auto_items = calculate_auto_items(floor_type, shape_type)

if floor_type_changed or shape_type_changed:
    st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()
    st.info(
        f"ë°”ë‹¥íŒ ì¢…ë¥˜({floor_type}) ë˜ëŠ” í˜•íƒœ({shape_type}) ë³€ê²½ìœ¼ë¡œ í’ˆëª© ìˆ˜ëŸ‰ì´ ì´ˆê¸°í™”ë˜ì—ˆìŠµë‹ˆë‹¤."
    )

if AUTO_ITEMS_KEY not in st.session_state:
    st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# UI: ã€Aã€‘ ìë™ì§€ì • í’ˆëª© ìˆ˜ëŸ‰ í¸ì§‘
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
if not NEW_AUTO_ITEMS_STRUCTURE:
    st.warning("âš ï¸ ìë™ì§€ì • í’ˆëª©ì„ ë¡œë“œí•˜ë ¤ë©´ ë‹¨ê°€í‘œ Excel íŒŒì¼ì„ ì—…ë¡œë“œí•˜ì„¸ìš”. (ì‹œíŠ¸: 'ìë™ì§€ì •í•­ëª©' í•„ìš”)")
elif pricebook_file is None:
    st.warning("âš ï¸ ë‹¨ê°€í‘œ íŒŒì¼ì´ ì—…ë¡œë“œë˜ì§€ ì•Šì•˜ìŠµë‹ˆë‹¤.")
else:
    with st.expander("ìë™ì§€ì • í’ˆëª© ìˆ˜ëŸ‰ í¸ì§‘", expanded=False):
        st.markdown("**ê¸°ë³¸ í¬í•¨ë˜ëŠ” í’ˆëª©ì˜ ìˆ˜ëŸ‰ì„ í¸ì§‘í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.**")
        st.caption(f"í˜„ì¬ ì„¤ì •: ë°”ë‹¥íŒ={floor_type}, í˜•íƒœ={shape_type}")

        # CSS ìŠ¤íƒ€ì¼: ê·œê²© ìˆ˜ëŸ‰ ì…ë ¥ ì¹¸ë§Œ ìŠ¤íƒ€ì¼ ì ìš©
        st.markdown(
            """
            <style>
            /* ê·œê²© ìˆ˜ëŸ‰ ì…ë ¥ ì¹¸ì˜ ë„ˆë¹„ë¥¼ ì¤„ì„ (columns ë‚´ë¶€ì˜ number_inputë§Œ) */
            div[data-testid="stExpander"] div[data-testid="column"] div[data-testid="stNumberInput"] {
                max-width: 180px !important;
            }

            /* ê·œê²© ìˆ˜ëŸ‰ ì…ë ¥ ì¹¸ì˜ ë ˆì´ë¸”ì„ íšŒìƒ‰ìœ¼ë¡œ (columns ë‚´ë¶€ì˜ number_inputë§Œ) */
            div[data-testid="stExpander"] div[data-testid="column"] div[data-testid="stNumberInput"] label {
                color: #808080 !important;
                font-size: 0.9rem !important;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        if st.button("ê¸°ë³¸ê°’ìœ¼ë¡œ ì´ˆê¸°í™”", key="reset_new_auto_items"):
            st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()
            st.success("ê¸°ë³¸ê°’ìœ¼ë¡œ ì´ˆê¸°í™”ë˜ì—ˆìŠµë‹ˆë‹¤.")
            st.rerun()

        # ì„¸ì…˜ ìŠ¤í…Œì´íŠ¸ì—ì„œ í˜„ì¬ ì„ íƒê°’ ê°€ì ¸ì˜¤ê¸°
        edited_items = st.session_state.get(AUTO_ITEMS_KEY, current_auto_items).copy()

        # ê° ëŒ€ë¶„ë¥˜ë³„ë¡œ ì²˜ë¦¬
        for major_category, config in NEW_AUTO_ITEMS_STRUCTURE.items():
            st.markdown(f"### {major_category}")

            if config["subcategories"] is None:
                # ì¼€ì´ìŠ¤ 1: ì¤‘ë¶„ë¥˜/ê·œê²© ì—†ìŒ (ì˜ˆ: GRPë°”ë‹¥íŒ, FRPë°”ë‹¥íŒ)
                if major_category not in edited_items:
                    edited_items[major_category] = {}

                current_qty = edited_items.get(major_category, {}).get("_self", 0)
                default_qty = current_auto_items.get(major_category, {}).get("_self", 0)

                new_qty = st.number_input(
                    f"{major_category} ìˆ˜ëŸ‰",
                    min_value=0.0,
                    max_value=999.0,
                    value=float(current_qty),
                    step=0.5,
                    key=f"qty_{major_category}",
                    help=f"ê¸°ë³¸ê°’: {default_qty}",
                )
                edited_items[major_category]["_self"] = new_qty

            else:
                # ì¼€ì´ìŠ¤ 2: ì¤‘ë¶„ë¥˜ ìˆìŒ
                for subcategory, specs in config["subcategories"].items():
                    if not specs:
                        # ì¼€ì´ìŠ¤ 2-1: ê·œê²© ì—†ìŒ (ì˜ˆ: PBë…ë¦½ë°°ê´€, ë„ì–´ë½)
                        if major_category not in edited_items:
                            edited_items[major_category] = {}
                        if subcategory not in edited_items[major_category]:
                            edited_items[major_category][subcategory] = {"_self": 0}

                        current_qty = edited_items[major_category][subcategory].get("_self", 0)
                        default_qty = current_auto_items.get(major_category, {}).get(subcategory, {}).get("_self", 0)

                        # ì¤‘ë¶„ë¥˜ëª…ì„ êµµê²Œ í‘œì‹œ
                        st.markdown(f"**{subcategory}**", unsafe_allow_html=True)

                        new_qty = st.number_input(
                            f"ìˆ˜ëŸ‰",
                            min_value=0.0,
                            max_value=999.0,
                            value=float(current_qty),
                            step=0.5,
                            key=f"qty_{major_category}_{subcategory}",
                            help=f"ê¸°ë³¸ê°’: {default_qty}",
                            label_visibility="collapsed",  # ë ˆì´ë¸” ìˆ¨ê¹€
                        )
                        edited_items[major_category][subcategory]["_self"] = new_qty

                    else:
                        # ì¼€ì´ìŠ¤ 2-2: ê·œê²© ìˆìŒ â†’ multiselect + ê° ê·œê²©ë³„ ìˆ˜ëŸ‰
                        # í˜„ì¬ ì„ íƒëœ ê·œê²© ì°¾ê¸° (ìˆ˜ëŸ‰ > 0ì¸ í•­ëª©)
                        if major_category not in edited_items:
                            edited_items[major_category] = {}
                        if subcategory not in edited_items[major_category]:
                            edited_items[major_category][subcategory] = {}

                        # ë””ë²„ê¹…: specs í™•ì¸
                        if not specs:
                            st.warning(f"âš ï¸ '{subcategory}'ì— ê·œê²©ì´ ì—†ìŠµë‹ˆë‹¤. Excelì˜ 'ìë™ì§€ì •í•­ëª©' ì‹œíŠ¸ë¥¼ í™•ì¸í•˜ì„¸ìš”.")
                            continue

                        default_selected = [
                            s for s in specs
                            if edited_items.get(major_category, {}).get(subcategory, {}).get(s, 0) > 0
                        ]

                        # ê¸°ë³¸ê°’ì´ ì—†ìœ¼ë©´ ì²« ë²ˆì§¸ í•­ëª©ì„ ì„ íƒ
                        if not default_selected and specs:
                            default_selected = [specs[0]]

                        # ë ˆì´ë¸”ì„ ë³„ë„ë¡œ í‘œì‹œ: ì¤‘ë¶„ë¥˜ëª…(ê²€ì€ìƒ‰) + "ì„ íƒ"(íšŒìƒ‰)
                        st.markdown(
                            f"**{subcategory}** <span style='color: #808080; font-size: 0.875rem;'>ì„ íƒ</span>",
                            unsafe_allow_html=True
                        )

                        selected_specs = st.multiselect(
                            f"{subcategory} ì„ íƒ",
                            options=specs,
                            default=default_selected,
                            key=f"multi_{major_category}_{subcategory}",
                            label_visibility="collapsed",  # ë ˆì´ë¸” ìˆ¨ê¹€
                        )

                        if selected_specs:
                            # ì„ íƒëœ ê·œê²©ë“¤ì— ëŒ€í•´ ìˆ˜ëŸ‰ ì…ë ¥
                            cols = st.columns(min(len(selected_specs), 3))
                            for idx, spec in enumerate(selected_specs):
                                with cols[idx % len(cols)]:
                                    current_qty = edited_items[major_category][subcategory].get(spec, 1)
                                    default_qty = current_auto_items.get(major_category, {}).get(subcategory, {}).get(spec, 0)

                                    new_qty = st.number_input(
                                        f"{spec}",
                                        min_value=0.0,
                                        max_value=999.0,
                                        value=float(current_qty) if current_qty > 0 else 1.0,
                                        step=0.5,
                                        key=f"qty_{major_category}_{subcategory}_{spec}",
                                        help=f"ê¸°ë³¸ê°’: {default_qty}",
                                    )
                                    edited_items[major_category][subcategory][spec] = new_qty

                        # ì„ íƒ í•´ì œëœ ê·œê²©ì€ 0ìœ¼ë¡œ ì„¤ì •
                        for spec in specs:
                            if spec not in selected_specs:
                                edited_items[major_category][subcategory][spec] = 0

        # ì²œì¥íŒ íƒ€ê³µ í•­ëª© (ë³„ë„ ìœ ì§€)
        st.divider()
        st.markdown("### ê°€ê³µ í’ˆëª© (ì²œì¥íŒ íƒ€ê³µ)")

        if "_drilling" not in edited_items:
            edited_items["_drilling"] = {}

        cols = st.columns(4)
        for idx, (item_name, default_qty) in enumerate(CEILING_DRILLING_ITEMS.items()):
            with cols[idx % 4]:
                current_qty = edited_items["_drilling"].get(item_name, default_qty)
                new_qty = st.number_input(
                    item_name,
                    min_value=0.0,
                    max_value=999.0,
                    value=float(current_qty),
                    step=1.0,
                    key=f"drilling_{item_name}",
                    help=f"ê¸°ë³¸ê°’: {default_qty}",
                )
                edited_items["_drilling"][item_name] = new_qty

        # ì„¸ì…˜ ìŠ¤í…Œì´íŠ¸ ì—…ë°ì´íŠ¸
        st.session_state[AUTO_ITEMS_KEY] = edited_items

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # 'ê²¬ì ì— í¬í•¨' ë¬¸ì¥ì—ì„œ í’ˆëª© ì¶”ê°€
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    quote_sentences = st.session_state.get("ai_quote_sentences", [])
    if quote_sentences:
        st.divider()
        st.markdown("### ğŸ“ ì‹œë°©ì„œì—ì„œ ëˆ„ë½ í•¨ëª© íƒì§€ ")
        st.caption("í’ˆëª©ëª…ì„ ì…ë ¥í•˜ì—¬ ì¶”ê°€í•˜ì„¸ìš”.")

        for idx, sent in enumerate(quote_sentences):
            with st.expander(f"ğŸ“„ {sent.get('sentence', '')[:50]}...", expanded=False):
                st.info(f"**ë¬¸ì¥:** {sent.get('sentence', '')}")
                if sent.get("context"):
                    st.caption(f"ìƒí™©: {sent.get('context', '')}")

                # AIê°€ ì¶”ì¶œí•œ í’ˆëª© ì œì•ˆ
                suggested_items = sent.get("items", [])
                default_name = suggested_items[0] if suggested_items else ""

                col1, col2, col3, col4, col5 = st.columns([2, 2, 2, 1, 1])
                with col1:
                    item_major = st.text_input(
                        "ëŒ€ë¶„ë¥˜",
                        key=f"quote_major_{idx}",
                        placeholder="ì˜ˆ: ì˜¤ë°°ìˆ˜ë°°ê´€"
                    )
                with col2:
                    item_sub = st.text_input(
                        "ì¤‘ë¶„ë¥˜",
                        value=default_name,
                        key=f"quote_sub_{idx}",
                        placeholder="ì˜ˆ: PVCë³¸ë“œ"
                    )
                with col3:
                    item_spec = st.text_input(
                        "ì‚¬ì–‘ ë° ê·œê²©",
                        key=f"quote_spec_{idx}",
                        placeholder="ì˜ˆ: 1kg"
                    )
                with col4:
                    item_qty = st.number_input(
                        "ìˆ˜ëŸ‰",
                        min_value=0.0,
                        value=1.0,
                        step=0.5,
                        key=f"quote_qty_{idx}"
                    )
                with col5:
                    st.write("")  # ê³µë°±
                    st.write("")  # ë ˆì´ë¸” ë†’ì´ ë§ì¶”ê¸°
                    if st.button(
                        "â• ì¶”ê°€", key=f"quote_add_{idx}", use_container_width=True, type="primary"
                    ):
                        if item_major.strip():
                            custom_items = st.session_state.get(CUSTOM_ITEMS_KEY, [])
                            custom_items.append({
                                "major": item_major.strip(),
                                "sub": item_sub.strip(),
                                "spec": item_spec.strip(),
                                "qty": item_qty,
                                "source": sent.get("sentence", "")[:50],
                            })
                            st.session_state[CUSTOM_ITEMS_KEY] = custom_items
                            st.success(f"âœ… '{item_major}' ì¶”ê°€ë¨!")
                            st.rerun()
                        else:
                            st.warning("âš ï¸ ëŒ€ë¶„ë¥˜ëŠ” í•„ìˆ˜ì…ë‹ˆë‹¤.")

    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    # ì‚¬ìš©ì ì •ì˜ í’ˆëª© ì¶”ê°€
    # â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
    st.divider()
    st.markdown("### ì‚¬ìš©ì ì •ì˜ í’ˆëª© ì¶”ê°€")
    st.caption("ë‹¨ê°€í‘œì— ìˆëŠ” í’ˆëª©ì„ ì§ì ‘ ì¶”ê°€í•  ìˆ˜ ìˆìŠµë‹ˆë‹¤.")

    # ì‚¬ìš©ì ì •ì˜ í’ˆëª© ì´ˆê¸°í™”
    if CUSTOM_ITEMS_KEY not in st.session_state:
        st.session_state[CUSTOM_ITEMS_KEY] = []

    # ìƒˆ í’ˆëª© ì¶”ê°€ í¼
    with st.form("add_custom_item_form", clear_on_submit=True):
        col1, col2, col3, col4 = st.columns([2, 2, 2, 1])
        with col1:
            new_major = st.text_input("ëŒ€ë¶„ë¥˜", placeholder="ì˜ˆ: ì•¡ì„¸ì„œë¦¬")
        with col2:
            new_sub = st.text_input("ì¤‘ë¶„ë¥˜", placeholder="ì˜ˆ: ìˆ˜ê±´ê±¸ì´")
        with col3:
            new_spec = st.text_input("ì‚¬ì–‘ ë° ê·œê²©", placeholder="ì˜ˆ: EL-400-1")
        with col4:
            new_qty = st.number_input(
                "ìˆ˜ëŸ‰", min_value=0.0, max_value=100.0, value=1.0, step=0.5
            )

        add_btn = st.form_submit_button("í’ˆëª© ì¶”ê°€", use_container_width=True)
        if add_btn:
            if not new_major.strip():
                st.warning("ëŒ€ë¶„ë¥˜ë¥¼ ì…ë ¥í•˜ì„¸ìš”.")
            else:
                st.session_state[CUSTOM_ITEMS_KEY].append(
                    {
                        "major": new_major.strip(),
                        "sub": new_sub.strip(),
                        "spec": new_spec.strip(),
                        "qty": new_qty,
                    }
                )
                display_name = f"{new_major}"
                if new_sub.strip():
                    display_name += f" > {new_sub}"
                if new_spec.strip():
                    display_name += f" ({new_spec})"
                st.success(f"'{display_name}' í’ˆëª©ì´ ì¶”ê°€ë˜ì—ˆìŠµë‹ˆë‹¤.")
                st.rerun()

    # ì¶”ê°€ëœ ì‚¬ìš©ì ì •ì˜ í’ˆëª© ëª©ë¡ í‘œì‹œ ë° ì‚­ì œ
    custom_items = st.session_state.get(CUSTOM_ITEMS_KEY, [])
    if custom_items:
        st.markdown("**ì¶”ê°€ëœ ì‚¬ìš©ì ì •ì˜ í’ˆëª©:**")
        items_to_remove = []
        for idx, item in enumerate(custom_items):
            col_info, col_del = st.columns([4, 1])
            with col_info:
                # ëŒ€ë¶„ë¥˜, ì¤‘ë¶„ë¥˜, ê·œê²© í‘œì‹œ
                major = item.get('major', item.get('category', ''))
                sub = item.get('sub', '')
                spec = item.get('spec', '')
                qty = item.get('qty', item.get('name', ''))

                display_text = f"{major}"
                if sub:
                    display_text += f" > {sub}"
                if spec:
                    display_text += f" ({spec})"
                display_text += f" - ìˆ˜ëŸ‰: {qty}"

                st.text(display_text)
            with col_del:
                if st.button("ì‚­ì œ", key=f"del_custom_{idx}"):
                    items_to_remove.append(idx)

        # ì‚­ì œ ì²˜ë¦¬
        if items_to_remove:
            for idx in sorted(items_to_remove, reverse=True):
                st.session_state[CUSTOM_ITEMS_KEY].pop(idx)
            st.rerun()

# ìµœì¢… ìë™ì§€ì • í’ˆëª©
final_auto_items = st.session_state.get(AUTO_ITEMS_KEY, current_auto_items)
final_custom_items = st.session_state.get(CUSTOM_ITEMS_KEY, [])

# ----------------------------
# ê²¬ì ì„œ ìƒì„±
# ----------------------------
rows: List[Dict[str, Any]] = []
warnings: List[str] = []

if price_df is None:
    st.warning("ë‹¨ê°€í‘œ(ì—‘ì…€)ë¥¼ ë¨¼ì € ì—…ë¡œë“œí•˜ì„¸ìš”.")
else:
    # 1) ë°”ë‹¥íŒ
    if floor_data:
        material = str(floor_data.get("ì¬ì§ˆ", "")).upper()
        spec_text = str(floor_data.get("ê·œê²©", "")).strip()
        qty = float(floor_data.get("ìˆ˜ëŸ‰", 1))
        unit_price = float(floor_data.get("ë‹¨ê°€", 0))
        senior = bool(floor_data.get("ì£¼ê±°ì•½ì", False))

        # í’ˆëª© 'ë°”ë‹¥íŒ' ë³¸ì²´
        add_row(rows, "ë°”ë‹¥íŒ", material, qty, unit_price)

        # ë¶€ì¬ë£Œ ìë™ í¬í•¨
        if material in ["GRP", "SMC/FRP", "PP/PE", "PVE"]:
            if material == "PVE":
                ë¶„ë¥˜ = "PP/PE ë¶€ì¬ë£Œ"
            elif material == "SMC/FRP":
                ë¶„ë¥˜ = "SMC/FRP ë¶€ì¬ë£Œ"
            elif material == "PP/PE":
                ë¶„ë¥˜ = "PP/PE ë¶€ì¬ë£Œ"
            else:
                ë¶„ë¥˜ = "GRPë¶€ì¬ë£Œ"
            add_all_by_category(rows, price_df, "ë°”ë‹¥íŒ", ë¶„ë¥˜)
        else:
            warnings.append(
                f"ë°”ë‹¥íŒ ì¬ì§ˆ '{material}'ì— ëŒ€í•œ ë¶„ë¥˜ ë§¤í•‘ì„ ì°¾ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤."
            )

        # ì£¼ê±°ì•½ì ì¶”ê°€
        if senior:
            for spec in [
                "ë§¤ë¦½í˜• íœ´ì§€ê±¸ì´(ë¹„ìƒí°)",
                "Lí˜• ì†ì¡ì´",
                "ã…¡í˜• ì†ì¡ì´",
                "ì ‘ì˜ì‹ ì˜ì",
            ]:
                rec = find_item(price_df, "ì•¡ì„¸ì„œë¦¬", f"ì£¼ê±°ì•½ì {spec}")
                if rec is not None:
                    add_row(
                        rows,
                        "ì•¡ì„¸ì„œë¦¬",
                        spec,
                        rec.get("ìˆ˜ëŸ‰", 1) or 1,
                        rec.get("ë‹¨ê°€", 0),
                    )
                else:
                    add_row(rows, "ì•¡ì„¸ì„œë¦¬", spec, 1, 0)
                    warnings.append(f"ì£¼ê±°ì•½ì '{spec}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

    # 2) ë²½íŒ & íƒ€ì¼
    if wall_data:
        # PUë²½íŒ - 1ê°œë¡œ í‘œì‹œ, ë‹¨ê°€ëŠ” ì´ ê¸ˆì•¡ (ë²½íŒ ì›ê°€ ê³„ì‚° ê²°ê³¼ ì‚¬ìš©)
        wall_spec = "PUë²½íŒ"
        # production_costê°€ ìˆìœ¼ë©´ ì§ì ‘ ì‚¬ìš©, ì—†ìœ¼ë©´ ê³„ì‚°
        total_wall_price = float(wall_data.get("production_cost", 0))
        if total_wall_price == 0:
            total_qty = float(wall_data.get("ì´ê°œìˆ˜", 0))
            unit_price_per_panel = float(wall_data.get("ë‹¨ê°€", 0))
            total_wall_price = total_qty * unit_price_per_panel
        add_row(rows, "ë²½íŒ", wall_spec, 1, total_wall_price)

        # NOTE: ë²½íƒ€ì¼ & ë°”ë‹¥íƒ€ì¼ì€ ì´ì œ ìë™ì§€ì • í’ˆëª©ì—ì„œ ì²˜ë¦¬ë¨
        # íƒ€ì¼ ì†Œìš”ê³„ì‚° ê²°ê³¼ëŠ” calculate_auto_items()ì—ì„œ ìë™ìœ¼ë¡œ ë°˜ì˜ë¨

    # 3) ì²œì¥íŒ - ì²œì¥íŒ ê³„ì‚° ê²°ê³¼ì˜ ì†Œê³„ + íƒ€ê³µë¹„ í•©ì‚°
    if ceiling_data:
        total_cnt = float(ceiling_data.get("ì´ê°œìˆ˜", 0))
        subtotal = float(ceiling_data.get("ì†Œê³„", 0))

        # íƒ€ê³µ ê°€ê³µë¹„ìš© í•©ì‚° (ìë™ì§€ì • í’ˆëª©ì—ì„œ ìˆ˜ëŸ‰ Ã— ë‹¨ê°€)
        drilling_items = {
            "í™˜í’ê¸°í™€": "í™˜í’ê¸°í™€",
            "ì‚¬ê°ë§¤ë¦½ë“±": "ì‚¬ê°ë§¤ë¦½ë“±",
            "ì›í˜•ë“± íƒ€ê³µ": "ì›í˜•ë“±",
            "ì§ì„  1íšŒ": "ì§ì„ 1íšŒ",
        }
        drilling_total = 0
        for auto_name, price_key in drilling_items.items():
            drill_qty = final_auto_items.get(auto_name, 0)
            if drill_qty > 0:
                drill_unit_price = ceiling_drilling_prices.get(price_key, 0)
                drilling_total += drill_qty * drill_unit_price

        # ì²œì¥íŒ ì´ ê¸ˆì•¡ = ì†Œê³„ + íƒ€ê³µë¹„
        total_price = subtotal + drilling_total

        # ë‹¨ê°€ì™€ ê¸ˆì•¡ ëª¨ë‘ ì´ ê¸ˆì•¡ìœ¼ë¡œ í‘œì‹œ (ìˆ˜ëŸ‰ 1)
        add_row(rows, "ì²œì¥íŒ", "GRPì²œì¥íŒ", 1, total_price)

    # 4) ìë™ì§€ì • í’ˆëª© ì¶”ê°€
    # ì´ë¯¸ ì¶”ê°€ëœ í’ˆëª© ì¶”ì  (ì¤‘ë³µ ë°©ì§€)
    added_specs = set()
    for r in rows:
        spec_key = f"{r['í’ˆëª©']}::{r['ì‚¬ì–‘ ë° ê·œê²©']}"
        added_specs.add(spec_key)

    # ìë™ì§€ì • í’ˆëª© ì¶”ê°€ (ìƒˆë¡œìš´ ê³„ì¸µ êµ¬ì¡°)
    for major_cat, subcats in final_auto_items.items():
        if major_cat == "_drilling":
            # ì²œì¥íŒ íƒ€ê³µ í•­ëª© (ë³„ë„ ì²˜ë¦¬)
            continue

        config = NEW_AUTO_ITEMS_STRUCTURE.get(major_cat)
        if not config:
            continue

        if config["subcategories"] is None:
            # ì¼€ì´ìŠ¤ 1: ì¤‘ë¶„ë¥˜ ì—†ìŒ (ì˜ˆ: GRPë°”ë‹¥íŒ, FRPë°”ë‹¥íŒ)
            qty = subcats.get("_self", 0)
            if qty > 0:
                # ë‹¨ê°€í‘œì—ì„œ ì°¾ê¸°
                category_name = config["category_map"]
                spec_key = f"{category_name}::{major_cat}"

                if spec_key not in added_specs:
                    rec = find_item(price_df, category_name, "")
                    if rec is not None:
                        unit_price = rec.get("ë‹¨ê°€", 0) or 0
                    else:
                        unit_price = 0
                        warnings.append(f"[ìë™] '{major_cat}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

                    add_row(rows, category_name, "", qty, unit_price)
                    added_specs.add(spec_key)

        else:
            # ì¼€ì´ìŠ¤ 2: ì¤‘ë¶„ë¥˜ ìˆìŒ
            for subcat, spec_dict in subcats.items():
                if not spec_dict:
                    continue

                if "_self" in spec_dict:
                    # ì¼€ì´ìŠ¤ 2-1: ê·œê²© ì—†ìŒ
                    qty = spec_dict.get("_self", 0)
                    if qty > 0:
                        category_name = config["category_map"]
                        spec_key = f"{category_name}::{subcat}"

                        if spec_key not in added_specs:
                            rec = find_item(price_df, category_name, subcat)
                            if rec is not None:
                                unit_price = rec.get("ë‹¨ê°€", 0) or 0
                            else:
                                unit_price = 0
                                warnings.append(f"[ìë™] '{major_cat} > {subcat}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

                            add_row(rows, category_name, subcat, qty, unit_price)
                            added_specs.add(spec_key)

                else:
                    # ì¼€ì´ìŠ¤ 2-2: ê·œê²© ìˆìŒ
                    for spec, qty in spec_dict.items():
                        if qty > 0:
                            category_name = config["category_map"]
                            # ì¤‘ë¶„ë¥˜ + ê·œê²©ì„ í•©ì³ì„œ ì‚¬ì–‘ ë° ê·œê²©ìœ¼ë¡œ ì‚¬ìš©
                            spec_text = f"{subcat} {spec}".strip()
                            spec_key = f"{category_name}::{spec_text}"

                            if spec_key not in added_specs:
                                rec = find_item(price_df, category_name, spec_text)
                                if rec is not None:
                                    unit_price = rec.get("ë‹¨ê°€", 0) or 0
                                else:
                                    unit_price = 0
                                    warnings.append(f"[ìë™] '{major_cat} > {subcat} > {spec}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

                                add_row(rows, category_name, spec_text, qty, unit_price)
                                added_specs.add(spec_key)

    # ì²œì¥íŒ íƒ€ê³µ í•­ëª© ì¶”ê°€
    drilling_items = final_auto_items.get("_drilling", {})
    for item_name, qty in drilling_items.items():
        if qty > 0 and item_name != "í™˜í’ê¸°í™€":  # í™˜í’ê¸°í™€ì€ ì²œì¥íŒ ì†Œê³„ì— ì´ë¯¸ í¬í•¨ë¨
            category_name = "ê°€ê³µ"
            spec_key = f"{category_name}::{item_name}"

            if spec_key not in added_specs:
                rec = find_item(price_df, category_name, item_name)
                if rec is not None:
                    unit_price = rec.get("ë‹¨ê°€", 0) or 0
                else:
                    unit_price = 0
                    warnings.append(f"[ê°€ê³µ] '{item_name}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

                add_row(rows, category_name, item_name, qty, unit_price)
                added_specs.add(spec_key)

    # 7) ì‚¬ìš©ì ì •ì˜ í’ˆëª© ì¶”ê°€
    for custom_item in final_custom_items:
        # ìƒˆë¡œìš´ êµ¬ì¡° (major, sub, spec) ë˜ëŠ” ê¸°ì¡´ êµ¬ì¡° (category, name) ì§€ì›
        major = custom_item.get("major", custom_item.get("category", "ê¸°íƒ€"))
        sub = custom_item.get("sub", "")
        spec = custom_item.get("spec", "")
        qty = custom_item.get("qty", 0)

        # ê¸°ì¡´ êµ¬ì¡° í˜¸í™˜ì„±
        if not major and "name" in custom_item:
            major = custom_item.get("category", "ê¸°íƒ€")
            spec = custom_item.get("name", "")

        if qty <= 0 or not major:
            continue

        # ì¤‘ë³µ ì²´í¬ë¥¼ ìœ„í•œ í‘œì‹œ í…ìŠ¤íŠ¸
        if sub and spec:
            display_text = f"{sub} {spec}"
        elif sub:
            display_text = sub
        elif spec:
            display_text = spec
        else:
            display_text = major

        spec_key = f"{major}::{display_text}"
        if spec_key in added_specs:
            continue

        # ë‹¨ê°€í‘œì—ì„œ ì°¾ê¸° (ëŒ€ë¶„ë¥˜ + í•©ì³ì§„ ì‚¬ì–‘)
        # display_textëŠ” ì´ë¯¸ "ì¤‘ë¶„ë¥˜ ì‚¬ì–‘" í˜•ì‹ìœ¼ë¡œ í•©ì³ì ¸ ìˆìŒ
        rec = find_item(price_df, major, display_text)
        if rec is not None:
            unit_price = rec.get("ë‹¨ê°€", 0) or 0
        else:
            unit_price = 0
            warnings.append(f"[ì‚¬ìš©ìì •ì˜] '{major} > {display_text}' ë‹¨ê°€ ë¯¸ë°œê²¬ â†’ 0 ì²˜ë¦¬")

        add_row(rows, major, display_text, qty, unit_price)
        added_specs.add(spec_key)

# ----------------------------
# ê²°ê³¼ í‘œ
# ----------------------------
if rows:
    est_df = pd.DataFrame(
        rows, columns=["í’ˆëª©", "ì‚¬ì–‘ ë° ê·œê²©", "ìˆ˜ëŸ‰", "ë‹¨ê°€", "ê¸ˆì•¡"]
    )
    est_df["ìˆ˜ëŸ‰"] = (
        pd.to_numeric(est_df["ìˆ˜ëŸ‰"], errors="coerce").fillna(0).astype(float)
    )
    est_df["ë‹¨ê°€"] = (
        pd.to_numeric(est_df["ë‹¨ê°€"], errors="coerce").fillna(0).astype(float)
    )
    est_df["ê¸ˆì•¡"] = (est_df["ìˆ˜ëŸ‰"] * est_df["ë‹¨ê°€"]).round(0)

    # ë””ë²„ê·¸: ë‹¨ê°€í‘œ ì „ì²´ êµ¬ì¡° í™•ì¸
    with st.expander("ğŸ” ë‹¨ê°€í‘œ ë””ë²„ê·¸ ì •ë³´", expanded=False):
        st.markdown("### ğŸ“‹ ìì¬ë‹¨ê°€ë‚´ì—­ ì‹œíŠ¸")
        st.write(f"**ì´ í–‰ ìˆ˜:** {len(price_df)}")
        st.write(f"**ì»¬ëŸ¼:** {list(price_df.columns)}")

        # í’ˆëª© ì»¬ëŸ¼ì˜ ê³ ìœ ê°’ í‘œì‹œ
        if "í’ˆëª©" in price_df.columns:
            unique_items = price_df["í’ˆëª©"].dropna().unique()
            st.write(f"**í’ˆëª©(ëŒ€ë¶„ë¥˜) ê³ ìœ ê°’ (ì´ {len(unique_items)}ê°œ):**")
            st.code(", ".join([f'"{x}"' for x in unique_items]))

        st.write("**ì „ì²´ ìƒ˜í”Œ (ì²˜ìŒ 30í–‰):**")
        st.dataframe(price_df[["í’ˆëª©", "ë¶„ë¥˜", "ì‚¬ì–‘ ë° ê·œê²©", "ë‹¨ê°€"]].head(30))

        st.divider()

        # ìë™ì§€ì •í•­ëª© ì‹œíŠ¸ë„ ì½ì–´ì„œ ë¹„êµ
        st.markdown("### ğŸ“‹ ìë™ì§€ì •í•­ëª© ì‹œíŠ¸ (ë¹„êµìš©)")
        try:
            if pricebook_file is not None:
                pricebook_file.seek(0)
                auto_items_df = pd.read_excel(io.BytesIO(pricebook_file.read()), sheet_name="ìë™ì§€ì •í•­ëª©")

                st.write(f"**ì´ í–‰ ìˆ˜:** {len(auto_items_df)}")
                st.write(f"**ì»¬ëŸ¼:** {list(auto_items_df.columns)}")

                # ì²« ë²ˆì§¸ ì»¬ëŸ¼(Aì—´)ì˜ ê³ ìœ ê°’ í‘œì‹œ
                first_col = auto_items_df.columns[0]
                auto_unique = auto_items_df[first_col].dropna().unique()
                st.write(f"**{first_col} ì»¬ëŸ¼ ê³ ìœ ê°’ (ì´ {len(auto_unique)}ê°œ):**")
                st.code(", ".join([f'"{x}"' for x in auto_unique]))

                st.write("**ì „ì²´ ìƒ˜í”Œ (ì²˜ìŒ 30í–‰):**")
                st.dataframe(auto_items_df.head(30))
        except Exception as e:
            st.warning(f"ìë™ì§€ì •í•­ëª© ì‹œíŠ¸ë¥¼ ì½ì„ ìˆ˜ ì—†ìŠµë‹ˆë‹¤: {e}")

    st.subheader("ê²¬ì ì„œ ë¯¸ë¦¬ë³´ê¸°")

    st.dataframe(est_df, use_container_width=True)

    totals = (
        est_df.groupby("í’ˆëª©", dropna=False)["ê¸ˆì•¡"]
        .sum()
        .reset_index()
        .sort_values("ê¸ˆì•¡", ascending=False)
    )
    st.markdown("#### í’ˆëª©ë³„ í•©ê³„")
    st.dataframe(totals, use_container_width=True)

    grand_total = est_df["ê¸ˆì•¡"].sum()
    st.metric("ì´ ê¸ˆì•¡ (ìƒì‚°ê´€ë¦¬ë¹„ ì œì™¸)", f"{grand_total:,.0f} ì›")

    # ----------------------------
    # ìƒì‚°ê´€ë¦¬ë¹„ ì„¤ì • UI (ê²¬ì ì„œ ë°ì´í„° ê¸°ë°˜)
    # ----------------------------
    st.markdown("---")
    st.subheader("ìƒì‚°ê´€ë¦¬ë¹„ ì„¤ì •")

    # í˜„ì¬ ê²¬ì ì„œì˜ ëª¨ë“  í’ˆëª©+ì‚¬ì–‘ ëª©ë¡ ì¶”ì¶œ
    available_items = []
    for _, row in est_df.iterrows():
        item_key = get_item_key(str(row["í’ˆëª©"]), str(row["ì‚¬ì–‘ ë° ê·œê²©"]))
        if item_key not in [i[0] for i in available_items]:
            available_items.append(
                (
                    item_key,
                    str(row["í’ˆëª©"]),
                    str(row["ì‚¬ì–‘ ë° ê·œê²©"]),
                    float(row["ê¸ˆì•¡"]),
                )
            )

    # ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™”
    if PROD_MGMT_SETTINGS_KEY not in st.session_state:
        st.session_state[PROD_MGMT_SETTINGS_KEY] = {
            cat: {"items": list(info["items"]), "rate": info["rate"]}
            for cat, info in DEFAULT_PROD_MGMT_CATEGORIES.items()
        }

    prod_mgmt_categories = st.session_state[PROD_MGMT_SETTINGS_KEY]

    # ì¹´í…Œê³ ë¦¬ ê´€ë¦¬ UI
    with st.expander("ì¹´í…Œê³ ë¦¬ ê´€ë¦¬ (ì¶”ê°€/ìˆ˜ì •/ì‚­ì œ)", expanded=False):
        # ìƒˆ ì¹´í…Œê³ ë¦¬ ì¶”ê°€
        st.markdown("##### ìƒˆ ì¹´í…Œê³ ë¦¬ ì¶”ê°€")
        col_new_name, col_new_rate, col_new_btn = st.columns([3, 1, 1])
        with col_new_name:
            new_cat_name = st.text_input(
                "ì¹´í…Œê³ ë¦¬ëª…", key="new_cat_name", placeholder="ì˜ˆ: ì‹ ê·œì¹´í…Œê³ ë¦¬"
            )
        with col_new_rate:
            new_cat_rate = st.number_input(
                "ë¹„ìœ¨(%)",
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                step=0.5,
                key="new_cat_rate",
            )
        with col_new_btn:
            st.write("")
            if st.button("â• ì¶”ê°€", key="add_cat_btn"):
                if new_cat_name and new_cat_name not in prod_mgmt_categories:
                    prod_mgmt_categories[new_cat_name] = {
                        "items": [],
                        "rate": new_cat_rate,
                    }
                    st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                    st.success(f"'{new_cat_name}' ì¹´í…Œê³ ë¦¬ ì¶”ê°€ë¨")
                    st.rerun()
                elif new_cat_name in prod_mgmt_categories:
                    st.error("ì´ë¯¸ ì¡´ì¬í•˜ëŠ” ì¹´í…Œê³ ë¦¬ëª…ì…ë‹ˆë‹¤.")

        st.markdown("---")

        # ê¸°ì¡´ ì¹´í…Œê³ ë¦¬ ì‚­ì œ
        st.markdown("##### ì¹´í…Œê³ ë¦¬ ì‚­ì œ")
        if prod_mgmt_categories:
            cat_to_delete = st.selectbox(
                "ì‚­ì œí•  ì¹´í…Œê³ ë¦¬",
                options=list(prod_mgmt_categories.keys()),
                key="cat_to_delete",
            )
            if st.button("ğŸ—‘ï¸ ì„ íƒ ì¹´í…Œê³ ë¦¬ ì‚­ì œ", key="delete_cat_btn"):
                if cat_to_delete in prod_mgmt_categories:
                    del prod_mgmt_categories[cat_to_delete]
                    st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                    st.success(f"'{cat_to_delete}' ì¹´í…Œê³ ë¦¬ ì‚­ì œë¨")
                    st.rerun()

        # ê¸°ë³¸ê°’ìœ¼ë¡œ ì´ˆê¸°í™”
        st.markdown("---")
        if st.button("ğŸ”„ ê¸°ë³¸ ì¹´í…Œê³ ë¦¬ë¡œ ì´ˆê¸°í™”", key="reset_cat_btn"):
            st.session_state[PROD_MGMT_SETTINGS_KEY] = {
                cat: {"items": list(info["items"]), "rate": info["rate"]}
                for cat, info in DEFAULT_PROD_MGMT_CATEGORIES.items()
            }
            st.success("ê¸°ë³¸ ì¹´í…Œê³ ë¦¬ë¡œ ì´ˆê¸°í™”ë¨")
            st.rerun()

    # ì¹´í…Œê³ ë¦¬ë³„ ì„¤ì • UI
    st.markdown("#### ì¹´í…Œê³ ë¦¬ë³„ ë¹„ìœ¨ ë° í¬í•¨ í•­ëª© ì„¤ì •")

    # ê° ì¹´í…Œê³ ë¦¬ë³„ ì„¤ì •
    updated_categories = {}

    for cat_name, cat_info in prod_mgmt_categories.items():
        with st.expander(f"ğŸ“ {cat_name}", expanded=True):
            col_rate, col_info = st.columns([1, 3])

            with col_rate:
                rate = st.number_input(
                    "ë¹„ìœ¨(%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=float(cat_info.get("rate", 0.0)),
                    step=0.5,
                    key=f"rate_{cat_name}",
                )

            # í˜„ì¬ ì¹´í…Œê³ ë¦¬ì— ë§¤ì¹­ë˜ëŠ” í•­ëª© ì°¾ê¸°
            matched_items = []
            unmatched_items = []
            cat_items = cat_info.get("items", [])

            for item_key, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡ in available_items:
                is_matched = False
                for pattern_í’ˆëª©, pattern_ì‚¬ì–‘ in cat_items:
                    if item_matches_pattern(í’ˆëª©, ì‚¬ì–‘, pattern_í’ˆëª©, pattern_ì‚¬ì–‘):
                        is_matched = True
                        break
                if is_matched:
                    matched_items.append((item_key, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡))
                else:
                    unmatched_items.append((item_key, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡))

            with col_info:
                matched_total = sum(ê¸ˆì•¡ for _, _, _, ê¸ˆì•¡ in matched_items)
                mgmt_fee = matched_total * (rate / 100.0)
                st.markdown(
                    f"**ì†Œê³„:** {matched_total:,.0f}ì› â†’ **ìƒì‚°ê´€ë¦¬ë¹„:** {mgmt_fee:,.0f}ì›"
                )

            # í¬í•¨ í•­ëª© í‘œì‹œ ë° í¸ì§‘
            st.markdown("**í¬í•¨ í•­ëª©:**")

            # íŒ¨í„´ ê¸°ë°˜ í•­ëª© í‘œì‹œ
            if cat_items:
                pattern_strs = []
                for p_í’ˆëª©, p_ì‚¬ì–‘ in cat_items:
                    if p_ì‚¬ì–‘:
                        pattern_strs.append(f"{p_í’ˆëª©}({p_ì‚¬ì–‘})")
                    else:
                        pattern_strs.append(f"{p_í’ˆëª©}(ì „ì²´)")
                st.caption(f"íŒ¨í„´: {', '.join(pattern_strs)}")

            # ì‹¤ì œ ë§¤ì¹­ëœ í•­ëª© í‘œì‹œ
            if matched_items:
                matched_df = pd.DataFrame(
                    [
                        {"í’ˆëª©": í’ˆëª©, "ì‚¬ì–‘": ì‚¬ì–‘, "ê¸ˆì•¡": f"{ê¸ˆì•¡:,.0f}"}
                        for _, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡ in matched_items
                    ]
                )
                st.dataframe(
                    matched_df,
                    use_container_width=True,
                    hide_index=True,
                    height=min(150, 35 * (len(matched_items) + 1)),
                )
            else:
                st.info("ë§¤ì¹­ëœ í•­ëª© ì—†ìŒ")

            # í•­ëª© ì¶”ê°€/ì œê±° (íŒ¨í„´ í¸ì§‘)
            st.markdown("**íŒ¨í„´ í¸ì§‘:**")
            col_add_pattern, col_remove_pattern = st.columns(2)

            with col_add_pattern:
                # ì¶”ê°€í•  í’ˆëª© ì„ íƒ
                unique_í’ˆëª©s = list(set(í’ˆëª© for _, í’ˆëª©, _, _ in available_items))
                add_í’ˆëª© = st.selectbox(
                    f"í’ˆëª© ì„ íƒ",
                    options=[""] + unique_í’ˆëª©s,
                    key=f"add_í’ˆëª©_{cat_name}",
                )

                if add_í’ˆëª©:
                    # í•´ë‹¹ í’ˆëª©ì˜ ì‚¬ì–‘ ëª©ë¡
                    í’ˆëª©_ì‚¬ì–‘s = [
                        ì‚¬ì–‘ for _, í’ˆëª©, ì‚¬ì–‘, _ in available_items if í’ˆëª© == add_í’ˆëª©
                    ]
                    add_ì‚¬ì–‘_option = st.selectbox(
                        "ì‚¬ì–‘ ë²”ìœ„",
                        options=["ì „ì²´(í’ˆëª© ì „ì²´ í¬í•¨)"] + í’ˆëª©_ì‚¬ì–‘s,
                        key=f"add_ì‚¬ì–‘_{cat_name}",
                    )

                    if st.button("â• íŒ¨í„´ ì¶”ê°€", key=f"add_pattern_btn_{cat_name}"):
                        if add_ì‚¬ì–‘_option == "ì „ì²´(í’ˆëª© ì „ì²´ í¬í•¨)":
                            new_pattern = (add_í’ˆëª©, None)
                        else:
                            new_pattern = (add_í’ˆëª©, add_ì‚¬ì–‘_option)

                        if new_pattern not in cat_items:
                            cat_items.append(new_pattern)
                            prod_mgmt_categories[cat_name]["items"] = cat_items
                            st.session_state[PROD_MGMT_SETTINGS_KEY] = (
                                prod_mgmt_categories
                            )
                            st.success(f"íŒ¨í„´ ì¶”ê°€ë¨")
                            st.rerun()

            with col_remove_pattern:
                if cat_items:
                    pattern_options = []
                    for p_í’ˆëª©, p_ì‚¬ì–‘ in cat_items:
                        if p_ì‚¬ì–‘:
                            pattern_options.append(f"{p_í’ˆëª©}({p_ì‚¬ì–‘})")
                        else:
                            pattern_options.append(f"{p_í’ˆëª©}(ì „ì²´)")

                    remove_pattern_str = st.selectbox(
                        "ì œê±°í•  íŒ¨í„´",
                        options=pattern_options,
                        key=f"remove_pattern_{cat_name}",
                    )

                    if st.button("â– íŒ¨í„´ ì œê±°", key=f"remove_pattern_btn_{cat_name}"):
                        # íŒ¨í„´ ë¬¸ìì—´ì„ ë‹¤ì‹œ íŠœí”Œë¡œ ë³€í™˜
                        idx = pattern_options.index(remove_pattern_str)
                        cat_items.pop(idx)
                        prod_mgmt_categories[cat_name]["items"] = cat_items
                        st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                        st.success(f"íŒ¨í„´ ì œê±°ë¨")
                        st.rerun()

            updated_categories[cat_name] = {"items": cat_items, "rate": rate}

    # ì„¤ì • ì—…ë°ì´íŠ¸
    st.session_state[PROD_MGMT_SETTINGS_KEY] = updated_categories
    prod_mgmt_categories = updated_categories

    # ----------------------------
    # ìƒì‚°ê´€ë¦¬ë¹„ ì¹´í…Œê³ ë¦¬ë³„ í•©ê³„ ê³„ì‚° ë° í‘œì‹œ
    # ----------------------------
    st.markdown("---")
    st.markdown("#### ìƒì‚°ê´€ë¦¬ë¹„ ì¹´í…Œê³ ë¦¬ë³„ í•©ê³„")

    # ê° í•­ëª©ì´ ì–´ëŠ ì¹´í…Œê³ ë¦¬ì— ì†í•˜ëŠ”ì§€ ë§¤í•‘
    item_to_category = {}
    for item_key, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡ in available_items:
        for cat_name, cat_info in prod_mgmt_categories.items():
            for pattern_í’ˆëª©, pattern_ì‚¬ì–‘ in cat_info.get("items", []):
                if item_matches_pattern(í’ˆëª©, ì‚¬ì–‘, pattern_í’ˆëª©, pattern_ì‚¬ì–‘):
                    item_to_category[item_key] = cat_name
                    break
            if item_key in item_to_category:
                break
        if item_key not in item_to_category:
            item_to_category[item_key] = "ë¯¸ë¶„ë¥˜"

    # ì¹´í…Œê³ ë¦¬ë³„ ì†Œê³„ ê³„ì‚°
    category_subtotals = {cat_name: 0.0 for cat_name in prod_mgmt_categories.keys()}
    category_subtotals["ë¯¸ë¶„ë¥˜"] = 0.0

    for item_key, í’ˆëª©, ì‚¬ì–‘, ê¸ˆì•¡ in available_items:
        cat = item_to_category.get(item_key, "ë¯¸ë¶„ë¥˜")
        category_subtotals[cat] += ê¸ˆì•¡

    # ìƒì‚°ê´€ë¦¬ë¹„ ê³„ì‚°
    category_mgmt_fees = {}
    total_mgmt_fee = 0.0

    for cat_name, subtotal in category_subtotals.items():
        if cat_name == "ë¯¸ë¶„ë¥˜":
            rate = 0.0
        else:
            rate = prod_mgmt_categories.get(cat_name, {}).get("rate", 0.0)
        mgmt_fee = subtotal * (rate / 100.0)
        category_mgmt_fees[cat_name] = mgmt_fee
        total_mgmt_fee += mgmt_fee

    # í‘œ í˜•ì‹ìœ¼ë¡œ í‘œì‹œ
    mgmt_summary_data = []
    for cat_name in list(prod_mgmt_categories.keys()) + (
        ["ë¯¸ë¶„ë¥˜"] if category_subtotals.get("ë¯¸ë¶„ë¥˜", 0) > 0 else []
    ):
        subtotal = category_subtotals.get(cat_name, 0)
        if cat_name == "ë¯¸ë¶„ë¥˜":
            rate = 0.0
        else:
            rate = prod_mgmt_categories.get(cat_name, {}).get("rate", 0.0)
        mgmt_fee = category_mgmt_fees.get(cat_name, 0)
        total_with_mgmt = subtotal + mgmt_fee
        mgmt_summary_data.append(
            {
                "ì¹´í…Œê³ ë¦¬": cat_name,
                "ì†Œê³„": f"{subtotal:,.0f}",
                "ë¹„ìœ¨(%)": f"{rate:.1f}",
                "ìƒì‚°ê´€ë¦¬ë¹„": f"{mgmt_fee:,.0f}",
                "ì´ê³„": f"{total_with_mgmt:,.0f}",
            }
        )

    mgmt_summary_df = pd.DataFrame(mgmt_summary_data)
    st.dataframe(mgmt_summary_df, use_container_width=True, hide_index=True)

    # ----------------------------
    # ì˜ì—…ê´€ë¦¬ë¹„ ì„¤ì • UI
    # ----------------------------
    st.markdown("---")
    st.subheader("ì˜ì—…ê´€ë¦¬ë¹„ ì„¤ì • (ì„ íƒ)")

    # ì˜ì—…ê´€ë¦¬ë¹„ ì„¸ì…˜ ìƒíƒœ ì´ˆê¸°í™”
    if SALES_MGMT_SETTINGS_KEY not in st.session_state:
        st.session_state[SALES_MGMT_SETTINGS_KEY] = {
            "enabled": False,
            "rate": 15.0,  # ê¸°ë³¸ê°’ 15%
        }

    sales_settings = st.session_state[SALES_MGMT_SETTINGS_KEY]

    col_sales_enable, col_sales_rate = st.columns([1, 2])
    with col_sales_enable:
        sales_enabled = st.checkbox(
            "ì˜ì—…ê´€ë¦¬ë¹„ ì¶”ê°€",
            value=sales_settings.get("enabled", False),
            help="ì²´í¬í•˜ë©´ ì˜ì—…ê´€ë¦¬ë¹„ê°€ ê²¬ì ì„œì— í¬í•¨ë©ë‹ˆë‹¤",
        )
    with col_sales_rate:
        if sales_enabled:
            sales_rate = st.number_input(
                "ì˜ì—…ê´€ë¦¬ë¹„ ë¹„ìœ¨(%)",
                min_value=0.0,
                max_value=100.0,
                value=float(sales_settings.get("rate", 15.0)),
                step=0.5,
                key="sales_mgmt_rate",
            )
        else:
            sales_rate = 0.0
            st.info("ì˜ì—…ê´€ë¦¬ë¹„ë¥¼ ì¶”ê°€í•˜ë ¤ë©´ ì²´í¬ë°•ìŠ¤ë¥¼ ì„ íƒí•˜ì„¸ìš”")

    # ì˜ì—…ê´€ë¦¬ë¹„ ì„¤ì • ì—…ë°ì´íŠ¸
    st.session_state[SALES_MGMT_SETTINGS_KEY] = {
        "enabled": sales_enabled,
        "rate": sales_rate if sales_enabled else 0.0,
    }

    # ì˜ì—…ê´€ë¦¬ë¹„ ê³„ì‚°
    total_before_sales = grand_total + total_mgmt_fee  # ì›ê°€ + ìƒì‚°ê´€ë¦¬ë¹„
    sales_mgmt_fee = total_before_sales * (sales_rate / 100.0) if sales_enabled else 0.0

    if sales_enabled:
        st.markdown(
            f"**ì˜ì—…ê´€ë¦¬ë¹„ ê¸°ì¤€ê¸ˆì•¡:** {total_before_sales:,.0f}ì› Ã— {sales_rate:.1f}% = **{sales_mgmt_fee:,.0f}ì›**"
        )

    # ìµœì¢… ì´ê³„ (ì˜ì—…ê´€ë¦¬ë¹„ í¬í•¨)
    final_total = grand_total + total_mgmt_fee + sales_mgmt_fee

    if sales_enabled:
        col_sub, col_mgmt, col_sales, col_final = st.columns(4)
        with col_sub:
            st.metric("ì›ê°€ ì†Œê³„", f"{grand_total:,.0f} ì›")
        with col_mgmt:
            st.metric("ìƒì‚°ê´€ë¦¬ë¹„", f"{total_mgmt_fee:,.0f} ì›")
        with col_sales:
            st.metric("ì˜ì—…ê´€ë¦¬ë¹„", f"{sales_mgmt_fee:,.0f} ì›")
        with col_final:
            st.metric("ìµœì¢… ì´ê³„", f"{final_total:,.0f} ì›")
    else:
        col_sub, col_mgmt, col_final = st.columns(3)
        with col_sub:
            st.metric("ì›ê°€ ì†Œê³„", f"{grand_total:,.0f} ì›")
        with col_mgmt:
            st.metric("ìƒì‚°ê´€ë¦¬ë¹„ í•©ê³„", f"{total_mgmt_fee:,.0f} ì›")
        with col_final:
            st.metric("ìµœì¢… ì´ê³„", f"{final_total:,.0f} ì›")

    # ----------------------------
    # ì„¸ëŒ€ íƒ€ì… ì €ì¥ ê¸°ëŠ¥
    # ----------------------------
    st.markdown("---")
    st.subheader("ì„¸ëŒ€ íƒ€ì… ì €ì¥")

    # ì €ì¥ëœ ê²¬ì  ëª©ë¡ ì´ˆê¸°í™”
    if SAVED_QUOTATIONS_KEY not in st.session_state:
        st.session_state[SAVED_QUOTATIONS_KEY] = []

    # í˜„ì¬ ì„¸ëŒ€ ì •ë³´
    current_spec = floor_data.get("ê·œê²©", "N/A") if floor_data else "N/A"
    current_units = floor_data.get("inputs", {}).get("units", 1) if floor_data else 1

    col_name, col_save = st.columns([3, 1])
    with col_name:
        type_name = st.text_input(
            "ì„¸ëŒ€ íƒ€ì… ì´ë¦„",
            value=f"íƒ€ì…{len(st.session_state[SAVED_QUOTATIONS_KEY]) + 1}",
            help="ì˜ˆ: 21A,B,E/22C,F",
        )
    with col_save:
        st.write("")  # ê³µë°±ìœ¼ë¡œ ë†’ì´ ë§ì¶¤
        st.write("")
        save_disabled = len(st.session_state[SAVED_QUOTATIONS_KEY]) >= 10
        if st.button(
            "ğŸ’¾ í˜„ì¬ ê²¬ì  ì €ì¥", disabled=save_disabled, help="ìµœëŒ€ 10ê°œê¹Œì§€ ì €ì¥ ê°€ëŠ¥"
        ):
            # í˜„ì¬ ê²¬ì  ë°ì´í„° ì €ì¥ (ìƒì‚°ê´€ë¦¬ë¹„ ì •ë³´ í¬í•¨)
            # prod_mgmt_settingsì˜ itemsë¥¼ ë¦¬ìŠ¤íŠ¸ë¡œ ë³€í™˜ (íŠœí”Œì€ JSON ì§ë ¬í™” ë¬¸ì œ)
            serializable_settings = {}
            for k, v in prod_mgmt_categories.items():
                serializable_settings[k] = {
                    "items": [list(item) for item in v.get("items", [])],
                    "rate": v.get("rate", 0.0),
                }
            quotation_data = {
                "name": type_name,
                "spec": current_spec,
                "units": current_units,
                "rows": rows.copy(),  # ê²¬ì  í•­ëª© ëª©ë¡
                "total": grand_total,  # ì›ê°€ ì†Œê³„
                "total_mgmt_fee": total_mgmt_fee,  # ìƒì‚°ê´€ë¦¬ë¹„ í•©ê³„
                "sales_mgmt_fee": sales_mgmt_fee,  # ì˜ì—…ê´€ë¦¬ë¹„
                "sales_mgmt_rate": (
                    sales_rate if sales_enabled else 0.0
                ),  # ì˜ì—…ê´€ë¦¬ë¹„ ë¹„ìœ¨
                "sales_mgmt_enabled": sales_enabled,  # ì˜ì—…ê´€ë¦¬ë¹„ í™œì„±í™” ì—¬ë¶€
                "final_total": final_total,  # ìµœì¢… ì´ê³„ (ì›ê°€ + ìƒì‚°ê´€ë¦¬ë¹„ + ì˜ì—…ê´€ë¦¬ë¹„)
                "category_subtotals": dict(category_subtotals),  # ì¹´í…Œê³ ë¦¬ë³„ ì†Œê³„
                "category_mgmt_fees": dict(category_mgmt_fees),  # ì¹´í…Œê³ ë¦¬ë³„ ìƒì‚°ê´€ë¦¬ë¹„
                "prod_mgmt_settings": serializable_settings,  # ìƒì‚°ê´€ë¦¬ë¹„ ì„¤ì •
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            st.session_state[SAVED_QUOTATIONS_KEY].append(quotation_data)
            st.success(
                f"âœ… '{type_name}' ì €ì¥ ì™„ë£Œ! (ê·œê²©: {current_spec}, {current_units}ì„¸ëŒ€, ìµœì¢…ë‹¨ê°€: {final_total:,.0f}ì›)"
            )
            st.rerun()

    # ì €ì¥ëœ ì„¸ëŒ€ íƒ€ì… ëª©ë¡ í‘œì‹œ
    saved_list = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
    if saved_list:
        st.markdown("#### ì €ì¥ëœ ì„¸ëŒ€ íƒ€ì… ëª©ë¡")

        # í…Œì´ë¸” í˜•ì‹ìœ¼ë¡œ í‘œì‹œ (ìƒì‚°ê´€ë¦¬ë¹„, ì˜ì—…ê´€ë¦¬ë¹„ í¬í•¨)
        saved_df = pd.DataFrame(
            [
                {
                    "ë²ˆí˜¸": i + 1,
                    "íƒ€ì…ëª…": q["name"],
                    "ê·œê²©": q["spec"],
                    "ì„¸ëŒ€ìˆ˜": q["units"],
                    "ì›ê°€ ì†Œê³„": f"{q['total']:,.0f}",
                    "ìƒì‚°ê´€ë¦¬ë¹„": f"{q.get('total_mgmt_fee', 0):,.0f}",
                    "ì˜ì—…ê´€ë¦¬ë¹„": (
                        f"{q.get('sales_mgmt_fee', 0):,.0f}"
                        if q.get("sales_mgmt_enabled", False)
                        else "-"
                    ),
                    "ì„¸ëŒ€ë‹¹ ìµœì¢…ë‹¨ê°€": f"{q.get('final_total', q['total']):,.0f}",
                    "ì´ ê¸ˆì•¡": f"{q.get('final_total', q['total']) * q['units']:,.0f}",
                }
                for i, q in enumerate(saved_list)
            ]
        )
        st.dataframe(saved_df, use_container_width=True, hide_index=True)

        # ì‚­ì œ ê¸°ëŠ¥
        col_del, col_clear = st.columns([2, 1])
        with col_del:
            if len(saved_list) > 0:
                del_idx = st.selectbox(
                    "ì‚­ì œí•  íƒ€ì… ì„ íƒ",
                    options=range(len(saved_list)),
                    format_func=lambda x: f"{x+1}. {saved_list[x]['name']} ({saved_list[x]['spec']})",
                )
                if st.button("ğŸ—‘ï¸ ì„ íƒ í•­ëª© ì‚­ì œ"):
                    del st.session_state[SAVED_QUOTATIONS_KEY][del_idx]
                    st.success("ì‚­ì œ ì™„ë£Œ!")
                    st.rerun()
        with col_clear:
            st.write("")
            if st.button("ğŸ—‘ï¸ ì „ì²´ ì‚­ì œ", type="secondary"):
                st.session_state[SAVED_QUOTATIONS_KEY] = []
                st.success("ì „ì²´ ì‚­ì œ ì™„ë£Œ!")
                st.rerun()

        # ì´ ì„¸ëŒ€ìˆ˜ ë° ì´ ê¸ˆì•¡ í•©ê³„ (ìƒì‚°ê´€ë¦¬ë¹„, ì˜ì—…ê´€ë¦¬ë¹„ í¬í•¨)
        total_all_units = sum(q["units"] for q in saved_list)
        total_all_amount = sum(
            q.get("final_total", q["total"]) * q["units"] for q in saved_list
        )
        total_all_cost = sum(q["total"] * q["units"] for q in saved_list)
        total_all_mgmt = sum(
            q.get("total_mgmt_fee", 0) * q["units"] for q in saved_list
        )
        total_all_sales = sum(
            q.get("sales_mgmt_fee", 0) * q["units"]
            for q in saved_list
            if q.get("sales_mgmt_enabled", False)
        )
        if total_all_sales > 0:
            st.markdown(
                f"**ì´ ì„¸ëŒ€ìˆ˜: {total_all_units}ì„¸ëŒ€ | ì›ê°€í•©ê³„: {total_all_cost:,.0f}ì› | ìƒì‚°ê´€ë¦¬ë¹„: {total_all_mgmt:,.0f}ì› | ì˜ì—…ê´€ë¦¬ë¹„: {total_all_sales:,.0f}ì› | ìµœì¢…í•©ê³„: {total_all_amount:,.0f}ì›**"
            )
        else:
            st.markdown(
                f"**ì´ ì„¸ëŒ€ìˆ˜: {total_all_units}ì„¸ëŒ€ | ì›ê°€í•©ê³„: {total_all_cost:,.0f}ì› | ìƒì‚°ê´€ë¦¬ë¹„í•©ê³„: {total_all_mgmt:,.0f}ì› | ìµœì¢…í•©ê³„: {total_all_amount:,.0f}ì›**"
            )

    st.markdown("---")

    # Excel ë‹¤ìš´ë¡œë“œ (LGE ì°½ì› ìŠ¤ë§ˆíŠ¸íŒŒí¬ í˜•ì‹)
    def df_to_excel_bytes(
        df: pd.DataFrame,
        total_units: int = 1,
        category_subtotals: dict = None,
        category_mgmt_fees: dict = None,
        prod_mgmt_settings: dict = None,
        total_mgmt_fee: float = 0.0,
        final_total: float = 0.0,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # ê°™ì€ ëŒ€ë¶„ë¥˜(í’ˆëª©)ë¼ë¦¬ ì—°ì† ë°°ì¹˜ë˜ë„ë¡ ì •ë ¬
        # ì›ë˜ ë“±ì¥ ìˆœì„œë¥¼ ìœ ì§€í•˜ë©´ì„œ ê°™ì€ í’ˆëª©ë¼ë¦¬ ëª¨ìŒ
        category_order = {}
        for i, cat in enumerate(df["í’ˆëª©"]):
            if cat not in category_order:
                category_order[cat] = i
        df = df.copy()
        df["_sort_key"] = df["í’ˆëª©"].map(category_order)
        df = (
            df.sort_values("_sort_key")
            .drop(columns=["_sort_key"])
            .reset_index(drop=True)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "ì›ìì¬ ì„¸ëŒ€ë‹¹ ë‹¨ê°€ë‚´ì—­"

        # A4 ê°€ë¡œ í˜•ì‹ ì„¤ì •
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # ë†’ì´ëŠ” ìë™

        # ê°€ìš´ë° ì •ë ¬ì„ ìœ„í•´ ì™¼ìª½ ì—¬ë°± ì»¬ëŸ¼ ì¶”ê°€
        LEFT_MARGIN = 3  # ì™¼ìª½ ì—¬ë°± ì»¬ëŸ¼ ìˆ˜ (ë” ë„“ê²Œ)

        # ìŠ¤íƒ€ì¼ ì •ì˜
        title_font = Font(name="ë§‘ì€ ê³ ë”•", size=18, bold=True)
        subtitle_font = Font(name="ë§‘ì€ ê³ ë”•", size=11, bold=True)
        header_font = Font(name="ë§‘ì€ ê³ ë”•", size=10, bold=True)
        data_font = Font(name="ë§‘ì€ ê³ ë”•", size=9)
        small_font = Font(name="ë§‘ì€ ê³ ë”•", size=8)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # íˆ¬ëª… ë°°ê²½ (fill ì œê±°)
        no_fill = PatternFill(fill_type=None)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # ì—¬ë°± ì»¬ëŸ¼ ì„¤ì •
        for i in range(1, LEFT_MARGIN + 1):
            ws.column_dimensions[chr(64 + i)].width = 2

        # ì‹¤ì œ ì‹œì‘ ì»¬ëŸ¼ (Cë¶€í„°)
        START_COL = LEFT_MARGIN + 1

        # 1í–‰: íƒ€ì´í‹€ - ê°€ë¡œë¡œ ë„“ê²Œ
        title_range = f"{chr(64+START_COL)}1:{chr(64+START_COL+7)}1"
        ws.merge_cells(title_range)
        title_cell = ws.cell(1, START_COL)
        title_cell.value = "ìš•ì‹¤ ì›ìì¬ ì„¸ëŒ€ë‹¹ ë‹¨ê°€ ë‚´ì—­"
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30

        # 2-3í–‰: ë¹ˆ í–‰
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 10

        # 4í–‰: ì„¸ëŒ€ ì •ë³´ ë° ë‚ ì§œ
        info_range = f"{chr(64+START_COL)}4:{chr(64+START_COL+2)}4"
        ws.merge_cells(info_range)
        info_cell = ws.cell(4, START_COL)
        info_cell.value = f"ì´ ì„¸ëŒ€ìˆ˜: {total_units}ì„¸ëŒ€"
        info_cell.font = subtitle_font
        info_cell.alignment = left_align

        date_range = f"{chr(64+START_COL+5)}4:{chr(64+START_COL+7)}4"
        ws.merge_cells(date_range)
        date_cell = ws.cell(4, START_COL + 5)
        date_cell.value = f"ì‘ì„±ì¼: {datetime.now():%Y. %m. %d}"
        date_cell.font = subtitle_font
        date_cell.alignment = right_align

        # 5í–‰: ì»¬ëŸ¼ í—¤ë” (ë‹¨ì¼ ì„¸ëŒ€ íƒ€ì…) - í…Œë‘ë¦¬ ì¶”ê°€, ë°°ê²½ íˆ¬ëª…
        # í’ˆëª© (C5:D5)
        í’ˆëª©_range = f"{chr(64+START_COL)}5:{chr(64+START_COL+1)}5"
        ws.merge_cells(í’ˆëª©_range)
        ws.cell(5, START_COL).value = "í’ˆëª©"
        ws.cell(5, START_COL).font = header_font
        ws.cell(5, START_COL).alignment = center_align
        for i in range(START_COL, START_COL + 2):
            ws.cell(5, i).border = thin_border

        # ì„¸ëŒ€ë‹¹ ë‹¨ê°€ (E5:G5)
        ì„¸ëŒ€ë‹¹_range = f"{chr(64+START_COL+2)}5:{chr(64+START_COL+4)}5"
        ws.merge_cells(ì„¸ëŒ€ë‹¹_range)
        ws.cell(5, START_COL + 2).value = "ì„¸ëŒ€ë‹¹ ë‹¨ê°€"
        ws.cell(5, START_COL + 2).font = header_font
        ws.cell(5, START_COL + 2).alignment = center_align
        for i in range(START_COL + 2, START_COL + 5):
            ws.cell(5, i).border = thin_border

        # ì´ ê¸ˆì•¡ (H5:J5)
        ì´ê¸ˆì•¡_range = f"{chr(64+START_COL+5)}5:{chr(64+START_COL+7)}5"
        ws.merge_cells(ì´ê¸ˆì•¡_range)
        ws.cell(5, START_COL + 5).value = f"ì´ ê¸ˆì•¡ ({total_units}ì„¸ëŒ€)"
        ws.cell(5, START_COL + 5).font = header_font
        ws.cell(5, START_COL + 5).alignment = center_align
        for i in range(START_COL + 5, START_COL + 8):
            ws.cell(5, i).border = thin_border

        # 6í–‰: ì„¸ë¶€ ì»¬ëŸ¼ í—¤ë” (ë°°ê²½ íˆ¬ëª…)
        headers_6 = [
            "ëŒ€ë¶„ë¥˜",
            "ì‚¬ì–‘ ë° ê·œê²©",
            "ìˆ˜ëŸ‰",
            "ë‹¨ê°€",
            "ê¸ˆì•¡",
            "ìˆ˜ëŸ‰",
            "ë‹¨ê°€",
            "ê¸ˆì•¡",
        ]
        for idx, header_text in enumerate(headers_6):
            cell = ws.cell(6, START_COL + idx)
            cell.value = header_text
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # ì»¬ëŸ¼ ë„ˆë¹„ ì„¤ì • (ê°€ë¡œë¡œ ë„“ê²Œ)
        ws.column_dimensions[chr(64 + START_COL)].width = 12  # ëŒ€ë¶„ë¥˜
        ws.column_dimensions[chr(64 + START_COL + 1)].width = 38  # ì‚¬ì–‘ ë° ê·œê²©
        ws.column_dimensions[chr(64 + START_COL + 2)].width = 9  # ìˆ˜ëŸ‰
        ws.column_dimensions[chr(64 + START_COL + 3)].width = 13  # ë‹¨ê°€
        ws.column_dimensions[chr(64 + START_COL + 4)].width = 15  # ê¸ˆì•¡
        ws.column_dimensions[chr(64 + START_COL + 5)].width = 9  # ìˆ˜ëŸ‰(ì´)
        ws.column_dimensions[chr(64 + START_COL + 6)].width = 13  # ë‹¨ê°€(ì´)
        ws.column_dimensions[chr(64 + START_COL + 7)].width = 17  # ê¸ˆì•¡(ì´)

        # ë°ì´í„° í–‰ ì‘ì„±
        row_num = 7
        current_category = None

        # ê° ëŒ€ë¶„ë¥˜ë³„ ì‹œì‘/ë í–‰ ê³„ì‚°
        category_rows = {}
        temp_row = 7
        for idx, row_data in df.iterrows():
            í’ˆëª© = str(row_data["í’ˆëª©"])
            if í’ˆëª© not in category_rows:
                category_rows[í’ˆëª©] = {"start": temp_row, "end": temp_row}
            else:
                category_rows[í’ˆëª©]["end"] = temp_row
            temp_row += 1

        # ëŒ€ë¶„ë¥˜ ì—´ í…Œë‘ë¦¬ ìŠ¤íƒ€ì¼ ì •ì˜
        top_only_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style=None),
        )
        middle_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style=None),
            bottom=Side(style=None),
        )
        bottom_only_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style=None),
            bottom=Side(style="thin"),
        )

        for idx, row_data in df.iterrows():
            í’ˆëª© = str(row_data["í’ˆëª©"])
            ì‚¬ì–‘ = str(row_data["ì‚¬ì–‘ ë° ê·œê²©"])
            ìˆ˜ëŸ‰ = float(row_data["ìˆ˜ëŸ‰"])
            ë‹¨ê°€ = float(row_data["ë‹¨ê°€"])
            ê¸ˆì•¡ = float(row_data["ê¸ˆì•¡"])

            # ëŒ€ë¶„ë¥˜ (í’ˆëª©ì´ ë°”ë€” ë•Œë§Œ í‘œì‹œ)
            cell_a = ws.cell(row=row_num, column=START_COL)
            if í’ˆëª© != current_category:
                cell_a.value = í’ˆëª©
                current_category = í’ˆëª©
            else:
                cell_a.value = ""
            cell_a.font = data_font
            cell_a.alignment = left_align

            # ëŒ€ë¶„ë¥˜ ì—´ í…Œë‘ë¦¬ ì„¤ì •
            cat_info = category_rows.get(í’ˆëª©, {})
            start_row = cat_info.get("start", row_num)
            end_row = cat_info.get("end", row_num)

            if start_row == end_row:
                # ë‹¨ì¼ í–‰ì´ë©´ ì „ì²´ í…Œë‘ë¦¬
                cell_a.border = thin_border
            elif row_num == start_row:
                # ì²« í–‰: ìœ„ í…Œë‘ë¦¬ë§Œ
                cell_a.border = top_only_border
            elif row_num == end_row:
                # ë§ˆì§€ë§‰ í–‰: ì•„ë˜ í…Œë‘ë¦¬ë§Œ
                cell_a.border = bottom_only_border
            else:
                # ì¤‘ê°„ í–‰: ì¢Œìš°ë§Œ
                cell_a.border = middle_border

            # ì‚¬ì–‘ ë° ê·œê²©
            ws.cell(row=row_num, column=START_COL + 1).value = ì‚¬ì–‘
            ws.cell(row=row_num, column=START_COL + 1).font = data_font
            ws.cell(row=row_num, column=START_COL + 1).alignment = left_align
            ws.cell(row=row_num, column=START_COL + 1).border = thin_border

            # ì„¸ëŒ€ë‹¹ ë‹¨ê°€ (C-E)
            ws.cell(row=row_num, column=START_COL + 2).value = ìˆ˜ëŸ‰
            ws.cell(row=row_num, column=START_COL + 2).font = data_font
            ws.cell(row=row_num, column=START_COL + 2).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 2).border = thin_border
            ws.cell(row=row_num, column=START_COL + 2).number_format = "#,##0.##"

            ws.cell(row=row_num, column=START_COL + 3).value = ë‹¨ê°€
            ws.cell(row=row_num, column=START_COL + 3).font = data_font
            ws.cell(row=row_num, column=START_COL + 3).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 3).border = thin_border
            ws.cell(row=row_num, column=START_COL + 3).number_format = "#,##0"

            ws.cell(row=row_num, column=START_COL + 4).value = ê¸ˆì•¡
            ws.cell(row=row_num, column=START_COL + 4).font = data_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            # ì´ ê¸ˆì•¡ (F-H) - ì„¸ëŒ€ìˆ˜ ê³±í•˜ê¸°
            ws.cell(row=row_num, column=START_COL + 5).value = ìˆ˜ëŸ‰ * total_units
            ws.cell(row=row_num, column=START_COL + 5).font = data_font
            ws.cell(row=row_num, column=START_COL + 5).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 5).border = thin_border
            ws.cell(row=row_num, column=START_COL + 5).number_format = "#,##0.##"

            ws.cell(row=row_num, column=START_COL + 6).value = ë‹¨ê°€
            ws.cell(row=row_num, column=START_COL + 6).font = data_font
            ws.cell(row=row_num, column=START_COL + 6).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 6).border = thin_border
            ws.cell(row=row_num, column=START_COL + 6).number_format = "#,##0"

            ws.cell(row=row_num, column=START_COL + 7).value = ê¸ˆì•¡ * total_units
            ws.cell(row=row_num, column=START_COL + 7).font = data_font
            ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

            row_num += 1

        # í•©ê³„ í–‰ (ë°°ê²½ íˆ¬ëª…)
        ws.cell(row=row_num, column=START_COL).value = "í•©ê³„"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border

        ws.cell(row=row_num, column=START_COL + 1).value = "(V.A.T ë³„ë„)"
        ws.cell(row=row_num, column=START_COL + 1).font = header_font
        ws.cell(row=row_num, column=START_COL + 1).alignment = center_align
        ws.cell(row=row_num, column=START_COL + 1).border = thin_border

        # ì„¸ëŒ€ë‹¹ í•©ê³„
        for col in [START_COL + 2, START_COL + 3]:
            ws.cell(row=row_num, column=col).value = ""
            ws.cell(row=row_num, column=col).border = thin_border

        ws.cell(row=row_num, column=START_COL + 4).value = df["ê¸ˆì•¡"].sum()
        ws.cell(row=row_num, column=START_COL + 4).font = header_font
        ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
        ws.cell(row=row_num, column=START_COL + 4).border = thin_border
        ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

        # ì´ í•©ê³„
        for col in [START_COL + 5, START_COL + 6]:
            ws.cell(row=row_num, column=col).value = ""
            ws.cell(row=row_num, column=col).border = thin_border

        ws.cell(row=row_num, column=START_COL + 7).value = (
            df["ê¸ˆì•¡"].sum() * total_units
        )
        ws.cell(row=row_num, column=START_COL + 7).font = header_font
        ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
        ws.cell(row=row_num, column=START_COL + 7).border = thin_border
        ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"
        row_num += 1

        # ----------------------------
        # ìƒì‚°ê´€ë¦¬ë¹„ ì¹´í…Œê³ ë¦¬ë³„ í•©ê³„ ì„¹ì…˜
        # ----------------------------
        if category_subtotals and prod_mgmt_settings:
            row_num += 1  # ë¹ˆ í–‰

            # ìƒì‚°ê´€ë¦¬ë¹„ ì„¹ì…˜ ì œëª©
            ws.merge_cells(
                start_row=row_num,
                start_column=START_COL,
                end_row=row_num,
                end_column=START_COL + 7,
            )
            ws.cell(row=row_num, column=START_COL).value = "ìƒì‚°ê´€ë¦¬ë¹„ ì¹´í…Œê³ ë¦¬ë³„ í•©ê³„"
            ws.cell(row=row_num, column=START_COL).font = subtitle_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            row_num += 1

            # í—¤ë” í–‰
            mgmt_headers = [
                "ì¹´í…Œê³ ë¦¬",
                "",
                "ì†Œê³„",
                "ë¹„ìœ¨(%)",
                "ìƒì‚°ê´€ë¦¬ë¹„",
                "",
                "",
                "ì´ê³„",
            ]
            for idx, h in enumerate(mgmt_headers):
                cell = ws.cell(row=row_num, column=START_COL + idx)
                cell.value = h
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1

            # ì¹´í…Œê³ ë¦¬ë³„ ë°ì´í„°
            for cat_name, cat_info in prod_mgmt_settings.items():
                subtotal = category_subtotals.get(cat_name, 0)
                rate = cat_info.get("rate", 0.0)
                mgmt_fee = (
                    category_mgmt_fees.get(cat_name, 0) if category_mgmt_fees else 0
                )
                cat_total = subtotal + mgmt_fee

                # ì¹´í…Œê³ ë¦¬ëª… (2ì¹¸ ë³‘í•©)
                ws.merge_cells(
                    start_row=row_num,
                    start_column=START_COL,
                    end_row=row_num,
                    end_column=START_COL + 1,
                )
                ws.cell(row=row_num, column=START_COL).value = cat_name
                ws.cell(row=row_num, column=START_COL).font = data_font
                ws.cell(row=row_num, column=START_COL).alignment = left_align
                ws.cell(row=row_num, column=START_COL).border = thin_border
                ws.cell(row=row_num, column=START_COL + 1).border = thin_border

                # ì†Œê³„
                ws.cell(row=row_num, column=START_COL + 2).value = subtotal
                ws.cell(row=row_num, column=START_COL + 2).font = data_font
                ws.cell(row=row_num, column=START_COL + 2).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 2).border = thin_border
                ws.cell(row=row_num, column=START_COL + 2).number_format = "#,##0"

                # ë¹„ìœ¨
                ws.cell(row=row_num, column=START_COL + 3).value = rate
                ws.cell(row=row_num, column=START_COL + 3).font = data_font
                ws.cell(row=row_num, column=START_COL + 3).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 3).border = thin_border
                ws.cell(row=row_num, column=START_COL + 3).number_format = "0.0"

                # ìƒì‚°ê´€ë¦¬ë¹„
                ws.cell(row=row_num, column=START_COL + 4).value = mgmt_fee
                ws.cell(row=row_num, column=START_COL + 4).font = data_font
                ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 4).border = thin_border
                ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

                # ë¹ˆ ì¹¸
                ws.cell(row=row_num, column=START_COL + 5).value = ""
                ws.cell(row=row_num, column=START_COL + 5).border = thin_border
                ws.cell(row=row_num, column=START_COL + 6).value = ""
                ws.cell(row=row_num, column=START_COL + 6).border = thin_border

                # ì´ê³„
                ws.cell(row=row_num, column=START_COL + 7).value = cat_total
                ws.cell(row=row_num, column=START_COL + 7).font = data_font
                ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 7).border = thin_border
                ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

                row_num += 1

            # ìƒì‚°ê´€ë¦¬ë¹„ í•©ê³„ í–‰
            ws.merge_cells(
                start_row=row_num,
                start_column=START_COL,
                end_row=row_num,
                end_column=START_COL + 1,
            )
            ws.cell(row=row_num, column=START_COL).value = "ìƒì‚°ê´€ë¦¬ë¹„ í•©ê³„"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=START_COL + 1).border = thin_border

            for col in [START_COL + 2, START_COL + 3]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            ws.cell(row=row_num, column=START_COL + 4).value = total_mgmt_fee
            ws.cell(row=row_num, column=START_COL + 4).font = header_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            for col in [START_COL + 5, START_COL + 6]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            ws.cell(row=row_num, column=START_COL + 7).value = ""
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            row_num += 1

            # ìµœì¢… ì´ê³„ í–‰
            row_num += 1
            ws.merge_cells(
                start_row=row_num,
                start_column=START_COL,
                end_row=row_num,
                end_column=START_COL + 3,
            )
            ws.cell(row=row_num, column=START_COL).value = (
                "ìµœì¢… ì´ê³„ (ì›ê°€ + ìƒì‚°ê´€ë¦¬ë¹„)"
            )
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            for col in range(START_COL + 1, START_COL + 4):
                ws.cell(row=row_num, column=col).border = thin_border

            # ì„¸ëŒ€ë‹¹ ìµœì¢…
            ws.cell(row=row_num, column=START_COL + 4).value = final_total
            ws.cell(row=row_num, column=START_COL + 4).font = header_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            for col in [START_COL + 5, START_COL + 6]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            # ì´ ì„¸ëŒ€ ìµœì¢…
            ws.cell(row=row_num, column=START_COL + 7).value = final_total * total_units
            ws.cell(row=row_num, column=START_COL + 7).font = header_font
            ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

        # BytesIOë¡œ ì €ì¥
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # ë°”ë‹¥íŒ ì„¸ëŒ€ìˆ˜ ì¶”ì¶œ
    total_units = 1  # ê¸°ë³¸ê°’
    if floor_data:
        # floor_data êµ¬ì¡°: {"inputs": {"units": N}, ...}
        inputs = floor_data.get("inputs", {})
        total_units = int(inputs.get("units", 1))

    xlsx_bytes = df_to_excel_bytes(
        est_df,
        total_units,
        category_subtotals=category_subtotals,
        category_mgmt_fees=category_mgmt_fees,
        prod_mgmt_settings=prod_mgmt_categories,
        total_mgmt_fee=total_mgmt_fee,
        final_total=final_total,
    )
    st.download_button(
        "ğŸ“¥ í˜„ì¬ ì„¸ëŒ€ ê²¬ì ì„œ ë‹¤ìš´ë¡œë“œ (ìƒì‚°ê´€ë¦¬ë¹„ í¬í•¨)",
        data=xlsx_bytes,
        file_name=f"ìš•ì‹¤_ì›ìì¬_ì„¸ëŒ€ë‹¹_ë‹¨ê°€ë‚´ì—­_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ----------------------------
    # í†µí•© ì—‘ì…€ ì¶œë ¥ (ë‹¤ì¤‘ ì„¸ëŒ€ íƒ€ì…)
    # ----------------------------
    def create_integrated_excel(saved_quotations: List[Dict]) -> bytes:
        """LGE ì°½ì› ìŠ¤ë§ˆíŠ¸íŒŒí¬ í˜•ì‹ì˜ í†µí•© ì—‘ì…€ ìƒì„±"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "ì„¸ëŒ€ë‹¹ ì›ìì¬ ë‹¨ê°€ë‚´ì—­"

        # ìŠ¤íƒ€ì¼ ì •ì˜
        title_font = Font(name="ë§‘ì€ ê³ ë”•", size=16, bold=True)
        header_font = Font(name="ë§‘ì€ ê³ ë”•", size=9, bold=True)
        data_font = Font(name="ë§‘ì€ ê³ ë”•", size=9)
        small_font = Font(name="ë§‘ì€ ê³ ë”•", size=8)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        num_types = len(saved_quotations)
        if num_types == 0:
            return b""

        # ëª¨ë“  í’ˆëª©/ì‚¬ì–‘ ì¡°í•© ìˆ˜ì§‘ (ìˆœì„œ ìœ ì§€)
        all_items = []
        seen = set()
        for q in saved_quotations:
            for row in q["rows"]:
                key = (row["í’ˆëª©"], row["ì‚¬ì–‘ ë° ê·œê²©"])
                if key not in seen:
                    seen.add(key)
                    all_items.append(key)

        # ê°™ì€ ëŒ€ë¶„ë¥˜(í’ˆëª©)ë¼ë¦¬ ì—°ì† ë°°ì¹˜ë˜ë„ë¡ ì •ë ¬
        category_order = {}
        for i, (í’ˆëª©, ì‚¬ì–‘) in enumerate(all_items):
            if í’ˆëª© not in category_order:
                category_order[í’ˆëª©] = i
        all_items = sorted(all_items, key=lambda x: category_order[x[0]])

        # ì»¬ëŸ¼ êµ¬ì¡° ê³„ì‚°
        # í’ˆëª©(1) + ì‚¬ì–‘(1) + [ìˆ˜ëŸ‰,ë‹¨ê°€,ê¸ˆì•¡] Ã— num_types + ë¹„ê³ (1)
        START_COL = 1
        SPEC_COL = 2
        DATA_START_COL = 3  # ì²« ë²ˆì§¸ ì„¸ëŒ€ íƒ€ì…ì˜ ìˆ˜ëŸ‰ ì»¬ëŸ¼

        # 1í–‰: íƒ€ì´í‹€
        ws.merge_cells(
            start_row=1, start_column=1, end_row=1, end_column=2 + num_types * 3 + 1
        )
        ws.cell(1, 1).value = "ìš•ì‹¤ ì›ìì¬ ì„¸ëŒ€ë‹¹ ë‹¨ê°€ ë‚´ì—­"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = center_align
        ws.row_dimensions[1].height = 25

        # 4í–‰: ì´ìˆ˜ëŸ‰ ë° ì‘ì„±ì¼
        total_all_units = sum(q["units"] for q in saved_quotations)
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        ws.cell(4, 1).font = header_font
        ws.merge_cells(
            start_row=4,
            start_column=DATA_START_COL + num_types * 3 - 2,
            end_row=4,
            end_column=DATA_START_COL + num_types * 3,
        )
        date_col = DATA_START_COL + num_types * 3 - 2
        ws.cell(4, date_col).value = f"ì´ìˆ˜ëŸ‰: {total_all_units}ê°œ"
        ws.cell(4, date_col).font = header_font
        ws.cell(4, date_col).alignment = right_align

        # 5í–‰: ì„¸ëŒ€ íƒ€ì… í—¤ë” (â— íƒ€ì…ëª… â— í˜•íƒœ)
        ws.cell(5, START_COL).value = "í’ˆëª©"
        ws.cell(5, START_COL).font = header_font
        ws.cell(5, START_COL).alignment = center_align
        ws.cell(5, START_COL).border = thin_border

        ws.cell(5, SPEC_COL).value = "ì‚¬ì–‘ ë° ê·œê²©"
        ws.cell(5, SPEC_COL).font = header_font
        ws.cell(5, SPEC_COL).alignment = center_align
        ws.cell(5, SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            # 3ì»¬ëŸ¼ ë³‘í•©
            ws.merge_cells(
                start_row=5, start_column=col_start, end_row=5, end_column=col_start + 2
            )
            ws.cell(5, col_start).value = f"â— {q['name']}"
            ws.cell(5, col_start).font = header_font
            ws.cell(5, col_start).alignment = center_align
            for c in range(col_start, col_start + 3):
                ws.cell(5, c).border = thin_border

        # ë¹„ê³  ì»¬ëŸ¼
        remark_col = DATA_START_COL + num_types * 3
        ws.cell(5, remark_col).value = "(V.A.T ì œì™¸)"
        ws.cell(5, remark_col).font = small_font
        ws.cell(5, remark_col).alignment = center_align
        ws.cell(5, remark_col).border = thin_border

        # 6í–‰: ê·œê²© ë° ì„¸ëŒ€ìˆ˜
        ws.cell(6, START_COL).value = ""
        ws.cell(6, START_COL).border = thin_border
        ws.cell(6, SPEC_COL).value = ""
        ws.cell(6, SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            ws.merge_cells(
                start_row=6, start_column=col_start, end_row=6, end_column=col_start + 2
            )
            ws.cell(6, col_start).value = f"â— ê·œê²©({q['spec']})  â— {q['units']}ì„¸ëŒ€"
            ws.cell(6, col_start).font = small_font
            ws.cell(6, col_start).alignment = center_align
            for c in range(col_start, col_start + 3):
                ws.cell(6, c).border = thin_border

        ws.cell(6, remark_col).value = "ë¹„ê³ "
        ws.cell(6, remark_col).font = header_font
        ws.cell(6, remark_col).alignment = center_align
        ws.cell(6, remark_col).border = thin_border

        # 7í–‰: ìˆ˜ëŸ‰/ë‹¨ê°€/ê¸ˆì•¡ í—¤ë”
        ws.cell(7, START_COL).value = ""
        ws.cell(7, START_COL).border = thin_border
        ws.cell(7, SPEC_COL).value = ""
        ws.cell(7, SPEC_COL).border = thin_border

        for i in range(num_types):
            col_start = DATA_START_COL + i * 3
            ws.cell(7, col_start).value = "ìˆ˜ëŸ‰"
            ws.cell(7, col_start).font = header_font
            ws.cell(7, col_start).alignment = center_align
            ws.cell(7, col_start).border = thin_border

            ws.cell(7, col_start + 1).value = "ë‹¨ê°€"
            ws.cell(7, col_start + 1).font = header_font
            ws.cell(7, col_start + 1).alignment = center_align
            ws.cell(7, col_start + 1).border = thin_border

            ws.cell(7, col_start + 2).value = "ê¸ˆì•¡"
            ws.cell(7, col_start + 2).font = header_font
            ws.cell(7, col_start + 2).alignment = center_align
            ws.cell(7, col_start + 2).border = thin_border

        ws.cell(7, remark_col).value = ""
        ws.cell(7, remark_col).border = thin_border

        # ë°ì´í„° í–‰ ì‘ì„±
        row_num = 8
        current_category = None

        # ê° ì„¸ëŒ€ë³„ ë°ì´í„°ë¥¼ ë”•ì…”ë„ˆë¦¬ë¡œ ë³€í™˜ (ë¹ ë¥¸ ì¡°íšŒìš©)
        type_data = []
        for q in saved_quotations:
            item_dict = {}
            for r in q["rows"]:
                key = (r["í’ˆëª©"], r["ì‚¬ì–‘ ë° ê·œê²©"])
                item_dict[key] = r
            type_data.append(item_dict)

        # ê° ëŒ€ë¶„ë¥˜ë³„ ì‹œì‘/ë í–‰ ê³„ì‚°
        category_rows = {}
        temp_row = 8
        for í’ˆëª©, ì‚¬ì–‘ in all_items:
            if í’ˆëª© not in category_rows:
                category_rows[í’ˆëª©] = {"start": temp_row, "end": temp_row}
            else:
                category_rows[í’ˆëª©]["end"] = temp_row
            temp_row += 1

        # ëŒ€ë¶„ë¥˜ ì—´ í…Œë‘ë¦¬ ìŠ¤íƒ€ì¼ ì •ì˜
        top_only_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style=None),
        )
        middle_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style=None),
            bottom=Side(style=None),
        )
        bottom_only_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style=None),
            bottom=Side(style="thin"),
        )

        for í’ˆëª©, ì‚¬ì–‘ in all_items:
            # í’ˆëª© (ì¹´í…Œê³ ë¦¬ ë³€ê²½ì‹œë§Œ í‘œì‹œ)
            cell_cat = ws.cell(row=row_num, column=START_COL)
            if í’ˆëª© != current_category:
                cell_cat.value = í’ˆëª©
                current_category = í’ˆëª©
            else:
                cell_cat.value = ""
            cell_cat.font = data_font
            cell_cat.alignment = left_align

            # ëŒ€ë¶„ë¥˜ ì—´ í…Œë‘ë¦¬ ì„¤ì •
            cat_info = category_rows.get(í’ˆëª©, {})
            start_row = cat_info.get("start", row_num)
            end_row = cat_info.get("end", row_num)

            if start_row == end_row:
                # ë‹¨ì¼ í–‰ì´ë©´ ì „ì²´ í…Œë‘ë¦¬
                cell_cat.border = thin_border
            elif row_num == start_row:
                # ì²« í–‰: ìœ„ í…Œë‘ë¦¬ë§Œ
                cell_cat.border = top_only_border
            elif row_num == end_row:
                # ë§ˆì§€ë§‰ í–‰: ì•„ë˜ í…Œë‘ë¦¬ë§Œ
                cell_cat.border = bottom_only_border
            else:
                # ì¤‘ê°„ í–‰: ì¢Œìš°ë§Œ
                cell_cat.border = middle_border

            # ì‚¬ì–‘ ë° ê·œê²©
            ws.cell(row=row_num, column=SPEC_COL).value = ì‚¬ì–‘
            ws.cell(row=row_num, column=SPEC_COL).font = data_font
            ws.cell(row=row_num, column=SPEC_COL).alignment = left_align
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            # ê° ì„¸ëŒ€ íƒ€ì…ë³„ ìˆ˜ëŸ‰/ë‹¨ê°€/ê¸ˆì•¡
            for i, td in enumerate(type_data):
                col_start = DATA_START_COL + i * 3
                key = (í’ˆëª©, ì‚¬ì–‘)
                if key in td:
                    r = td[key]
                    qty = r.get("ìˆ˜ëŸ‰", 0) or 0
                    price = r.get("ë‹¨ê°€", 0) or 0
                    amount = r.get("ê¸ˆì•¡", 0) or 0
                else:
                    qty, price, amount = 0, 0, 0

                ws.cell(row=row_num, column=col_start).value = qty if qty else 0
                ws.cell(row=row_num, column=col_start).font = data_font
                ws.cell(row=row_num, column=col_start).alignment = right_align
                ws.cell(row=row_num, column=col_start).border = thin_border
                ws.cell(row=row_num, column=col_start).number_format = "#,##0.##"

                ws.cell(row=row_num, column=col_start + 1).value = price if price else 0
                ws.cell(row=row_num, column=col_start + 1).font = data_font
                ws.cell(row=row_num, column=col_start + 1).alignment = right_align
                ws.cell(row=row_num, column=col_start + 1).border = thin_border
                ws.cell(row=row_num, column=col_start + 1).number_format = "#,##0"

                ws.cell(row=row_num, column=col_start + 2).value = (
                    amount if amount else 0
                )
                ws.cell(row=row_num, column=col_start + 2).font = data_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            # ë¹„ê³ 
            ws.cell(row=row_num, column=remark_col).value = ""
            ws.cell(row=row_num, column=remark_col).border = thin_border

            row_num += 1

        # í•©ê³„ í–‰: ì„¸íŠ¸ë‹¹ ë‹¨ê°€
        ws.cell(row=row_num, column=START_COL).value = "ì„¸íŠ¸ë‹¹ ë‹¨ê°€"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border
        ws.cell(row=row_num, column=SPEC_COL).value = ""
        ws.cell(row=row_num, column=SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            ws.cell(row=row_num, column=col_start).value = 1
            ws.cell(row=row_num, column=col_start).font = header_font
            ws.cell(row=row_num, column=col_start).alignment = right_align
            ws.cell(row=row_num, column=col_start).border = thin_border

            ws.cell(row=row_num, column=col_start + 1).value = ""
            ws.cell(row=row_num, column=col_start + 1).border = thin_border

            ws.cell(row=row_num, column=col_start + 2).value = q["total"]
            ws.cell(row=row_num, column=col_start + 2).font = header_font
            ws.cell(row=row_num, column=col_start + 2).alignment = right_align
            ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

        ws.cell(row=row_num, column=remark_col).value = ""
        ws.cell(row=row_num, column=remark_col).border = thin_border
        row_num += 1

        # ì„¸ëŒ€ ì´ í•©ê³„ í–‰
        ws.cell(row=row_num, column=START_COL).value = "ì„¸ëŒ€ ì´ í•©ê³„"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border
        ws.cell(row=row_num, column=SPEC_COL).value = ""
        ws.cell(row=row_num, column=SPEC_COL).border = thin_border

        grand_total = 0
        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            type_total = q["total"] * q["units"]
            grand_total += type_total

            ws.cell(row=row_num, column=col_start).value = q["units"]
            ws.cell(row=row_num, column=col_start).font = header_font
            ws.cell(row=row_num, column=col_start).alignment = right_align
            ws.cell(row=row_num, column=col_start).border = thin_border

            ws.cell(row=row_num, column=col_start + 1).value = ""
            ws.cell(row=row_num, column=col_start + 1).border = thin_border

            ws.cell(row=row_num, column=col_start + 2).value = type_total
            ws.cell(row=row_num, column=col_start + 2).font = header_font
            ws.cell(row=row_num, column=col_start + 2).alignment = right_align
            ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

        # ì´ í•©ê³„ í‘œì‹œ
        ws.cell(row=row_num, column=remark_col).value = f"{grand_total:,.0f}"
        ws.cell(row=row_num, column=remark_col).font = header_font
        ws.cell(row=row_num, column=remark_col).alignment = right_align
        ws.cell(row=row_num, column=remark_col).border = thin_border
        row_num += 1

        # ----------------------------
        # ê³µì‚¬ ì›ê°€ ì¡°ì • ë° ê´€ë¦¬ë¹„ ì„¹ì…˜ (ì´ë¯¸ì§€ í˜•ì‹)
        # ----------------------------
        # ìƒì‚°ê´€ë¦¬ë¹„ ì •ë³´ê°€ ìˆëŠ”ì§€ í™•ì¸
        has_mgmt_fee = any(
            q.get("total_mgmt_fee", 0) > 0 or q.get("prod_mgmt_settings")
            for q in saved_quotations
        )

        if has_mgmt_fee:
            row_num += 2

            # ì„¹ì…˜ íƒ€ì´í‹€: â— ê³µì‚¬ ì›ê°€ ì¡°ì • ë° ê´€ë¦¬ë¹„
            ws.merge_cells(
                start_row=row_num,
                start_column=START_COL,
                end_row=row_num,
                end_column=remark_col,
            )
            ws.cell(row=row_num, column=START_COL).value = "â— ê³µì‚¬ ì›ê°€ ì¡°ì • ë° ê´€ë¦¬ë¹„"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = left_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            row_num += 1

            # ì¹´í…Œê³ ë¦¬ ê·¸ë£¹ ì •ì˜ (ì´ë¯¸ì§€ì— ë§ì¶˜ êµ¬ì¡°)
            # ê·¸ë£¹: (ê·¸ë£¹ëª…, [(ì¹´í…Œê³ ë¦¬í‚¤, í‘œì‹œëª…), ...])
            mgmt_groups = [
                (
                    "íšŒì‚¬ ìƒì‚°ê´€ë¦¬ë¹„",
                    [
                        ("íšŒì‚¬ìƒì‚°í’ˆ(ë°”ë‹¥íŒ,ìš•ì¡°)", "ë°”ë‹¥íŒ, ìš•ì¡°(20~25%)"),
                        ("íšŒì‚¬ìƒì‚°í’ˆ(ì²œì¥íŒ)", "ì²œì¥íŒ(15~20%)"),
                    ],
                ),
                (
                    "ëª…ì§„ ìƒì‚°ê´€ë¦¬ë¹„",
                    [
                        ("íšŒì‚¬-ëª…ì§„(ë²½,PVEë°”ë‹¥íŒ)", "PVEë°”ë‹¥íŒ, íƒ€ì¼ë²½ì²´(15~20%)"),
                    ],
                ),
                (
                    "íƒ€ì‚¬ êµ¬ë§¤í’ˆ",
                    [
                        ("íƒ€ì‚¬(ì²œì¥,ë°”ë‹¥íŒ,íƒ€ì¼)", "ë°”ë‹¥íŒ, íƒ€ì¼(5~10%)"),
                        ("íƒ€ì‚¬(ë„ê¸°,ìˆ˜ì „,ê¸°íƒ€)", "ë„ê¸°, ìˆ˜ì „ë¥˜, ê¸°íƒ€(5~10%)"),
                    ],
                ),
            ]

            # ì¹´í…Œê³ ë¦¬ë³„ ë°ì´í„°ë¥¼ first quotation ê¸°ì¤€ìœ¼ë¡œ ìˆ˜ì§‘
            def get_cat_data(q, cat_key):
                """ì£¼ì–´ì§„ ê²¬ì ì—ì„œ ì¹´í…Œê³ ë¦¬ ë°ì´í„° ì¶”ì¶œ"""
                cat_subtotals = q.get("category_subtotals", {})
                cat_mgmt_fees = q.get("category_mgmt_fees", {})
                prod_settings = q.get("prod_mgmt_settings", {})

                subtotal = cat_subtotals.get(cat_key, 0)
                mgmt_fee = cat_mgmt_fees.get(cat_key, 0)
                rate = 0
                if isinstance(prod_settings.get(cat_key), dict):
                    rate = prod_settings[cat_key].get("rate", 0)
                return subtotal, mgmt_fee, rate

            # ê·¸ë£¹ë³„ë¡œ ë Œë”ë§
            for group_name, categories in mgmt_groups:
                num_rows = len(categories)
                start_row_for_group = row_num

                for idx, (cat_key, display_name) in enumerate(categories):
                    # ì²« ë²ˆì§¸ í–‰ì—ë§Œ ê·¸ë£¹ëª… í‘œì‹œ (ì„¸ë¡œ ë³‘í•©)
                    if idx == 0:
                        if num_rows > 1:
                            ws.merge_cells(
                                start_row=start_row_for_group,
                                start_column=START_COL,
                                end_row=start_row_for_group + num_rows - 1,
                                end_column=START_COL,
                            )
                        ws.cell(row=start_row_for_group, column=START_COL).value = (
                            group_name
                        )
                        ws.cell(row=start_row_for_group, column=START_COL).font = (
                            data_font
                        )
                        ws.cell(row=start_row_for_group, column=START_COL).alignment = (
                            center_align
                        )
                        # ë³‘í•©ëœ ì…€ì˜ í…Œë‘ë¦¬ ì„¤ì •
                        for r in range(
                            start_row_for_group, start_row_for_group + num_rows
                        ):
                            ws.cell(row=r, column=START_COL).border = thin_border

                    # ì¹´í…Œê³ ë¦¬ í‘œì‹œëª…
                    ws.cell(row=row_num, column=SPEC_COL).value = display_name
                    ws.cell(row=row_num, column=SPEC_COL).font = data_font
                    ws.cell(row=row_num, column=SPEC_COL).alignment = left_align
                    ws.cell(row=row_num, column=SPEC_COL).border = thin_border

                    # ê° ì„¸ëŒ€ íƒ€ì…ë³„ ë°ì´í„°
                    for i, q in enumerate(saved_quotations):
                        col_start = DATA_START_COL + i * 3
                        subtotal, mgmt_fee, rate = get_cat_data(q, cat_key)

                        # ë¹„ìœ¨(%)
                        ws.cell(row=row_num, column=col_start).value = f"{rate}%"
                        ws.cell(row=row_num, column=col_start).font = small_font
                        ws.cell(row=row_num, column=col_start).alignment = right_align
                        ws.cell(row=row_num, column=col_start).border = thin_border

                        # ë¹ˆ ì¹¸ (ì¤‘ê°„)
                        ws.cell(row=row_num, column=col_start + 1).value = ""
                        ws.cell(row=row_num, column=col_start + 1).border = thin_border

                        # ìƒì‚°ê´€ë¦¬ë¹„
                        ws.cell(row=row_num, column=col_start + 2).value = (
                            mgmt_fee if mgmt_fee else ""
                        )
                        ws.cell(row=row_num, column=col_start + 2).font = data_font
                        ws.cell(row=row_num, column=col_start + 2).alignment = (
                            right_align
                        )
                        ws.cell(row=row_num, column=col_start + 2).border = thin_border
                        if mgmt_fee:
                            ws.cell(row=row_num, column=col_start + 2).number_format = (
                                "#,##0"
                            )

                    # ë¹„ê³  (ê° í–‰ë³„ í•©ê³„ í‘œì‹œ)
                    row_total = sum(
                        get_cat_data(q, cat_key)[1] for q in saved_quotations
                    )
                    ws.cell(row=row_num, column=remark_col).value = (
                        row_total if row_total else ""
                    )
                    ws.cell(row=row_num, column=remark_col).font = data_font
                    ws.cell(row=row_num, column=remark_col).alignment = right_align
                    ws.cell(row=row_num, column=remark_col).border = thin_border
                    if row_total:
                        ws.cell(row=row_num, column=remark_col).number_format = "#,##0"

                    row_num += 1

            # ì„¤ì¹˜(AS) ê´€ë¦¬ë¹„(15% ê³ ì •) í–‰ - ì˜ì—…ê´€ë¦¬ë¹„ê°€ ìˆëŠ” ê²½ìš°
            has_sales_fee = any(
                q.get("sales_mgmt_enabled", False) for q in saved_quotations
            )
            if has_sales_fee:
                ws.cell(row=row_num, column=START_COL).value = (
                    "ì„¤ì¹˜(AS) ê´€ë¦¬ë¹„(15% ê³ ì •)"
                )
                ws.cell(row=row_num, column=START_COL).font = data_font
                ws.cell(row=row_num, column=START_COL).alignment = left_align
                ws.cell(row=row_num, column=START_COL).border = thin_border
                ws.cell(row=row_num, column=SPEC_COL).value = ""
                ws.cell(row=row_num, column=SPEC_COL).border = thin_border

                for i, q in enumerate(saved_quotations):
                    col_start = DATA_START_COL + i * 3
                    sales_rate = q.get("sales_mgmt_rate", 0)
                    sales_fee = (
                        q.get("sales_mgmt_fee", 0)
                        if q.get("sales_mgmt_enabled", False)
                        else 0
                    )

                    ws.cell(row=row_num, column=col_start).value = (
                        f"{sales_rate}%" if sales_fee else ""
                    )
                    ws.cell(row=row_num, column=col_start).font = small_font
                    ws.cell(row=row_num, column=col_start).alignment = right_align
                    ws.cell(row=row_num, column=col_start).border = thin_border

                    ws.cell(row=row_num, column=col_start + 1).value = ""
                    ws.cell(row=row_num, column=col_start + 1).border = thin_border

                    ws.cell(row=row_num, column=col_start + 2).value = (
                        sales_fee if sales_fee else ""
                    )
                    ws.cell(row=row_num, column=col_start + 2).font = data_font
                    ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                    ws.cell(row=row_num, column=col_start + 2).border = thin_border
                    if sales_fee:
                        ws.cell(row=row_num, column=col_start + 2).number_format = (
                            "#,##0"
                        )

                total_sales = sum(
                    q.get("sales_mgmt_fee", 0)
                    for q in saved_quotations
                    if q.get("sales_mgmt_enabled", False)
                )
                ws.cell(row=row_num, column=remark_col).value = (
                    total_sales if total_sales else ""
                )
                ws.cell(row=row_num, column=remark_col).font = data_font
                ws.cell(row=row_num, column=remark_col).alignment = right_align
                ws.cell(row=row_num, column=remark_col).border = thin_border
                if total_sales:
                    ws.cell(row=row_num, column=remark_col).number_format = "#,##0"
                row_num += 1

            # ë§¤ì…ì„¸ ì°¨ì´ì•¡(0~3%) í–‰ - ë¹ˆ í–‰
            ws.cell(row=row_num, column=START_COL).value = "ë§¤ì…ì„¸ ì°¨ì´ì•¡(0~3%)"
            ws.cell(row=row_num, column=START_COL).font = data_font
            ws.cell(row=row_num, column=START_COL).alignment = left_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border
            for i in range(num_types):
                col_start = DATA_START_COL + i * 3
                ws.cell(row=row_num, column=col_start).value = "0%"
                ws.cell(row=row_num, column=col_start).font = small_font
                ws.cell(row=row_num, column=col_start).alignment = right_align
                ws.cell(row=row_num, column=col_start).border = thin_border
                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).value = ""
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=remark_col).value = ""
            ws.cell(row=row_num, column=remark_col).border = thin_border
            row_num += 1

            # ì†Œê³„ í–‰
            ws.cell(row=row_num, column=START_COL).value = "ì†Œê³„"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            total_all_mgmt = 0
            total_all_sales = 0
            for i, q in enumerate(saved_quotations):
                col_start = DATA_START_COL + i * 3
                mgmt_fee = q.get("total_mgmt_fee", 0)
                sales_fee = (
                    q.get("sales_mgmt_fee", 0)
                    if q.get("sales_mgmt_enabled", False)
                    else 0
                )
                subtotal = mgmt_fee + sales_fee
                total_all_mgmt += mgmt_fee
                total_all_sales += sales_fee

                ws.cell(row=row_num, column=col_start).value = ""
                ws.cell(row=row_num, column=col_start).border = thin_border

                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border

                ws.cell(row=row_num, column=col_start + 2).value = subtotal
                ws.cell(row=row_num, column=col_start + 2).font = header_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            ws.cell(row=row_num, column=remark_col).value = (
                total_all_mgmt + total_all_sales
            )
            ws.cell(row=row_num, column=remark_col).font = header_font
            ws.cell(row=row_num, column=remark_col).alignment = right_align
            ws.cell(row=row_num, column=remark_col).border = thin_border
            ws.cell(row=row_num, column=remark_col).number_format = "#,##0"
            row_num += 1

            # ì´ ê¸ˆì•¡ í•©ê³„ í–‰
            row_num += 1
            ws.cell(row=row_num, column=START_COL).value = "ì´ ê¸ˆì•¡ í•©ê³„"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            final_grand_total = 0
            for i, q in enumerate(saved_quotations):
                col_start = DATA_START_COL + i * 3
                final_per_unit = q.get("final_total", q["total"])
                type_final_total = final_per_unit * q["units"]
                final_grand_total += type_final_total

                ws.cell(row=row_num, column=col_start).value = ""
                ws.cell(row=row_num, column=col_start).border = thin_border

                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border

                ws.cell(row=row_num, column=col_start + 2).value = final_per_unit
                ws.cell(row=row_num, column=col_start + 2).font = header_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            # ë…¹ìƒ‰ ë°°ê²½ì˜ ìµœì¢… í•©ê³„
            green_fill = PatternFill(
                start_color="90EE90", end_color="90EE90", fill_type="solid"
            )
            ws.cell(row=row_num, column=remark_col).value = final_grand_total
            ws.cell(row=row_num, column=remark_col).font = header_font
            ws.cell(row=row_num, column=remark_col).alignment = right_align
            ws.cell(row=row_num, column=remark_col).border = thin_border
            ws.cell(row=row_num, column=remark_col).fill = green_fill
            ws.cell(row=row_num, column=remark_col).number_format = "#,##0"

        # ì»¬ëŸ¼ ë„ˆë¹„ ì„¤ì •
        ws.column_dimensions[get_column_letter(START_COL)].width = 12
        ws.column_dimensions[get_column_letter(SPEC_COL)].width = 30
        for i in range(num_types):
            col_start = DATA_START_COL + i * 3
            ws.column_dimensions[get_column_letter(col_start)].width = 7  # ìˆ˜ëŸ‰
            ws.column_dimensions[get_column_letter(col_start + 1)].width = 10  # ë‹¨ê°€
            ws.column_dimensions[get_column_letter(col_start + 2)].width = 12  # ê¸ˆì•¡
        ws.column_dimensions[get_column_letter(remark_col)].width = 15

        # BytesIOë¡œ ì €ì¥
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # í†µí•© ì—‘ì…€ ë‹¤ìš´ë¡œë“œ ë²„íŠ¼
    saved_list = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
    if saved_list and len(saved_list) >= 1:
        st.markdown("### í†µí•© ê²¬ì ì„œ ë‹¤ìš´ë¡œë“œ")
        integrated_bytes = create_integrated_excel(saved_list)
        if integrated_bytes:
            st.download_button(
                "ğŸ“¥ í†µí•© ê²¬ì ì„œ Excel ë‹¤ìš´ë¡œë“œ (LGE í˜•ì‹)",
                data=integrated_bytes,
                file_name=f"ìš•ì‹¤_ì›ìì¬_í†µí•©_ë‹¨ê°€ë‚´ì—­_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
            st.info(f"ì´ {len(saved_list)}ê°œ ì„¸ëŒ€ íƒ€ì… í¬í•¨")

if warnings:
    with st.expander("âš ï¸ ê²½ê³ /ì°¸ê³ ", expanded=False):
        for w in warnings:
            st.warning(w)
