# 욕실 견적서 생성기
# session_state 연동 버전 - 바닥/벽/천장 계산 결과를 자동으로 가져옵니다.

from common_styles import apply_common_styles, set_page_config
import auth

import json
import io
from typing import Dict, Any, List, Optional
from datetime import datetime
import pandas as pd
import streamlit as st

# Session state keys
FLOOR_RESULT_KEY = "floor_result"
WALL_RESULT_KEY = "wall_result"
CEIL_RESULT_KEY = "ceil_result"
SAVED_QUOTATIONS_KEY = "saved_quotations"  # 저장된 세대 타입별 견적 목록 (최대 10개)
PROD_MGMT_SETTINGS_KEY = "prod_mgmt_settings"  # 생산관리비 설정

# 생산관리비 기본 카테고리 정의 (품목+사양 단위로 세부 지정)
# items: [(품목, 사양패턴), ...] - 사양패턴이 None이면 해당 품목 전체, 문자열이면 contains 매칭
DEFAULT_PROD_MGMT_CATEGORIES = {
    "회사생산품(바닥판,욕조)": {
        "items": [
            ("바닥판", "GRP"),
            ("바닥판", "FRP"),  # FRP 포함 (SMC/FRP 등)
            ("바닥판", "SMC"),  # SMC 포함
            ("바닥판", "PP"),   # PP/PE 포함
            ("욕조", None),  # 욕조 전체
        ],
        "rate": 20.0,  # 기본값 20%
    },
    "회사생산품(천장판)": {
        "items": [
            ("천장판", None),  # 천장판 전체
        ],
        "rate": 15.0,  # 기본값 15%
    },
    "회사-명진(벽,PVE바닥판)": {
        "items": [
            ("벽판", None),  # 벽판 전체
            ("바닥판", "PVE"),  # PVE 바닥판만
        ],
        "rate": 15.0,  # 기본값 15%
    },
    "타사(천장,바닥판,타일)": {
        "items": [
            ("타일", None),  # 타일 전체
        ],
        "rate": 5.0,  # 기본값 5%
    },
    "타사(도기,수전,기타)": {
        "items": [
            ("도기류", None),
            ("수전", None),
            ("액세서리", None),
            ("문세트", None),
            ("욕실등", None),
            ("공통자재", None),
            ("냉온수배관", None),
            ("문틀규격", None),
            ("은경", None),
            ("욕실장", None),
            ("칸막이", None),
            ("환기류", None),
        ],
        "rate": 5.0,  # 기본값 5%
    },
}

# 영업관리비 설정 키
SALES_MGMT_SETTINGS_KEY = "sales_mgmt_settings"


def get_item_key(품목: str, 사양: str) -> str:
    """품목+사양을 고유 키로 변환"""
    return f"{품목}::{사양}"


def parse_item_key(key: str) -> tuple:
    """고유 키를 품목, 사양으로 분리"""
    parts = key.split("::", 1)
    return (parts[0], parts[1]) if len(parts) == 2 else (parts[0], "")


def item_matches_pattern(품목: str, 사양: str, pattern_품목: str, pattern_사양: Optional[str]) -> bool:
    """품목+사양이 패턴과 매칭되는지 확인"""
    if 품목 != pattern_품목:
        return False
    if pattern_사양 is None:
        return True  # 품목만 매칭하면 전체 포함
    return pattern_사양.upper() in 사양.upper()

set_page_config(page_title="욕실 견적서 생성기", layout="wide")
apply_common_styles()

auth.require_auth()

# ----------------------------
# Helper Functions
# ----------------------------
REQ_COLUMNS = ["품목", "분류", "사양 및 규격", "단가", "수량"]


@st.cache_data(show_spinner=False)
def load_pricebook_from_excel(
    file_bytes: bytes, sheet_name: str = "자재단가내역"
) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
    # Normalize columns
    colmap = {}
    for c in df.columns:
        c2 = str(c).strip()
        if c2 in ["품목", "폼목"]:
            colmap[c] = "품목"
        elif c2 in ["분류"]:
            colmap[c] = "분류"
        elif c2 in ["사양 및 규격", "사양", "규격"]:
            colmap[c] = "사양 및 규격"
        elif c2 in ["단가"]:
            colmap[c] = "단가"
        elif c2 in ["수량"]:
            colmap[c] = "수량"
        elif c2 in ["금액"]:
            colmap[c] = "금액"
    df = df.rename(columns=colmap)
    # Ensure required columns exist
    for c in ["품목", "분류", "사양 및 규격", "단가", "수량"]:
        if c not in df.columns:
            df[c] = None
    # Clean values
    for c in ["품목", "분류", "사양 및 규격"]:
        df[c] = df[c].astype(str).str.strip()
    for c in ["단가", "수량"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    if "금액" not in df.columns:
        df["금액"] = df["단가"].fillna(0) * df["수량"].fillna(0)
    return df


def find_item(
    df: pd.DataFrame,
    품목: str,
    분류: Optional[str] = None,
    spec_contains: Optional[str] = None,
) -> Optional[pd.Series]:
    q = df["품목"] == 품목
    if 분류 is not None:
        q &= df["분류"] == 분류
    if spec_contains:
        q &= df["사양 및 규격"].str.contains(str(spec_contains), case=False, na=False)
    candidates = df[q]
    if len(candidates) == 0:
        return None
    # If multiple, prefer exact spec match first
    if spec_contains:
        exact = candidates[candidates["사양 및 규격"].str.strip() == spec_contains]
        if len(exact) == 1:
            return exact.iloc[0]
    return candidates.iloc[0]


def add_row(
    rows: List[Dict[str, Any]],
    품목: str,
    spec: str,
    qty: float,
    unit_price: Optional[float],
) -> None:
    unit_price = unit_price if unit_price is not None else 0
    amount = (qty or 0) * (unit_price or 0)
    rows.append(
        {
            "품목": 품목,
            "사양 및 규격": spec,
            "수량": qty,
            "단가": unit_price,
            "금액": amount,
        }
    )


def add_all_by_category(
    rows: List[Dict[str, Any]], df: pd.DataFrame, 품목: str, 분류: str
):
    sub = df[(df["품목"] == 품목) & (df["분류"] == 분류)]
    for _, r in sub.iterrows():
        add_row(
            rows,
            품목,
            str(r["사양 및 규격"]),
            r["수량"] if pd.notna(r["수량"]) else 1,
            r["단가"] if pd.notna(r["단가"]) else 0,
        )


# ----------------------------
# Convert session_state to quotation format
# ----------------------------
def convert_floor_data(floor_result: dict) -> dict:
    """Convert floor_result to quotation format"""
    if not floor_result:
        return {}

    # session_state 구조: {"section", "inputs", "result", "decision_log"}
    inputs = floor_result.get("inputs", {})
    result = floor_result.get("result", {})

    # 소재 정보 추출 (result에서)
    material = result.get("소재", "")
    # "PP/PE 바닥판" -> "PP/PE" 변환
    material_clean = material.replace(" 바닥판", "").replace("바닥판", "").strip()

    # 가격 정보 추출 (result에서) - 소계 사용
    단가 = result.get("소계", 0)

    # 세대수 정보 (inputs에서)
    units = inputs.get("units", 1)

    # 규격 문자열 생성
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)
    spec = f"{W}×{L}" if W and L else ""

    return {
        "재질": material_clean,
        "규격": spec,
        "수량": units,  # 세대수 = 수량
        "단가": 단가,
        "주거약자": inputs.get("user_type", "") == "주거약자",
        "inputs": inputs,  # inputs 정보 유지 (세대수 등)
    }


def convert_wall_data(wall_result: dict) -> dict:
    """Convert wall_result to quotation format"""
    if not wall_result:
        return {}

    result = wall_result.get("result", {})
    counts = result.get("counts", {})
    inputs = wall_result.get("inputs", {})

    # 소계 사용 (관리비 제외)
    unit_price = result.get("소계", 0)

    return {
        "총개수": counts.get("n_panels", 0),
        "단가": unit_price,
        "벽타일": inputs.get("tile", "300×600"),
    }


def convert_ceiling_data(ceil_result: dict) -> dict:
    """Convert ceil_result to quotation format"""
    if not ceil_result:
        return {}

    # ceil_panel_final.py의 session_state 구조에 맞춰 파싱
    inputs = ceil_result.get("inputs", {})
    result = ceil_result.get("result", {})

    # 재질 정보 추출 (inputs에서)
    material = inputs.get("material", "GRP")  # GRP/FRP/기타

    # 소계 사용 (관리비 제외)
    subtotal = result.get("소계", 0)

    # JSON export 데이터 사용 (이미 변환된 포맷)
    json_export = result.get("json_export", {})
    if json_export:
        # 점검구가 딕셔너리 형태인 경우 개수만 추출
        jgm = json_export.get("점검구", 1)
        hole_count = jgm.get("개수", 1) if isinstance(jgm, dict) else jgm

        return {
            "재질": json_export.get("재질", material),
            "총개수": json_export.get("총개수", 0),
            "바디판넬": json_export.get("바디판넬", {}),
            "사이드판넬": json_export.get("사이드판넬", {}),
            "천공구": hole_count,
            "단가": subtotal or json_export.get("소계", 0),
        }

    # Fallback: summary 데이터에서 추출
    summary = result.get("summary", {})
    elements = result.get("elements", [])

    # 바디/사이드 개수 카운트
    body_cnt = sum(1 for e in elements if e.get("kind") == "BODY")
    side_cnt = sum(1 for e in elements if e.get("kind") == "SIDE")

    # 대표 모델명 추출
    body_models = [e.get("model", "") for e in elements if e.get("kind") == "BODY"]
    side_models = [e.get("model", "") for e in elements if e.get("kind") == "SIDE"]

    body_info = {}
    if body_models:
        # 가장 많이 나온 모델
        from collections import Counter
        body_top = Counter(body_models).most_common(1)
        if body_top:
            body_info = {"종류": body_top[0][0].replace("(rot)", ""), "개수": body_cnt}

    side_info = {}
    if side_models:
        from collections import Counter
        side_top = Counter(side_models).most_common(1)
        if side_top:
            side_info = {"종류": side_top[0][0].replace("(rot)", ""), "개수": side_cnt}

    total_cnt = summary.get("총판넬수", body_cnt + side_cnt)

    return {
        "재질": material,
        "총개수": int(total_cnt),
        "바디판넬": body_info,
        "사이드판넬": side_info,
        "천공구": 1,  # 기본값, json_export 없으면 1로 가정
        "단가": int(subtotal),
    }


# ----------------------------
# UI
# ----------------------------
st.title("🛁 욕실 견적서 생성기")

# Check for calculation results
floor_result = st.session_state.get(FLOOR_RESULT_KEY)
wall_result = st.session_state.get(WALL_RESULT_KEY)
ceil_result = st.session_state.get(CEIL_RESULT_KEY)

has_floor = bool(floor_result)
has_wall = bool(wall_result)
has_ceil = bool(ceil_result)

# Status display
st.markdown("### 계산 결과 상태")
col1, col2, col3, col4 = st.columns(4)
with col1:
    status = "✅ 완료" if has_floor else "❌ 미완료"
    st.metric("바닥판", status)
with col2:
    status = "✅ 완료" if has_wall else "❌ 미완료"
    st.metric("벽판", status)
with col3:
    status = "✅ 완료" if has_ceil else "❌ 미완료"
    st.metric("천장판", status)
with col4:
    # 바닥판 세대수 표시
    units_display = 1
    if floor_result:
        inputs = floor_result.get("inputs", {})
        units_display = int(inputs.get("units", 1))
    st.metric("공사 세대수", f"{units_display}세대")

# ========== 바닥판, 벽판, 천장판 계산 의존성 체크 ==========
missing_steps = []
if not has_floor:
    missing_steps.append("🟦 바닥판 계산")
if not has_wall:
    missing_steps.append("🟩 벽판 계산")
if not has_ceil:
    missing_steps.append("🟨 천장판 계산")

if missing_steps:
    st.warning(
        f"⚠️ 견적서를 생성하려면 먼저 **{', '.join(missing_steps)}**을(를) 완료해야 합니다."
    )

    # 안내 카드
    st.markdown(
        """
    <div style="
        border: 1px solid #f59e0b;
        border-radius: 12px;
        padding: 20px;
        margin: 16px 0;
        background: linear-gradient(135deg, #fef3c7 0%, #fde68a 100%);
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
    ">
        <div style="display: flex; align-items: center; gap: 12px; margin-bottom: 12px;">
            <span style="font-size: 24px;">📋</span>
            <h3 style="margin: 0; color: #0f172a; font-weight: 700;">계산 순서 안내</h3>
        </div>
        <p style="margin: 0 0 12px 36px; color: #78350f; line-height: 1.6;">
            견적서 생성은 모든 계산이 완료된 후 진행할 수 있습니다:
        </p>
        <div style="margin-left: 36px; padding: 12px; background: white; border-radius: 8px; border: 1px solid #f59e0b;">
            <p style="margin: 0; color: #92400e; font-size: 0.95rem; line-height: 1.6;">
                <strong>1단계:</strong> 🟦 바닥판 계산"""
        + (" ← <em style='color:#dc2626;'>미완료</em>" if not has_floor else " ✅")
        + """<br>
                <strong>2단계:</strong> 🟩 벽판 계산"""
        + (" ← <em style='color:#dc2626;'>미완료</em>" if not has_wall else " ✅")
        + """<br>
                <strong>3단계:</strong> 🟨 천장판 계산"""
        + (" ← <em style='color:#dc2626;'>미완료</em>" if not has_ceil else " ✅")
        + """<br>
                <strong>4단계:</strong> 📋 견적서 생성 ← <em>현재 페이지</em>
            </p>
        </div>
    </div>
    """,
        unsafe_allow_html=True,
    )

    # 미완료 단계로 이동하는 버튼
    col_spacer, col_btn, col_spacer2 = st.columns([1, 2, 1])
    with col_btn:
        if not has_floor:
            st.page_link(
                "pages/1_바닥판_계산.py", label="🟦 바닥판 계산 시작하기", icon=None
            )
        elif not has_wall:
            st.page_link(
                "pages/2_벽판_계산.py", label="🟩 벽판 계산 시작하기", icon=None
            )
        elif not has_ceil:
            st.page_link(
                "pages/3_천장판_계산.py", label="🟨 천장판 계산 시작하기", icon=None
            )

    st.stop()  # 이전 단계 미완료 시 이후 UI 차단

# 모든 단계 완료 시 성공 메시지
st.success("✅ 모든 계산이 완료되었습니다. 견적서를 생성할 수 있습니다.")

# Convert session_state data
floor_data = convert_floor_data(floor_result)
wall_data = convert_wall_data(wall_result)
ceiling_data = convert_ceiling_data(ceil_result)

# Sidebar: Pricebook upload
with st.sidebar:
    st.markdown("### ① 단가표 업로드")
    pricebook_file = st.file_uploader(
        "Sungil_DB2_new.xlsx (시트명: 자재단가내역)", type=["xlsx"]
    )

    st.markdown("---")
    st.markdown("### ② 계산 결과 (자동 연동)")
    st.success(f"✅ 바닥판: {floor_data.get('재질', 'N/A')}")
    st.success(f"✅ 벽판: {wall_data.get('총개수', 0)}장")
    st.success(f"✅ 천장판: {ceiling_data.get('총개수', 0)}장")

    st.markdown("---")
    st.markdown("### ③ 옵션 선택")

# 생산관리비 설정은 견적서 데이터가 생성된 후에 표시 (아래로 이동)
# 먼저 rows 데이터를 생성한 후 UI를 표시

# Load pricebook
price_df: Optional[pd.DataFrame] = None
if pricebook_file is not None:
    try:
        price_df = load_pricebook_from_excel(pricebook_file.read())
        st.sidebar.success(f"단가표 로드 완료: {len(price_df)}행")
    except Exception as e:
        st.sidebar.error(f"단가표 로드 실패: {e}")

# ----------------------------
# 공통 품목 자동지정 정의 (통합)
# ----------------------------

# Session State 키
AUTO_ITEMS_KEY = "auto_assigned_items"
AUTO_FLOOR_TYPE_KEY = "auto_floor_type"
AUTO_SHAPE_TYPE_KEY = "auto_shape_type"
SELECT_ITEMS_KEY = "select_items"
OPTIONAL_ITEMS_KEY = "optional_items"

# ═══════════════════════════════════════════════════════════════
# 【A】 자동지정 품목 (기본 포함, 수량 편집 가능)
# ═══════════════════════════════════════════════════════════════

# 【A-1】 완전 고정 수량 (바닥판/규격/형태 무관)
FIXED_QUANTITY_ITEMS = {
    # 오수/배수 배관류
    "엘보(Φ100)": 1,
    "엘보(Φ50)": 2,
    "오수구덮개": 1,
    "PVC접착제": 0.15,
    # 도기류 (양변기만 자동지정, 세면기는 선택)
    "양변기": 1,
    # 문세트 3종
    "PVC 4방문틀": 1,
    "ABS 문짝": 1,
    "도어하드웨어": 1,  # 도어락+경첩
    # 수전 3종 (세면기수전, 샤워수전, 슬라이드바)
    "세면기 수전": 1,
    "샤워수전": 1,
    "슬라이드바": 1,
    # 액세서리 5종
    "은경(거울)": 1,
    "수건걸이": 1,
    "휴지걸이": 1,
    "일자유리선반": 1,
    "코너선반": 1,
    # 욕실등 내함
    "욕실등, 콘센트 내함": 1,
    # 공통자재
    "실리콘(내항균성)": 4.5,
    "실리콘(외장용)": 1,
    "우레탄폼": 0.5,
    "이면지지클립": 1,
    "타일 평탄클립": 1,
    "에폭시 접착제": 1,
}

# 【A-2】 바닥판 종류에 따라 달라지는 수량
FLOOR_TYPE_ITEMS = {
    "PP": {
        "직관(Φ100)": 1,
        "직관(Φ50)": 2,
        "배수트랩(습식용)": 2,
        "배수트랩(상하용)": 0,
        "드레인커버(세면부)": 0,
        "드레인커버(샤워부)": 0,
        # PP전용 품목
        "양변기(오수구) 소켓(Φ100)": 1,
        "세면,바닥,샤워 배수세트(Φ175)": 2,
        "난방배관 소켓(Φ16)": 2,
        "클럽메쉬 세트(클립포함)": 1,
        "벽체코너 받침대": 5,
        "볼트": 5,
        "성형슬리브(오수)Φ125": 1,
        "성형슬리브(세면,바닥,샤워)Φ175": 2,
        "슬리브용 몰탈막음 스펀지": 3,
    },
    "GRP": {
        "직관(Φ100)": 2,
        "직관(Φ50)": 2.5,
        "배수트랩(습식용)": 0,
        "배수트랩(상하용)": 2,
        "드레인커버(세면부)": 1,
        "드레인커버(샤워부)": 1,
        # PP전용 품목은 GRP에서 0
        "양변기(오수구) 소켓(Φ100)": 0,
        "세면,바닥,샤워 배수세트(Φ175)": 0,
        "난방배관 소켓(Φ16)": 0,
        "클럽메쉬 세트(클립포함)": 0,
        "벽체코너 받침대": 0,
        "볼트": 0,
        "성형슬리브(오수)Φ125": 0,
        "성형슬리브(세면,바닥,샤워)Φ175": 0,
        "슬리브용 몰탈막음 스펀지": 0,
    },
}

# 【A-3】 형태(사각형/코너형)에 따라 달라지는 수량
SHAPE_TYPE_ITEMS = {
    "사각형": {
        "코너마감재": 3,
        "코너비드": 0,
    },
    "코너형": {
        "코너마감재": 5,
        "코너비드": 1,
    },
}

# ═══════════════════════════════════════════════════════════════
# 【B】 선택 유지 품목 (종류 선택 필요)
# ═══════════════════════════════════════════════════════════════

# 【B-1】 기본값 있음 (4개)
SELECT_ITEMS_WITH_DEFAULT = {
    "냉온수배관": {
        "options": ["선택안함", "PB 독립배관", "PB 세대 세트 배관", "PB+이중관(오픈수전함)"],
        "default": "PB+이중관(오픈수전함)",
        "category": "냉온수배관",
    },
    "세면기": {
        "options": ["선택안함", "긴다리 세면기", "반다리 세면기"],
        "default": "긴다리 세면기",
        "category": "도기류",
    },
    "욕실장": {
        "options": ["선택안함", "욕실장(일반형)", "PS장(600*900)", "슬라이딩 욕실장"],
        "default": "욕실장(일반형)",
        "category": "욕실장",
    },
    "문틀규격": {
        "options": ["선택안함", "110m/m", "130m/m", "140m/m", "155m/m", "175m/m", "195m/m", "210m/m", "230m/m"],
        "default": "선택안함",  # 필수 선택 (벽체 두께에 따라)
        "category": "문세트",
    },
}

# 【B-2】 기본값 = 선택안함 (옵션 품목, 12개)
OPTIONAL_ITEMS = {
    "칸막이": {
        "options": ["선택안함", "샤워부스", "샤워파티션"],
        "default": "선택안함",
        "category": "칸막이",
    },
    "욕조": {
        "options": ["선택안함", "SQ욕조", "세라믹 욕조"],
        "default": "선택안함",
        "category": "욕조",
    },
    "환기류": {
        "options": ["선택안함", "환풍기", "후렉시블 호스, 서스밴드"],
        "default": "선택안함",
        "category": "환기류",
    },
    "도어스토퍼": {
        "options": ["선택안함", "도어스토퍼"],
        "default": "선택안함",
        "category": "문세트",
    },
    "손끼임방지": {
        "options": ["선택안함", "손끼임방지"],
        "default": "선택안함",
        "category": "문세트",
    },
    "청소건": {
        "options": ["선택안함", "청소건"],
        "default": "선택안함",
        "category": "수전",
    },
    "레인샤워수전": {
        "options": ["선택안함", "레인 샤워수전", "선반형 레인 샤워수전"],
        "default": "선택안함",
        "category": "수전",
    },
    "세탁기수전": {
        "options": ["선택안함", "세탁기 수전"],
        "default": "선택안함",
        "category": "수전",
    },
    "매립형휴지걸이": {
        "options": ["선택안함", "매립형 휴지걸이"],
        "default": "선택안함",
        "category": "액세서리",
    },
    "청소솔": {
        "options": ["선택안함", "청소솔"],
        "default": "선택안함",
        "category": "액세서리",
    },
    "2단수건선반": {
        "options": ["선택안함", "2단 수건선반"],
        "default": "선택안함",
        "category": "액세서리",
    },
    "욕실등(등기구)": {
        "options": ["선택안함", "천장 매립등(사각)", "천장 매립등(원형)", "벽부등"],
        "default": "선택안함",
        "category": "욕실등",
    },
}

# ═══════════════════════════════════════════════════════════════
# 자동지정 품목 계산 함수
# ═══════════════════════════════════════════════════════════════

def calculate_auto_items(floor_type: str, shape_type: str) -> Dict[str, float]:
    """바닥판 종류와 형태에 따라 자동지정 품목 수량 계산"""
    items = {}

    # 1단계: 고정 품목
    items.update(FIXED_QUANTITY_ITEMS.copy())

    # 2단계: 바닥판 종류별 품목
    if floor_type in FLOOR_TYPE_ITEMS:
        items.update(FLOOR_TYPE_ITEMS[floor_type])

    # 3단계: 형태별 품목
    if shape_type in SHAPE_TYPE_ITEMS:
        items.update(SHAPE_TYPE_ITEMS[shape_type])

    return items

# ═══════════════════════════════════════════════════════════════
# UI: 바닥판 종류 및 형태 선택
# ═══════════════════════════════════════════════════════════════
st.markdown("---")
st.subheader("품목 설정")

col_floor_type, col_shape_type = st.columns(2)

with col_floor_type:
    # 바닥판 재질에서 종류 추출
    floor_material = floor_data.get("재질", "").upper() if floor_data else ""

    # PP/PE, PVE는 PP로, GRP/SMC/FRP는 GRP로 매핑
    if "PP" in floor_material or "PE" in floor_material or "PVE" in floor_material:
        default_floor_type = "PP"
    else:
        default_floor_type = "GRP"

    prev_floor_type = st.session_state.get(AUTO_FLOOR_TYPE_KEY, default_floor_type)

    floor_type = st.radio(
        "바닥판 종류",
        options=["PP", "GRP"],
        index=0 if prev_floor_type == "PP" else 1,
        horizontal=True,
        help="PP: PP/PE/PVE 바닥판, GRP: GRP/SMC/FRP 바닥판",
        key="floor_type_radio"
    )

with col_shape_type:
    prev_shape_type = st.session_state.get(AUTO_SHAPE_TYPE_KEY, "사각형")

    shape_type = st.radio(
        "욕실 형태",
        options=["사각형", "코너형"],
        index=0 if prev_shape_type == "사각형" else 1,
        horizontal=True,
        help="사각형: 코너마감재 3개, 코너형: 코너마감재 5개 + 코너비드 1개",
        key="shape_type_radio"
    )

# 변경 감지
floor_type_changed = st.session_state.get(AUTO_FLOOR_TYPE_KEY) != floor_type
shape_type_changed = st.session_state.get(AUTO_SHAPE_TYPE_KEY) != shape_type

st.session_state[AUTO_FLOOR_TYPE_KEY] = floor_type
st.session_state[AUTO_SHAPE_TYPE_KEY] = shape_type

current_auto_items = calculate_auto_items(floor_type, shape_type)

if floor_type_changed or shape_type_changed:
    st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()
    st.info(f"바닥판 종류({floor_type}) 또는 형태({shape_type}) 변경으로 품목 수량이 초기화되었습니다.")

if AUTO_ITEMS_KEY not in st.session_state:
    st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()

# ═══════════════════════════════════════════════════════════════
# UI: 【B-1】 선택 유지 품목 (기본값 있음)
# ═══════════════════════════════════════════════════════════════
st.markdown("---")
st.markdown("### 필수 선택 품목")
st.caption("종류를 선택해야 하는 품목입니다. 기본값이 설정되어 있습니다.")

# 선택값 초기화
if SELECT_ITEMS_KEY not in st.session_state:
    st.session_state[SELECT_ITEMS_KEY] = {
        name: info["default"] for name, info in SELECT_ITEMS_WITH_DEFAULT.items()
    }

select_cols = st.columns(2)
select_values = {}

for idx, (name, info) in enumerate(SELECT_ITEMS_WITH_DEFAULT.items()):
    with select_cols[idx % 2]:
        options = info["options"]
        default_val = st.session_state[SELECT_ITEMS_KEY].get(name, info["default"])
        default_idx = options.index(default_val) if default_val in options else 0

        selected = st.selectbox(
            name,
            options=options,
            index=default_idx,
            key=f"select_{name}"
        )
        select_values[name] = selected

st.session_state[SELECT_ITEMS_KEY] = select_values

# ═══════════════════════════════════════════════════════════════
# UI: 【B-2】 옵션 품목 (기본값 = 선택안함)
# ═══════════════════════════════════════════════════════════════
with st.expander("옵션 품목 (선택사항)", expanded=False):
    st.caption("필요시 선택하세요. 기본값은 '선택안함'입니다.")

    if OPTIONAL_ITEMS_KEY not in st.session_state:
        st.session_state[OPTIONAL_ITEMS_KEY] = {
            name: info["default"] for name, info in OPTIONAL_ITEMS.items()
        }

    opt_cols = st.columns(3)
    opt_values = {}

    for idx, (name, info) in enumerate(OPTIONAL_ITEMS.items()):
        with opt_cols[idx % 3]:
            options = info["options"]
            default_val = st.session_state[OPTIONAL_ITEMS_KEY].get(name, info["default"])
            default_idx = options.index(default_val) if default_val in options else 0

            selected = st.selectbox(
                name,
                options=options,
                index=default_idx,
                key=f"opt_{name}"
            )
            opt_values[name] = selected

    st.session_state[OPTIONAL_ITEMS_KEY] = opt_values

# ═══════════════════════════════════════════════════════════════
# UI: 【A】 자동지정 품목 수량 편집
# ═══════════════════════════════════════════════════════════════
with st.expander("자동지정 품목 수량 편집", expanded=False):
    st.markdown("**기본 포함되는 품목의 수량을 편집할 수 있습니다.**")
    st.caption(f"현재 설정: 바닥판={floor_type}, 형태={shape_type}")

    if st.button("기본값으로 초기화", key="reset_auto_items"):
        st.session_state[AUTO_ITEMS_KEY] = current_auto_items.copy()
        st.success("기본값으로 초기화되었습니다.")
        st.rerun()

    # 카테고리별 분류
    auto_categories = {
        "오수/배수 배관류": ["엘보(Φ100)", "엘보(Φ50)", "직관(Φ100)", "직관(Φ50)",
                          "오수구덮개", "PVC접착제", "배수트랩(습식용)", "배수트랩(상하용)",
                          "드레인커버(세면부)", "드레인커버(샤워부)"],
        "PP바닥판 전용": ["양변기(오수구) 소켓(Φ100)", "세면,바닥,샤워 배수세트(Φ175)",
                        "난방배관 소켓(Φ16)", "클럽메쉬 세트(클립포함)", "벽체코너 받침대",
                        "볼트", "성형슬리브(오수)Φ125", "성형슬리브(세면,바닥,샤워)Φ175",
                        "슬리브용 몰탈막음 스펀지"],
        "도기류": ["양변기"],
        "문세트": ["PVC 4방문틀", "ABS 문짝", "도어하드웨어"],
        "수전": ["세면기 수전", "샤워수전", "슬라이드바"],
        "액세서리": ["은경(거울)", "수건걸이", "휴지걸이", "일자유리선반", "코너선반"],
        "욕실등": ["욕실등, 콘센트 내함"],
        "공통자재": ["실리콘(내항균성)", "실리콘(외장용)", "우레탄폼",
                    "이면지지클립", "타일 평탄클립", "에폭시 접착제",
                    "코너마감재", "코너비드"],
    }

    edited_items = st.session_state.get(AUTO_ITEMS_KEY, current_auto_items).copy()

    for cat_name, item_list in auto_categories.items():
        st.markdown(f"**{cat_name}**")
        cols = st.columns(3)
        col_idx = 0

        for item_name in item_list:
            if item_name in current_auto_items:
                saved_qty = edited_items.get(item_name, current_auto_items[item_name])
                default_qty = current_auto_items[item_name]

                with cols[col_idx % 3]:
                    help_text = "현재 바닥판에서 미사용 (기본: 0)" if default_qty == 0 else f"기본값: {default_qty}"
                    edited_qty = st.number_input(
                        item_name,
                        min_value=0.0,
                        max_value=100.0,
                        value=float(saved_qty),
                        step=0.5,
                        key=f"auto_{item_name}",
                        help=help_text
                    )
                    edited_items[item_name] = edited_qty
                col_idx += 1

    st.session_state[AUTO_ITEMS_KEY] = edited_items

# 최종 자동지정 품목
final_auto_items = st.session_state.get(AUTO_ITEMS_KEY, current_auto_items)
final_select_items = st.session_state.get(SELECT_ITEMS_KEY, {})
final_optional_items = st.session_state.get(OPTIONAL_ITEMS_KEY, {})

# ----------------------------
# 견적서 생성
# ----------------------------
rows: List[Dict[str, Any]] = []
warnings: List[str] = []

if price_df is None:
    st.warning("단가표(엑셀)를 먼저 업로드하세요.")
else:
    # 1) 바닥판
    if floor_data:
        material = str(floor_data.get("재질", "")).upper()
        spec_text = str(floor_data.get("규격", "")).strip()
        qty = float(floor_data.get("수량", 1))
        unit_price = float(floor_data.get("단가", 0))
        senior = bool(floor_data.get("주거약자", False))

        # 품목 '바닥판' 본체
        add_row(rows, "바닥판", material, qty, unit_price)

        # 부재료 자동 포함
        if material in ["GRP", "SMC/FRP", "PP/PE", "PVE"]:
            if material == "PVE":
                분류 = "PP/PE 부재료"
            elif material == "SMC/FRP":
                분류 = "SMC/FRP 부재료"
            elif material == "PP/PE":
                분류 = "PP/PE 부재료"
            else:
                분류 = "GRP부재료"
            add_all_by_category(rows, price_df, "바닥판", 분류)
        else:
            warnings.append(
                f"바닥판 재질 '{material}'에 대한 분류 매핑을 찾을 수 없습니다."
            )

        # 주거약자 추가
        if senior:
            for spec in [
                "매립형 휴지걸이(비상폰)",
                "L형 손잡이",
                "ㅡ형 손잡이",
                "접의식 의자",
            ]:
                rec = find_item(price_df, "액세서리", "주거약자", spec_contains=spec)
                if rec is not None:
                    add_row(
                        rows,
                        "액세서리",
                        spec,
                        rec.get("수량", 1) or 1,
                        rec.get("단가", 0),
                    )
                else:
                    add_row(rows, "액세서리", spec, 1, 0)
                    warnings.append(f"주거약자 '{spec}' 단가 미발견 → 0 처리")

    # 2) 벽판 & 타일
    if wall_data:
        # PU벽판
        wall_spec = "PU벽판"
        rec = find_item(price_df, "벽판", "PU타일 벽체", spec_contains="PU벽판")
        qty = float(wall_data.get("총개수", 0))
        unit_price = None
        if rec is not None:
            unit_price = rec.get("단가", None)
        else:
            unit_price = float(wall_data.get("단가", 0))
            warnings.append(
                "벽판(PU벽판) 단가를 엑셀에서 찾지 못해 기본값 0으로 설정했습니다."
            )
        add_row(rows, "벽판", wall_spec, qty, unit_price)

        # 벽타일 & 바닥타일 규격 연동
        tile_str = str(wall_data.get("벽타일", "")).replace("×", "x").replace(" ", "")
        wall_tile_spec = None
        if tile_str in ["250x400", "250*400"]:
            wall_tile_spec = "벽타일 250*400"
            floor_tile_spec = "바닥타일 200*200"
        else:
            wall_tile_spec = "벽타일 300*600"
            floor_tile_spec = "바닥타일 300*300"

        # 벽타일
        rec = find_item(
            price_df, "타일", "PU타일 벽체 타일", spec_contains=wall_tile_spec
        )
        if rec is not None:
            add_row(
                rows,
                "타일",
                wall_tile_spec,
                rec.get("수량", 1) or 1,
                rec.get("단가", 0),
            )
        else:
            add_row(rows, "타일", wall_tile_spec, 1, 0)
            warnings.append(f"'{wall_tile_spec}' 단가 미발견 → 0 처리")

        # 바닥타일
        rec = find_item(
            price_df, "타일", "바닥타일", spec_contains=floor_tile_spec.split()[-1]
        )
        if rec is None:
            rec = find_item(price_df, "타일", "바닥타일", spec_contains=floor_tile_spec)
        if rec is not None:
            add_row(
                rows,
                "타일",
                floor_tile_spec,
                rec.get("수량", 1) or 1,
                rec.get("단가", 0),
            )
        else:
            add_row(rows, "타일", floor_tile_spec, 1, 0)
            warnings.append(f"'{floor_tile_spec}' 단가 미발견 → 0 처리")

    # 3) 천장판
    if ceiling_data:
        material = str(ceiling_data.get("재질", "")).upper()
        body = ceiling_data.get("바디판넬", {}) or {}
        side = ceiling_data.get("사이드판넬", {}) or {}
        total_cnt = float(ceiling_data.get("총개수", 0))

        # 천공구: 딕셔너리인 경우 개수 필드 추출
        hole_data = ceiling_data.get("천공구", 0)
        hole_cnt = float(hole_data.get("개수", 0) if isinstance(hole_data, dict) else hole_data)

        # 메인 판
        if material == "ABS":
            rec = find_item(price_df, "천장판", None, spec_contains="ABS천장판")
            add_row(
                rows,
                "천장판",
                "ABS천장판",
                total_cnt or (body.get("개수", 0) + side.get("개수", 0)),
                rec.get("단가", 0) if rec is not None else 0,
            )
            if rec is None:
                warnings.append("ABS천장판 단가 미발견 → 0 처리")
        elif material == "GRP":
            rec = find_item(price_df, "천장판", None, spec_contains="GRP천장판")
            add_row(
                rows,
                "천장판",
                "GRP천장판",
                total_cnt or (body.get("개수", 0) + side.get("개수", 0)),
                rec.get("단가", 0) if rec is not None else 0,
            )
            if rec is None:
                warnings.append("GRP천장판 단가 미발견 → 0 처리")
        else:
            add_row(rows, "천장판", material, total_cnt, 0)
            warnings.append(f"천장판 재질 '{material}' 단가 미발견 → 0 처리")

        # 세부 수량 표기 (정보용)
        if body.get("개수", 0):
            add_row(
                rows,
                "천장판",
                f"바디판넬 ({body.get('종류','')})",
                float(body.get("개수", 0)),
                float(ceiling_data.get("단가", 0)),
            )
        if side.get("개수", 0):
            add_row(
                rows,
                "천장판",
                f"사이드판넬 ({side.get('종류','')})",
                float(side.get("개수", 0)),
                float(ceiling_data.get("단가", 0)),
            )
        if hole_cnt:
            add_row(rows, "천장판", "천공구", hole_cnt, 0)

    # 4) 필수 선택 품목 반영 (SELECT_ITEMS_WITH_DEFAULT)
    for name, spec in final_select_items.items():
        if spec == "선택안함":
            continue

        item_info = SELECT_ITEMS_WITH_DEFAULT.get(name, {})
        category = item_info.get("category", name)

        rec = find_item(price_df, category, None, spec_contains=spec)
        if rec is not None:
            add_row(rows, category, spec, rec.get("수량", 1) or 1, rec.get("단가", 0))
        else:
            add_row(rows, category, spec, 1, 0)
            warnings.append(f"[필수선택] '{name} - {spec}' 단가 미발견 → 0 처리")

    # 5) 옵션 품목 반영 (OPTIONAL_ITEMS)
    for name, spec in final_optional_items.items():
        if spec == "선택안함":
            continue

        item_info = OPTIONAL_ITEMS.get(name, {})
        category = item_info.get("category", name)

        rec = find_item(price_df, category, None, spec_contains=spec)
        if rec is not None:
            add_row(rows, category, spec, rec.get("수량", 1) or 1, rec.get("단가", 0))
        else:
            add_row(rows, category, spec, 1, 0)
            warnings.append(f"[옵션] '{name} - {spec}' 단가 미발견 → 0 처리")

    # 6) 자동지정 품목 추가
    # 품목명을 단가표에서 찾기 위한 매핑
    AUTO_ITEM_CATEGORY_MAP = {
        # 오수/배수 배관류
        "엘보(Φ100)": ("오,배수배관류", "엘보"),
        "엘보(Φ50)": ("오,배수배관류", "엘보"),
        "직관(Φ100)": ("오,배수배관류", "직관"),
        "직관(Φ50)": ("오,배수배관류", "직관"),
        "오수구덮개": ("오,배수배관류", None),
        "PVC접착제": ("오,배수배관류", None),
        "배수트랩(습식용)": ("오,배수배관류", "배수트랩"),
        "배수트랩(상하용)": ("오,배수배관류", "배수트랩"),
        "드레인커버(세면부)": ("오,배수배관류", "드레인커버"),
        "드레인커버(샤워부)": ("오,배수배관류", "드레인커버"),
        # PP바닥판 전용
        "양변기(오수구) 소켓(Φ100)": ("오,배수배관류", "소켓"),
        "세면,바닥,샤워 배수세트(Φ175)": ("오,배수배관류", "배수세트"),
        "난방배관 소켓(Φ16)": ("오,배수배관류", "소켓"),
        "클럽메쉬 세트(클립포함)": ("오,배수배관류", "클럽메쉬"),
        "벽체코너 받침대": ("오,배수배관류", "받침대"),
        "볼트": ("오,배수배관류", "볼트"),
        "성형슬리브(오수)Φ125": ("오,배수배관류", "슬리브"),
        "성형슬리브(세면,바닥,샤워)Φ175": ("오,배수배관류", "슬리브"),
        "슬리브용 몰탈막음 스펀지": ("오,배수배관류", "스펀지"),
        # 도기류 (양변기만 자동지정)
        "양변기": ("도기류", "양변기"),
        # 문세트 3종
        "PVC 4방문틀": ("문세트", "4방"),
        "ABS 문짝": ("문세트", "문짝"),
        "도어하드웨어": ("문세트", "도어락"),  # 도어락+경첩
        # 수전 3종
        "세면기 수전": ("수전", "세면기"),
        "샤워수전": ("수전", "샤워"),
        "슬라이드바": ("수전", "슬라이드바"),
        # 액세서리 5종
        "은경(거울)": ("은경", None),
        "수건걸이": ("액세서리", "수건걸이"),
        "휴지걸이": ("액세서리", "휴지걸이"),
        "일자유리선반": ("액세서리", "유리선반"),
        "코너선반": ("액세서리", "코너선반"),
        # 욕실등
        "욕실등, 콘센트 내함": ("공통자재", "내함"),
        # 공통자재
        "실리콘(내항균성)": ("공통자재", "내항균"),
        "실리콘(외장용)": ("공통자재", "외장"),
        "우레탄폼": ("공통자재", "우레탄폼"),
        "이면지지클립": ("공통자재", "이면지지"),
        "타일 평탄클립": ("공통자재", "평탄클립"),
        "에폭시 접착제": ("공통자재", "에폭시"),
        "코너마감재": ("공통자재", "코너마감재"),
        "코너비드": ("공통자재", "코너비드"),
    }

    # 이미 추가된 품목 추적 (중복 방지)
    added_specs = set()
    for r in rows:
        spec_key = f"{r['품목']}::{r['사양 및 규격']}"
        added_specs.add(spec_key)

    # 자동지정 품목 추가
    for item_name, qty in final_auto_items.items():
        if qty <= 0:
            continue  # 수량이 0이면 스킵

        # 카테고리 매핑 조회
        cat_info = AUTO_ITEM_CATEGORY_MAP.get(item_name)
        if cat_info is None:
            continue

        품목_cat, spec_hint = cat_info

        # 중복 체크
        spec_key = f"{품목_cat}::{item_name}"
        if spec_key in added_specs:
            continue  # 이미 추가됨

        # 단가표에서 찾기
        rec = None
        if spec_hint:
            rec = find_item(price_df, 품목_cat, None, spec_contains=spec_hint)
            if rec is None:
                rec = find_item(price_df, 품목_cat, None, spec_contains=item_name)
        else:
            rec = find_item(price_df, 품목_cat, None, spec_contains=item_name)

        # 단가 설정
        if rec is not None:
            unit_price = rec.get("단가", 0) or 0
        else:
            unit_price = 0
            warnings.append(f"[자동지정] '{item_name}' 단가 미발견 → 0 처리")

        # rows에 추가
        add_row(rows, 품목_cat, item_name, qty, unit_price)
        added_specs.add(spec_key)

# ----------------------------
# 결과 표
# ----------------------------
if rows:
    est_df = pd.DataFrame(
        rows, columns=["품목", "사양 및 규격", "수량", "단가", "금액"]
    )
    est_df["수량"] = (
        pd.to_numeric(est_df["수량"], errors="coerce").fillna(0).astype(float)
    )
    est_df["단가"] = (
        pd.to_numeric(est_df["단가"], errors="coerce").fillna(0).astype(float)
    )
    est_df["금액"] = (est_df["수량"] * est_df["단가"]).round(0)

    st.subheader("견적서 미리보기")
    st.dataframe(est_df, use_container_width=True)

    totals = (
        est_df.groupby("품목", dropna=False)["금액"]
        .sum()
        .reset_index()
        .sort_values("금액", ascending=False)
    )
    st.markdown("#### 품목별 합계")
    st.dataframe(totals, use_container_width=True)

    grand_total = est_df["금액"].sum()
    st.metric("총 금액 (생산관리비 제외)", f"{grand_total:,.0f} 원")

    # ----------------------------
    # 생산관리비 설정 UI (견적서 데이터 기반)
    # ----------------------------
    st.markdown("---")
    st.subheader("생산관리비 설정")

    # 현재 견적서의 모든 품목+사양 목록 추출
    available_items = []
    for _, row in est_df.iterrows():
        item_key = get_item_key(str(row["품목"]), str(row["사양 및 규격"]))
        if item_key not in [i[0] for i in available_items]:
            available_items.append((item_key, str(row["품목"]), str(row["사양 및 규격"]), float(row["금액"])))

    # 세션 상태 초기화
    if PROD_MGMT_SETTINGS_KEY not in st.session_state:
        st.session_state[PROD_MGMT_SETTINGS_KEY] = {
            cat: {"items": list(info["items"]), "rate": info["rate"]}
            for cat, info in DEFAULT_PROD_MGMT_CATEGORIES.items()
        }

    prod_mgmt_categories = st.session_state[PROD_MGMT_SETTINGS_KEY]

    # 카테고리 관리 UI
    with st.expander("카테고리 관리 (추가/수정/삭제)", expanded=False):
        # 새 카테고리 추가
        st.markdown("##### 새 카테고리 추가")
        col_new_name, col_new_rate, col_new_btn = st.columns([3, 1, 1])
        with col_new_name:
            new_cat_name = st.text_input("카테고리명", key="new_cat_name", placeholder="예: 신규카테고리")
        with col_new_rate:
            new_cat_rate = st.number_input("비율(%)", min_value=0.0, max_value=100.0, value=0.0, step=0.5, key="new_cat_rate")
        with col_new_btn:
            st.write("")
            if st.button("➕ 추가", key="add_cat_btn"):
                if new_cat_name and new_cat_name not in prod_mgmt_categories:
                    prod_mgmt_categories[new_cat_name] = {"items": [], "rate": new_cat_rate}
                    st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                    st.success(f"'{new_cat_name}' 카테고리 추가됨")
                    st.rerun()
                elif new_cat_name in prod_mgmt_categories:
                    st.error("이미 존재하는 카테고리명입니다.")

        st.markdown("---")

        # 기존 카테고리 삭제
        st.markdown("##### 카테고리 삭제")
        if prod_mgmt_categories:
            cat_to_delete = st.selectbox(
                "삭제할 카테고리",
                options=list(prod_mgmt_categories.keys()),
                key="cat_to_delete"
            )
            if st.button("🗑️ 선택 카테고리 삭제", key="delete_cat_btn"):
                if cat_to_delete in prod_mgmt_categories:
                    del prod_mgmt_categories[cat_to_delete]
                    st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                    st.success(f"'{cat_to_delete}' 카테고리 삭제됨")
                    st.rerun()

        # 기본값으로 초기화
        st.markdown("---")
        if st.button("🔄 기본 카테고리로 초기화", key="reset_cat_btn"):
            st.session_state[PROD_MGMT_SETTINGS_KEY] = {
                cat: {"items": list(info["items"]), "rate": info["rate"]}
                for cat, info in DEFAULT_PROD_MGMT_CATEGORIES.items()
            }
            st.success("기본 카테고리로 초기화됨")
            st.rerun()

    # 카테고리별 설정 UI
    st.markdown("#### 카테고리별 비율 및 포함 항목 설정")

    # 각 카테고리별 설정
    updated_categories = {}

    for cat_name, cat_info in prod_mgmt_categories.items():
        with st.expander(f"📁 {cat_name}", expanded=True):
            col_rate, col_info = st.columns([1, 3])

            with col_rate:
                rate = st.number_input(
                    "비율(%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=float(cat_info.get("rate", 0.0)),
                    step=0.5,
                    key=f"rate_{cat_name}"
                )

            # 현재 카테고리에 매칭되는 항목 찾기
            matched_items = []
            unmatched_items = []
            cat_items = cat_info.get("items", [])

            for item_key, 품목, 사양, 금액 in available_items:
                is_matched = False
                for pattern_품목, pattern_사양 in cat_items:
                    if item_matches_pattern(품목, 사양, pattern_품목, pattern_사양):
                        is_matched = True
                        break
                if is_matched:
                    matched_items.append((item_key, 품목, 사양, 금액))
                else:
                    unmatched_items.append((item_key, 품목, 사양, 금액))

            with col_info:
                matched_total = sum(금액 for _, _, _, 금액 in matched_items)
                mgmt_fee = matched_total * (rate / 100.0)
                st.markdown(f"**소계:** {matched_total:,.0f}원 → **생산관리비:** {mgmt_fee:,.0f}원")

            # 포함 항목 표시 및 편집
            st.markdown("**포함 항목:**")

            # 패턴 기반 항목 표시
            if cat_items:
                pattern_strs = []
                for p_품목, p_사양 in cat_items:
                    if p_사양:
                        pattern_strs.append(f"{p_품목}({p_사양})")
                    else:
                        pattern_strs.append(f"{p_품목}(전체)")
                st.caption(f"패턴: {', '.join(pattern_strs)}")

            # 실제 매칭된 항목 표시
            if matched_items:
                matched_df = pd.DataFrame([
                    {"품목": 품목, "사양": 사양, "금액": f"{금액:,.0f}"}
                    for _, 품목, 사양, 금액 in matched_items
                ])
                st.dataframe(matched_df, use_container_width=True, hide_index=True, height=min(150, 35 * (len(matched_items) + 1)))
            else:
                st.info("매칭된 항목 없음")

            # 항목 추가/제거 (패턴 편집)
            st.markdown("**패턴 편집:**")
            col_add_pattern, col_remove_pattern = st.columns(2)

            with col_add_pattern:
                # 추가할 품목 선택
                unique_품목s = list(set(품목 for _, 품목, _, _ in available_items))
                add_품목 = st.selectbox(f"품목 선택", options=[""] + unique_품목s, key=f"add_품목_{cat_name}")

                if add_품목:
                    # 해당 품목의 사양 목록
                    품목_사양s = [사양 for _, 품목, 사양, _ in available_items if 품목 == add_품목]
                    add_사양_option = st.selectbox(
                        "사양 범위",
                        options=["전체(품목 전체 포함)"] + 품목_사양s,
                        key=f"add_사양_{cat_name}"
                    )

                    if st.button("➕ 패턴 추가", key=f"add_pattern_btn_{cat_name}"):
                        if add_사양_option == "전체(품목 전체 포함)":
                            new_pattern = (add_품목, None)
                        else:
                            new_pattern = (add_품목, add_사양_option)

                        if new_pattern not in cat_items:
                            cat_items.append(new_pattern)
                            prod_mgmt_categories[cat_name]["items"] = cat_items
                            st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                            st.success(f"패턴 추가됨")
                            st.rerun()

            with col_remove_pattern:
                if cat_items:
                    pattern_options = []
                    for p_품목, p_사양 in cat_items:
                        if p_사양:
                            pattern_options.append(f"{p_품목}({p_사양})")
                        else:
                            pattern_options.append(f"{p_품목}(전체)")

                    remove_pattern_str = st.selectbox(
                        "제거할 패턴",
                        options=pattern_options,
                        key=f"remove_pattern_{cat_name}"
                    )

                    if st.button("➖ 패턴 제거", key=f"remove_pattern_btn_{cat_name}"):
                        # 패턴 문자열을 다시 튜플로 변환
                        idx = pattern_options.index(remove_pattern_str)
                        cat_items.pop(idx)
                        prod_mgmt_categories[cat_name]["items"] = cat_items
                        st.session_state[PROD_MGMT_SETTINGS_KEY] = prod_mgmt_categories
                        st.success(f"패턴 제거됨")
                        st.rerun()

            updated_categories[cat_name] = {
                "items": cat_items,
                "rate": rate
            }

    # 설정 업데이트
    st.session_state[PROD_MGMT_SETTINGS_KEY] = updated_categories
    prod_mgmt_categories = updated_categories

    # ----------------------------
    # 생산관리비 카테고리별 합계 계산 및 표시
    # ----------------------------
    st.markdown("---")
    st.markdown("#### 생산관리비 카테고리별 합계")

    # 각 항목이 어느 카테고리에 속하는지 매핑
    item_to_category = {}
    for item_key, 품목, 사양, 금액 in available_items:
        for cat_name, cat_info in prod_mgmt_categories.items():
            for pattern_품목, pattern_사양 in cat_info.get("items", []):
                if item_matches_pattern(품목, 사양, pattern_품목, pattern_사양):
                    item_to_category[item_key] = cat_name
                    break
            if item_key in item_to_category:
                break
        if item_key not in item_to_category:
            item_to_category[item_key] = "미분류"

    # 카테고리별 소계 계산
    category_subtotals = {cat_name: 0.0 for cat_name in prod_mgmt_categories.keys()}
    category_subtotals["미분류"] = 0.0

    for item_key, 품목, 사양, 금액 in available_items:
        cat = item_to_category.get(item_key, "미분류")
        category_subtotals[cat] += 금액

    # 생산관리비 계산
    category_mgmt_fees = {}
    total_mgmt_fee = 0.0

    for cat_name, subtotal in category_subtotals.items():
        if cat_name == "미분류":
            rate = 0.0
        else:
            rate = prod_mgmt_categories.get(cat_name, {}).get("rate", 0.0)
        mgmt_fee = subtotal * (rate / 100.0)
        category_mgmt_fees[cat_name] = mgmt_fee
        total_mgmt_fee += mgmt_fee

    # 표 형식으로 표시
    mgmt_summary_data = []
    for cat_name in list(prod_mgmt_categories.keys()) + (["미분류"] if category_subtotals.get("미분류", 0) > 0 else []):
        subtotal = category_subtotals.get(cat_name, 0)
        if cat_name == "미분류":
            rate = 0.0
        else:
            rate = prod_mgmt_categories.get(cat_name, {}).get("rate", 0.0)
        mgmt_fee = category_mgmt_fees.get(cat_name, 0)
        total_with_mgmt = subtotal + mgmt_fee
        mgmt_summary_data.append({
            "카테고리": cat_name,
            "소계": f"{subtotal:,.0f}",
            "비율(%)": f"{rate:.1f}",
            "생산관리비": f"{mgmt_fee:,.0f}",
            "총계": f"{total_with_mgmt:,.0f}",
        })

    mgmt_summary_df = pd.DataFrame(mgmt_summary_data)
    st.dataframe(mgmt_summary_df, use_container_width=True, hide_index=True)

    # ----------------------------
    # 영업관리비 설정 UI
    # ----------------------------
    st.markdown("---")
    st.subheader("영업관리비 설정 (선택)")

    # 영업관리비 세션 상태 초기화
    if SALES_MGMT_SETTINGS_KEY not in st.session_state:
        st.session_state[SALES_MGMT_SETTINGS_KEY] = {
            "enabled": False,
            "rate": 15.0,  # 기본값 15%
        }

    sales_settings = st.session_state[SALES_MGMT_SETTINGS_KEY]

    col_sales_enable, col_sales_rate = st.columns([1, 2])
    with col_sales_enable:
        sales_enabled = st.checkbox(
            "영업관리비 추가",
            value=sales_settings.get("enabled", False),
            help="체크하면 영업관리비가 견적서에 포함됩니다"
        )
    with col_sales_rate:
        if sales_enabled:
            sales_rate = st.number_input(
                "영업관리비 비율(%)",
                min_value=0.0,
                max_value=100.0,
                value=float(sales_settings.get("rate", 15.0)),
                step=0.5,
                key="sales_mgmt_rate"
            )
        else:
            sales_rate = 0.0
            st.info("영업관리비를 추가하려면 체크박스를 선택하세요")

    # 영업관리비 설정 업데이트
    st.session_state[SALES_MGMT_SETTINGS_KEY] = {
        "enabled": sales_enabled,
        "rate": sales_rate if sales_enabled else 0.0,
    }

    # 영업관리비 계산
    total_before_sales = grand_total + total_mgmt_fee  # 원가 + 생산관리비
    sales_mgmt_fee = total_before_sales * (sales_rate / 100.0) if sales_enabled else 0.0

    if sales_enabled:
        st.markdown(f"**영업관리비 기준금액:** {total_before_sales:,.0f}원 × {sales_rate:.1f}% = **{sales_mgmt_fee:,.0f}원**")

    # 최종 총계 (영업관리비 포함)
    final_total = grand_total + total_mgmt_fee + sales_mgmt_fee

    if sales_enabled:
        col_sub, col_mgmt, col_sales, col_final = st.columns(4)
        with col_sub:
            st.metric("원가 소계", f"{grand_total:,.0f} 원")
        with col_mgmt:
            st.metric("생산관리비", f"{total_mgmt_fee:,.0f} 원")
        with col_sales:
            st.metric("영업관리비", f"{sales_mgmt_fee:,.0f} 원")
        with col_final:
            st.metric("최종 총계", f"{final_total:,.0f} 원")
    else:
        col_sub, col_mgmt, col_final = st.columns(3)
        with col_sub:
            st.metric("원가 소계", f"{grand_total:,.0f} 원")
        with col_mgmt:
            st.metric("생산관리비 합계", f"{total_mgmt_fee:,.0f} 원")
        with col_final:
            st.metric("최종 총계", f"{final_total:,.0f} 원")

    # ----------------------------
    # 세대 타입 저장 기능
    # ----------------------------
    st.markdown("---")
    st.subheader("세대 타입 저장")

    # 저장된 견적 목록 초기화
    if SAVED_QUOTATIONS_KEY not in st.session_state:
        st.session_state[SAVED_QUOTATIONS_KEY] = []

    # 현재 세대 정보
    current_spec = floor_data.get("규격", "N/A") if floor_data else "N/A"
    current_units = floor_data.get("inputs", {}).get("units", 1) if floor_data else 1

    col_name, col_save = st.columns([3, 1])
    with col_name:
        type_name = st.text_input(
            "세대 타입 이름",
            value=f"타입{len(st.session_state[SAVED_QUOTATIONS_KEY]) + 1}",
            help="예: 21A,B,E/22C,F"
        )
    with col_save:
        st.write("")  # 공백으로 높이 맞춤
        st.write("")
        save_disabled = len(st.session_state[SAVED_QUOTATIONS_KEY]) >= 10
        if st.button(
            "💾 현재 견적 저장",
            disabled=save_disabled,
            help="최대 10개까지 저장 가능"
        ):
            # 현재 견적 데이터 저장 (생산관리비 정보 포함)
            # prod_mgmt_settings의 items를 리스트로 변환 (튜플은 JSON 직렬화 문제)
            serializable_settings = {}
            for k, v in prod_mgmt_categories.items():
                serializable_settings[k] = {
                    "items": [list(item) for item in v.get("items", [])],
                    "rate": v.get("rate", 0.0)
                }
            quotation_data = {
                "name": type_name,
                "spec": current_spec,
                "units": current_units,
                "rows": rows.copy(),  # 견적 항목 목록
                "total": grand_total,  # 원가 소계
                "total_mgmt_fee": total_mgmt_fee,  # 생산관리비 합계
                "sales_mgmt_fee": sales_mgmt_fee,  # 영업관리비
                "sales_mgmt_rate": sales_rate if sales_enabled else 0.0,  # 영업관리비 비율
                "sales_mgmt_enabled": sales_enabled,  # 영업관리비 활성화 여부
                "final_total": final_total,  # 최종 총계 (원가 + 생산관리비 + 영업관리비)
                "category_subtotals": dict(category_subtotals),  # 카테고리별 소계
                "category_mgmt_fees": dict(category_mgmt_fees),  # 카테고리별 생산관리비
                "prod_mgmt_settings": serializable_settings,  # 생산관리비 설정
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            st.session_state[SAVED_QUOTATIONS_KEY].append(quotation_data)
            st.success(f"✅ '{type_name}' 저장 완료! (규격: {current_spec}, {current_units}세대, 최종단가: {final_total:,.0f}원)")
            st.rerun()

    # 저장된 세대 타입 목록 표시
    saved_list = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
    if saved_list:
        st.markdown("#### 저장된 세대 타입 목록")

        # 테이블 형식으로 표시 (생산관리비, 영업관리비 포함)
        saved_df = pd.DataFrame([
            {
                "번호": i + 1,
                "타입명": q["name"],
                "규격": q["spec"],
                "세대수": q["units"],
                "원가 소계": f"{q['total']:,.0f}",
                "생산관리비": f"{q.get('total_mgmt_fee', 0):,.0f}",
                "영업관리비": f"{q.get('sales_mgmt_fee', 0):,.0f}" if q.get('sales_mgmt_enabled', False) else "-",
                "세대당 최종단가": f"{q.get('final_total', q['total']):,.0f}",
                "총 금액": f"{q.get('final_total', q['total']) * q['units']:,.0f}",
            }
            for i, q in enumerate(saved_list)
        ])
        st.dataframe(saved_df, use_container_width=True, hide_index=True)

        # 삭제 기능
        col_del, col_clear = st.columns([2, 1])
        with col_del:
            if len(saved_list) > 0:
                del_idx = st.selectbox(
                    "삭제할 타입 선택",
                    options=range(len(saved_list)),
                    format_func=lambda x: f"{x+1}. {saved_list[x]['name']} ({saved_list[x]['spec']})"
                )
                if st.button("🗑️ 선택 항목 삭제"):
                    del st.session_state[SAVED_QUOTATIONS_KEY][del_idx]
                    st.success("삭제 완료!")
                    st.rerun()
        with col_clear:
            st.write("")
            if st.button("🗑️ 전체 삭제", type="secondary"):
                st.session_state[SAVED_QUOTATIONS_KEY] = []
                st.success("전체 삭제 완료!")
                st.rerun()

        # 총 세대수 및 총 금액 합계 (생산관리비, 영업관리비 포함)
        total_all_units = sum(q["units"] for q in saved_list)
        total_all_amount = sum(q.get("final_total", q["total"]) * q["units"] for q in saved_list)
        total_all_cost = sum(q["total"] * q["units"] for q in saved_list)
        total_all_mgmt = sum(q.get("total_mgmt_fee", 0) * q["units"] for q in saved_list)
        total_all_sales = sum(q.get("sales_mgmt_fee", 0) * q["units"] for q in saved_list if q.get("sales_mgmt_enabled", False))
        if total_all_sales > 0:
            st.markdown(f"**총 세대수: {total_all_units}세대 | 원가합계: {total_all_cost:,.0f}원 | 생산관리비: {total_all_mgmt:,.0f}원 | 영업관리비: {total_all_sales:,.0f}원 | 최종합계: {total_all_amount:,.0f}원**")
        else:
            st.markdown(f"**총 세대수: {total_all_units}세대 | 원가합계: {total_all_cost:,.0f}원 | 생산관리비합계: {total_all_mgmt:,.0f}원 | 최종합계: {total_all_amount:,.0f}원**")

    st.markdown("---")

    # Excel 다운로드 (LGE 창원 스마트파크 형식)
    def df_to_excel_bytes(
        df: pd.DataFrame,
        total_units: int = 1,
        category_subtotals: dict = None,
        category_mgmt_fees: dict = None,
        prod_mgmt_settings: dict = None,
        total_mgmt_fee: float = 0.0,
        final_total: float = 0.0,
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "원자재 세대당 단가내역"

        # A4 가로 형식 설정
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0  # 높이는 자동

        # 가운데 정렬을 위해 왼쪽 여백 컬럼 추가
        LEFT_MARGIN = 3  # 왼쪽 여백 컬럼 수 (더 넓게)

        # 스타일 정의
        title_font = Font(name="맑은 고딕", size=18, bold=True)
        subtitle_font = Font(name="맑은 고딕", size=11, bold=True)
        header_font = Font(name="맑은 고딕", size=10, bold=True)
        data_font = Font(name="맑은 고딕", size=9)
        small_font = Font(name="맑은 고딕", size=8)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        # 투명 배경 (fill 제거)
        no_fill = PatternFill(fill_type=None)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 여백 컬럼 설정
        for i in range(1, LEFT_MARGIN + 1):
            ws.column_dimensions[chr(64 + i)].width = 2

        # 실제 시작 컬럼 (C부터)
        START_COL = LEFT_MARGIN + 1

        # 1행: 타이틀 - 가로로 넓게
        title_range = f"{chr(64+START_COL)}1:{chr(64+START_COL+7)}1"
        ws.merge_cells(title_range)
        title_cell = ws.cell(1, START_COL)
        title_cell.value = "욕실 원자재 세대당 단가 내역"
        title_cell.font = title_font
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30

        # 2-3행: 빈 행
        ws.row_dimensions[2].height = 10
        ws.row_dimensions[3].height = 10

        # 4행: 세대 정보 및 날짜
        info_range = f"{chr(64+START_COL)}4:{chr(64+START_COL+2)}4"
        ws.merge_cells(info_range)
        info_cell = ws.cell(4, START_COL)
        info_cell.value = f"총 세대수: {total_units}세대"
        info_cell.font = subtitle_font
        info_cell.alignment = left_align

        date_range = f"{chr(64+START_COL+5)}4:{chr(64+START_COL+7)}4"
        ws.merge_cells(date_range)
        date_cell = ws.cell(4, START_COL + 5)
        date_cell.value = f"작성일: {datetime.now():%Y. %m. %d}"
        date_cell.font = subtitle_font
        date_cell.alignment = right_align

        # 5행: 컬럼 헤더 (단일 세대 타입) - 테두리 추가, 배경 투명
        # 품목 (C5:D5)
        품목_range = f"{chr(64+START_COL)}5:{chr(64+START_COL+1)}5"
        ws.merge_cells(품목_range)
        ws.cell(5, START_COL).value = "품목"
        ws.cell(5, START_COL).font = header_font
        ws.cell(5, START_COL).alignment = center_align
        for i in range(START_COL, START_COL + 2):
            ws.cell(5, i).border = thin_border

        # 세대당 단가 (E5:G5)
        세대당_range = f"{chr(64+START_COL+2)}5:{chr(64+START_COL+4)}5"
        ws.merge_cells(세대당_range)
        ws.cell(5, START_COL + 2).value = "세대당 단가"
        ws.cell(5, START_COL + 2).font = header_font
        ws.cell(5, START_COL + 2).alignment = center_align
        for i in range(START_COL + 2, START_COL + 5):
            ws.cell(5, i).border = thin_border

        # 총 금액 (H5:J5)
        총금액_range = f"{chr(64+START_COL+5)}5:{chr(64+START_COL+7)}5"
        ws.merge_cells(총금액_range)
        ws.cell(5, START_COL + 5).value = f"총 금액 ({total_units}세대)"
        ws.cell(5, START_COL + 5).font = header_font
        ws.cell(5, START_COL + 5).alignment = center_align
        for i in range(START_COL + 5, START_COL + 8):
            ws.cell(5, i).border = thin_border

        # 6행: 세부 컬럼 헤더 (배경 투명)
        headers_6 = [
            "대분류",
            "사양 및 규격",
            "수량",
            "단가",
            "금액",
            "수량",
            "단가",
            "금액",
        ]
        for idx, header_text in enumerate(headers_6):
            cell = ws.cell(6, START_COL + idx)
            cell.value = header_text
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        # 컬럼 너비 설정 (가로로 넓게)
        ws.column_dimensions[chr(64 + START_COL)].width = 12  # 대분류
        ws.column_dimensions[chr(64 + START_COL + 1)].width = 38  # 사양 및 규격
        ws.column_dimensions[chr(64 + START_COL + 2)].width = 9  # 수량
        ws.column_dimensions[chr(64 + START_COL + 3)].width = 13  # 단가
        ws.column_dimensions[chr(64 + START_COL + 4)].width = 15  # 금액
        ws.column_dimensions[chr(64 + START_COL + 5)].width = 9  # 수량(총)
        ws.column_dimensions[chr(64 + START_COL + 6)].width = 13  # 단가(총)
        ws.column_dimensions[chr(64 + START_COL + 7)].width = 17  # 금액(총)

        # 데이터 행 작성
        row_num = 7
        current_category = None

        for idx, row_data in df.iterrows():
            품목 = str(row_data["품목"])
            사양 = str(row_data["사양 및 규격"])
            수량 = float(row_data["수량"])
            단가 = float(row_data["단가"])
            금액 = float(row_data["금액"])

            # 대분류 (품목이 바뀔 때만 표시)
            cell_a = ws.cell(row=row_num, column=START_COL)
            if 품목 != current_category:
                cell_a.value = 품목
                current_category = 품목
            else:
                cell_a.value = ""
            cell_a.font = data_font
            cell_a.alignment = left_align
            cell_a.border = thin_border

            # 사양 및 규격
            ws.cell(row=row_num, column=START_COL + 1).value = 사양
            ws.cell(row=row_num, column=START_COL + 1).font = data_font
            ws.cell(row=row_num, column=START_COL + 1).alignment = left_align
            ws.cell(row=row_num, column=START_COL + 1).border = thin_border

            # 세대당 단가 (C-E)
            ws.cell(row=row_num, column=START_COL + 2).value = 수량
            ws.cell(row=row_num, column=START_COL + 2).font = data_font
            ws.cell(row=row_num, column=START_COL + 2).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 2).border = thin_border
            ws.cell(row=row_num, column=START_COL + 2).number_format = "#,##0.##"

            ws.cell(row=row_num, column=START_COL + 3).value = 단가
            ws.cell(row=row_num, column=START_COL + 3).font = data_font
            ws.cell(row=row_num, column=START_COL + 3).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 3).border = thin_border
            ws.cell(row=row_num, column=START_COL + 3).number_format = "#,##0"

            ws.cell(row=row_num, column=START_COL + 4).value = 금액
            ws.cell(row=row_num, column=START_COL + 4).font = data_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            # 총 금액 (F-H) - 세대수 곱하기
            ws.cell(row=row_num, column=START_COL + 5).value = 수량 * total_units
            ws.cell(row=row_num, column=START_COL + 5).font = data_font
            ws.cell(row=row_num, column=START_COL + 5).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 5).border = thin_border
            ws.cell(row=row_num, column=START_COL + 5).number_format = "#,##0.##"

            ws.cell(row=row_num, column=START_COL + 6).value = 단가
            ws.cell(row=row_num, column=START_COL + 6).font = data_font
            ws.cell(row=row_num, column=START_COL + 6).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 6).border = thin_border
            ws.cell(row=row_num, column=START_COL + 6).number_format = "#,##0"

            ws.cell(row=row_num, column=START_COL + 7).value = 금액 * total_units
            ws.cell(row=row_num, column=START_COL + 7).font = data_font
            ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

            row_num += 1

        # 합계 행 (배경 투명)
        ws.cell(row=row_num, column=START_COL).value = "합계"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border

        ws.cell(row=row_num, column=START_COL + 1).value = "(V.A.T 별도)"
        ws.cell(row=row_num, column=START_COL + 1).font = header_font
        ws.cell(row=row_num, column=START_COL + 1).alignment = center_align
        ws.cell(row=row_num, column=START_COL + 1).border = thin_border

        # 세대당 합계
        for col in [START_COL + 2, START_COL + 3]:
            ws.cell(row=row_num, column=col).value = ""
            ws.cell(row=row_num, column=col).border = thin_border

        ws.cell(row=row_num, column=START_COL + 4).value = df["금액"].sum()
        ws.cell(row=row_num, column=START_COL + 4).font = header_font
        ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
        ws.cell(row=row_num, column=START_COL + 4).border = thin_border
        ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

        # 총 합계
        for col in [START_COL + 5, START_COL + 6]:
            ws.cell(row=row_num, column=col).value = ""
            ws.cell(row=row_num, column=col).border = thin_border

        ws.cell(row=row_num, column=START_COL + 7).value = (
            df["금액"].sum() * total_units
        )
        ws.cell(row=row_num, column=START_COL + 7).font = header_font
        ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
        ws.cell(row=row_num, column=START_COL + 7).border = thin_border
        ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"
        row_num += 1

        # ----------------------------
        # 생산관리비 카테고리별 합계 섹션
        # ----------------------------
        if category_subtotals and prod_mgmt_settings:
            row_num += 1  # 빈 행

            # 생산관리비 섹션 제목
            ws.merge_cells(start_row=row_num, start_column=START_COL, end_row=row_num, end_column=START_COL + 7)
            ws.cell(row=row_num, column=START_COL).value = "생산관리비 카테고리별 합계"
            ws.cell(row=row_num, column=START_COL).font = subtitle_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            row_num += 1

            # 헤더 행
            mgmt_headers = ["카테고리", "", "소계", "비율(%)", "생산관리비", "", "", "총계"]
            for idx, h in enumerate(mgmt_headers):
                cell = ws.cell(row=row_num, column=START_COL + idx)
                cell.value = h
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            row_num += 1

            # 카테고리별 데이터
            for cat_name, cat_info in prod_mgmt_settings.items():
                subtotal = category_subtotals.get(cat_name, 0)
                rate = cat_info.get("rate", 0.0)
                mgmt_fee = category_mgmt_fees.get(cat_name, 0) if category_mgmt_fees else 0
                cat_total = subtotal + mgmt_fee

                # 카테고리명 (2칸 병합)
                ws.merge_cells(start_row=row_num, start_column=START_COL, end_row=row_num, end_column=START_COL + 1)
                ws.cell(row=row_num, column=START_COL).value = cat_name
                ws.cell(row=row_num, column=START_COL).font = data_font
                ws.cell(row=row_num, column=START_COL).alignment = left_align
                ws.cell(row=row_num, column=START_COL).border = thin_border
                ws.cell(row=row_num, column=START_COL + 1).border = thin_border

                # 소계
                ws.cell(row=row_num, column=START_COL + 2).value = subtotal
                ws.cell(row=row_num, column=START_COL + 2).font = data_font
                ws.cell(row=row_num, column=START_COL + 2).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 2).border = thin_border
                ws.cell(row=row_num, column=START_COL + 2).number_format = "#,##0"

                # 비율
                ws.cell(row=row_num, column=START_COL + 3).value = rate
                ws.cell(row=row_num, column=START_COL + 3).font = data_font
                ws.cell(row=row_num, column=START_COL + 3).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 3).border = thin_border
                ws.cell(row=row_num, column=START_COL + 3).number_format = "0.0"

                # 생산관리비
                ws.cell(row=row_num, column=START_COL + 4).value = mgmt_fee
                ws.cell(row=row_num, column=START_COL + 4).font = data_font
                ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 4).border = thin_border
                ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

                # 빈 칸
                ws.cell(row=row_num, column=START_COL + 5).value = ""
                ws.cell(row=row_num, column=START_COL + 5).border = thin_border
                ws.cell(row=row_num, column=START_COL + 6).value = ""
                ws.cell(row=row_num, column=START_COL + 6).border = thin_border

                # 총계
                ws.cell(row=row_num, column=START_COL + 7).value = cat_total
                ws.cell(row=row_num, column=START_COL + 7).font = data_font
                ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
                ws.cell(row=row_num, column=START_COL + 7).border = thin_border
                ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

                row_num += 1

            # 생산관리비 합계 행
            ws.merge_cells(start_row=row_num, start_column=START_COL, end_row=row_num, end_column=START_COL + 1)
            ws.cell(row=row_num, column=START_COL).value = "생산관리비 합계"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=START_COL + 1).border = thin_border

            for col in [START_COL + 2, START_COL + 3]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            ws.cell(row=row_num, column=START_COL + 4).value = total_mgmt_fee
            ws.cell(row=row_num, column=START_COL + 4).font = header_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            for col in [START_COL + 5, START_COL + 6]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            ws.cell(row=row_num, column=START_COL + 7).value = ""
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            row_num += 1

            # 최종 총계 행
            row_num += 1
            ws.merge_cells(start_row=row_num, start_column=START_COL, end_row=row_num, end_column=START_COL + 3)
            ws.cell(row=row_num, column=START_COL).value = "최종 총계 (원가 + 생산관리비)"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            for col in range(START_COL + 1, START_COL + 4):
                ws.cell(row=row_num, column=col).border = thin_border

            # 세대당 최종
            ws.cell(row=row_num, column=START_COL + 4).value = final_total
            ws.cell(row=row_num, column=START_COL + 4).font = header_font
            ws.cell(row=row_num, column=START_COL + 4).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 4).border = thin_border
            ws.cell(row=row_num, column=START_COL + 4).number_format = "#,##0"

            for col in [START_COL + 5, START_COL + 6]:
                ws.cell(row=row_num, column=col).value = ""
                ws.cell(row=row_num, column=col).border = thin_border

            # 총 세대 최종
            ws.cell(row=row_num, column=START_COL + 7).value = final_total * total_units
            ws.cell(row=row_num, column=START_COL + 7).font = header_font
            ws.cell(row=row_num, column=START_COL + 7).alignment = right_align
            ws.cell(row=row_num, column=START_COL + 7).border = thin_border
            ws.cell(row=row_num, column=START_COL + 7).number_format = "#,##0"

        # BytesIO로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 바닥판 세대수 추출
    total_units = 1  # 기본값
    if floor_data:
        # floor_data 구조: {"inputs": {"units": N}, ...}
        inputs = floor_data.get("inputs", {})
        total_units = int(inputs.get("units", 1))

    xlsx_bytes = df_to_excel_bytes(
        est_df,
        total_units,
        category_subtotals=category_subtotals,
        category_mgmt_fees=category_mgmt_fees,
        prod_mgmt_settings=prod_mgmt_categories,
        total_mgmt_fee=total_mgmt_fee,
        final_total=final_total,
    )
    st.download_button(
        "📥 현재 세대 견적서 다운로드 (생산관리비 포함)",
        data=xlsx_bytes,
        file_name=f"욕실_원자재_세대당_단가내역_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # ----------------------------
    # 통합 엑셀 출력 (다중 세대 타입)
    # ----------------------------
    def create_integrated_excel(saved_quotations: List[Dict]) -> bytes:
        """LGE 창원 스마트파크 형식의 통합 엑셀 생성"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "세대당 원자재 단가내역"

        # 스타일 정의
        title_font = Font(name="맑은 고딕", size=16, bold=True)
        header_font = Font(name="맑은 고딕", size=9, bold=True)
        data_font = Font(name="맑은 고딕", size=9)
        small_font = Font(name="맑은 고딕", size=8)

        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        num_types = len(saved_quotations)
        if num_types == 0:
            return b""

        # 모든 품목/사양 조합 수집 (순서 유지)
        all_items = []
        seen = set()
        for q in saved_quotations:
            for row in q["rows"]:
                key = (row["품목"], row["사양 및 규격"])
                if key not in seen:
                    seen.add(key)
                    all_items.append(key)

        # 컬럼 구조 계산
        # 품목(1) + 사양(1) + [수량,단가,금액] × num_types + 비고(1)
        START_COL = 1
        SPEC_COL = 2
        DATA_START_COL = 3  # 첫 번째 세대 타입의 수량 컬럼

        # 1행: 타이틀
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + num_types * 3 + 1)
        ws.cell(1, 1).value = "욕실 원자재 세대당 단가 내역"
        ws.cell(1, 1).font = title_font
        ws.cell(1, 1).alignment = center_align
        ws.row_dimensions[1].height = 25

        # 4행: 총수량 및 작성일
        total_all_units = sum(q["units"] for q in saved_quotations)
        ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=2)
        ws.cell(4, 1).font = header_font
        ws.merge_cells(start_row=4, start_column=DATA_START_COL + num_types * 3 - 2,
                       end_row=4, end_column=DATA_START_COL + num_types * 3)
        date_col = DATA_START_COL + num_types * 3 - 2
        ws.cell(4, date_col).value = f"총수량: {total_all_units}개"
        ws.cell(4, date_col).font = header_font
        ws.cell(4, date_col).alignment = right_align

        # 5행: 세대 타입 헤더 (◎ 타입명 ◎ 형태)
        ws.cell(5, START_COL).value = "품목"
        ws.cell(5, START_COL).font = header_font
        ws.cell(5, START_COL).alignment = center_align
        ws.cell(5, START_COL).border = thin_border

        ws.cell(5, SPEC_COL).value = "사양 및 규격"
        ws.cell(5, SPEC_COL).font = header_font
        ws.cell(5, SPEC_COL).alignment = center_align
        ws.cell(5, SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            # 3컬럼 병합
            ws.merge_cells(start_row=5, start_column=col_start, end_row=5, end_column=col_start + 2)
            ws.cell(5, col_start).value = f"◎ {q['name']}"
            ws.cell(5, col_start).font = header_font
            ws.cell(5, col_start).alignment = center_align
            for c in range(col_start, col_start + 3):
                ws.cell(5, c).border = thin_border

        # 비고 컬럼
        remark_col = DATA_START_COL + num_types * 3
        ws.cell(5, remark_col).value = "(V.A.T 제외)"
        ws.cell(5, remark_col).font = small_font
        ws.cell(5, remark_col).alignment = center_align
        ws.cell(5, remark_col).border = thin_border

        # 6행: 규격 및 세대수
        ws.cell(6, START_COL).value = ""
        ws.cell(6, START_COL).border = thin_border
        ws.cell(6, SPEC_COL).value = ""
        ws.cell(6, SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            ws.merge_cells(start_row=6, start_column=col_start, end_row=6, end_column=col_start + 2)
            ws.cell(6, col_start).value = f"◎ 규격({q['spec']})  ◎ {q['units']}세대"
            ws.cell(6, col_start).font = small_font
            ws.cell(6, col_start).alignment = center_align
            for c in range(col_start, col_start + 3):
                ws.cell(6, c).border = thin_border

        ws.cell(6, remark_col).value = "비고"
        ws.cell(6, remark_col).font = header_font
        ws.cell(6, remark_col).alignment = center_align
        ws.cell(6, remark_col).border = thin_border

        # 7행: 수량/단가/금액 헤더
        ws.cell(7, START_COL).value = ""
        ws.cell(7, START_COL).border = thin_border
        ws.cell(7, SPEC_COL).value = ""
        ws.cell(7, SPEC_COL).border = thin_border

        for i in range(num_types):
            col_start = DATA_START_COL + i * 3
            ws.cell(7, col_start).value = "수량"
            ws.cell(7, col_start).font = header_font
            ws.cell(7, col_start).alignment = center_align
            ws.cell(7, col_start).border = thin_border

            ws.cell(7, col_start + 1).value = "단가"
            ws.cell(7, col_start + 1).font = header_font
            ws.cell(7, col_start + 1).alignment = center_align
            ws.cell(7, col_start + 1).border = thin_border

            ws.cell(7, col_start + 2).value = "금액"
            ws.cell(7, col_start + 2).font = header_font
            ws.cell(7, col_start + 2).alignment = center_align
            ws.cell(7, col_start + 2).border = thin_border

        ws.cell(7, remark_col).value = ""
        ws.cell(7, remark_col).border = thin_border

        # 데이터 행 작성
        row_num = 8
        current_category = None

        # 각 세대별 데이터를 딕셔너리로 변환 (빠른 조회용)
        type_data = []
        for q in saved_quotations:
            item_dict = {}
            for r in q["rows"]:
                key = (r["품목"], r["사양 및 규격"])
                item_dict[key] = r
            type_data.append(item_dict)

        for 품목, 사양 in all_items:
            # 품목 (카테고리 변경시만 표시)
            cell_cat = ws.cell(row=row_num, column=START_COL)
            if 품목 != current_category:
                cell_cat.value = 품목
                current_category = 품목
            else:
                cell_cat.value = ""
            cell_cat.font = data_font
            cell_cat.alignment = left_align
            cell_cat.border = thin_border

            # 사양 및 규격
            ws.cell(row=row_num, column=SPEC_COL).value = 사양
            ws.cell(row=row_num, column=SPEC_COL).font = data_font
            ws.cell(row=row_num, column=SPEC_COL).alignment = left_align
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            # 각 세대 타입별 수량/단가/금액
            for i, td in enumerate(type_data):
                col_start = DATA_START_COL + i * 3
                key = (품목, 사양)
                if key in td:
                    r = td[key]
                    qty = r.get("수량", 0) or 0
                    price = r.get("단가", 0) or 0
                    amount = r.get("금액", 0) or 0
                else:
                    qty, price, amount = 0, 0, 0

                ws.cell(row=row_num, column=col_start).value = qty if qty else 0
                ws.cell(row=row_num, column=col_start).font = data_font
                ws.cell(row=row_num, column=col_start).alignment = right_align
                ws.cell(row=row_num, column=col_start).border = thin_border
                ws.cell(row=row_num, column=col_start).number_format = "#,##0.##"

                ws.cell(row=row_num, column=col_start + 1).value = price if price else 0
                ws.cell(row=row_num, column=col_start + 1).font = data_font
                ws.cell(row=row_num, column=col_start + 1).alignment = right_align
                ws.cell(row=row_num, column=col_start + 1).border = thin_border
                ws.cell(row=row_num, column=col_start + 1).number_format = "#,##0"

                ws.cell(row=row_num, column=col_start + 2).value = amount if amount else 0
                ws.cell(row=row_num, column=col_start + 2).font = data_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            # 비고
            ws.cell(row=row_num, column=remark_col).value = ""
            ws.cell(row=row_num, column=remark_col).border = thin_border

            row_num += 1

        # 합계 행: 세트당 단가
        ws.cell(row=row_num, column=START_COL).value = "세트당 단가"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border
        ws.cell(row=row_num, column=SPEC_COL).value = ""
        ws.cell(row=row_num, column=SPEC_COL).border = thin_border

        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            ws.cell(row=row_num, column=col_start).value = 1
            ws.cell(row=row_num, column=col_start).font = header_font
            ws.cell(row=row_num, column=col_start).alignment = right_align
            ws.cell(row=row_num, column=col_start).border = thin_border

            ws.cell(row=row_num, column=col_start + 1).value = ""
            ws.cell(row=row_num, column=col_start + 1).border = thin_border

            ws.cell(row=row_num, column=col_start + 2).value = q["total"]
            ws.cell(row=row_num, column=col_start + 2).font = header_font
            ws.cell(row=row_num, column=col_start + 2).alignment = right_align
            ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

        ws.cell(row=row_num, column=remark_col).value = ""
        ws.cell(row=row_num, column=remark_col).border = thin_border
        row_num += 1

        # 세대 총 합계 행
        ws.cell(row=row_num, column=START_COL).value = "세대 총 합계"
        ws.cell(row=row_num, column=START_COL).font = header_font
        ws.cell(row=row_num, column=START_COL).alignment = center_align
        ws.cell(row=row_num, column=START_COL).border = thin_border
        ws.cell(row=row_num, column=SPEC_COL).value = ""
        ws.cell(row=row_num, column=SPEC_COL).border = thin_border

        grand_total = 0
        for i, q in enumerate(saved_quotations):
            col_start = DATA_START_COL + i * 3
            type_total = q["total"] * q["units"]
            grand_total += type_total

            ws.cell(row=row_num, column=col_start).value = q["units"]
            ws.cell(row=row_num, column=col_start).font = header_font
            ws.cell(row=row_num, column=col_start).alignment = right_align
            ws.cell(row=row_num, column=col_start).border = thin_border

            ws.cell(row=row_num, column=col_start + 1).value = ""
            ws.cell(row=row_num, column=col_start + 1).border = thin_border

            ws.cell(row=row_num, column=col_start + 2).value = type_total
            ws.cell(row=row_num, column=col_start + 2).font = header_font
            ws.cell(row=row_num, column=col_start + 2).alignment = right_align
            ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

        # 총 합계 표시
        ws.cell(row=row_num, column=remark_col).value = f"{grand_total:,.0f}"
        ws.cell(row=row_num, column=remark_col).font = header_font
        ws.cell(row=row_num, column=remark_col).alignment = right_align
        ws.cell(row=row_num, column=remark_col).border = thin_border
        row_num += 1

        # ----------------------------
        # 공사 원가 조정 및 관리비 섹션 (이미지 형식)
        # ----------------------------
        # 생산관리비 정보가 있는지 확인
        has_mgmt_fee = any(q.get("total_mgmt_fee", 0) > 0 or q.get("prod_mgmt_settings") for q in saved_quotations)

        if has_mgmt_fee:
            row_num += 2

            # 섹션 타이틀: ◎ 공사 원가 조정 및 관리비
            ws.merge_cells(start_row=row_num, start_column=START_COL, end_row=row_num, end_column=remark_col)
            ws.cell(row=row_num, column=START_COL).value = "◎ 공사 원가 조정 및 관리비"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = left_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            row_num += 1

            # 카테고리 그룹 정의 (이미지에 맞춘 구조)
            # 그룹: (그룹명, [(카테고리키, 표시명), ...])
            mgmt_groups = [
                ("회사 생산관리비", [
                    ("회사생산품(바닥판,욕조)", "바닥판, 욕조(20~25%)"),
                    ("회사생산품(천장판)", "천장판(15~20%)"),
                ]),
                ("명진 생산관리비", [
                    ("회사-명진(벽,PVE바닥판)", "PVE바닥판, 타일벽체(15~20%)"),
                ]),
                ("타사 구매품", [
                    ("타사(천장,바닥판,타일)", "바닥판, 타일(5~10%)"),
                    ("타사(도기,수전,기타)", "도기, 수전류, 기타(5~10%)"),
                ]),
            ]

            # 카테고리별 데이터를 first quotation 기준으로 수집
            def get_cat_data(q, cat_key):
                """주어진 견적에서 카테고리 데이터 추출"""
                cat_subtotals = q.get("category_subtotals", {})
                cat_mgmt_fees = q.get("category_mgmt_fees", {})
                prod_settings = q.get("prod_mgmt_settings", {})

                subtotal = cat_subtotals.get(cat_key, 0)
                mgmt_fee = cat_mgmt_fees.get(cat_key, 0)
                rate = 0
                if isinstance(prod_settings.get(cat_key), dict):
                    rate = prod_settings[cat_key].get("rate", 0)
                return subtotal, mgmt_fee, rate

            # 그룹별로 렌더링
            for group_name, categories in mgmt_groups:
                num_rows = len(categories)
                start_row_for_group = row_num

                for idx, (cat_key, display_name) in enumerate(categories):
                    # 첫 번째 행에만 그룹명 표시 (세로 병합)
                    if idx == 0:
                        if num_rows > 1:
                            ws.merge_cells(
                                start_row=start_row_for_group,
                                start_column=START_COL,
                                end_row=start_row_for_group + num_rows - 1,
                                end_column=START_COL
                            )
                        ws.cell(row=start_row_for_group, column=START_COL).value = group_name
                        ws.cell(row=start_row_for_group, column=START_COL).font = data_font
                        ws.cell(row=start_row_for_group, column=START_COL).alignment = center_align
                        # 병합된 셀의 테두리 설정
                        for r in range(start_row_for_group, start_row_for_group + num_rows):
                            ws.cell(row=r, column=START_COL).border = thin_border

                    # 카테고리 표시명
                    ws.cell(row=row_num, column=SPEC_COL).value = display_name
                    ws.cell(row=row_num, column=SPEC_COL).font = data_font
                    ws.cell(row=row_num, column=SPEC_COL).alignment = left_align
                    ws.cell(row=row_num, column=SPEC_COL).border = thin_border

                    # 각 세대 타입별 데이터
                    for i, q in enumerate(saved_quotations):
                        col_start = DATA_START_COL + i * 3
                        subtotal, mgmt_fee, rate = get_cat_data(q, cat_key)

                        # 비율(%)
                        ws.cell(row=row_num, column=col_start).value = f"{rate}%"
                        ws.cell(row=row_num, column=col_start).font = small_font
                        ws.cell(row=row_num, column=col_start).alignment = right_align
                        ws.cell(row=row_num, column=col_start).border = thin_border

                        # 빈 칸 (중간)
                        ws.cell(row=row_num, column=col_start + 1).value = ""
                        ws.cell(row=row_num, column=col_start + 1).border = thin_border

                        # 생산관리비
                        ws.cell(row=row_num, column=col_start + 2).value = mgmt_fee if mgmt_fee else ""
                        ws.cell(row=row_num, column=col_start + 2).font = data_font
                        ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                        ws.cell(row=row_num, column=col_start + 2).border = thin_border
                        if mgmt_fee:
                            ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

                    # 비고 (각 행별 합계 표시)
                    row_total = sum(get_cat_data(q, cat_key)[1] for q in saved_quotations)
                    ws.cell(row=row_num, column=remark_col).value = row_total if row_total else ""
                    ws.cell(row=row_num, column=remark_col).font = data_font
                    ws.cell(row=row_num, column=remark_col).alignment = right_align
                    ws.cell(row=row_num, column=remark_col).border = thin_border
                    if row_total:
                        ws.cell(row=row_num, column=remark_col).number_format = "#,##0"

                    row_num += 1

            # 설치(AS) 관리비(15% 고정) 행 - 영업관리비가 있는 경우
            has_sales_fee = any(q.get("sales_mgmt_enabled", False) for q in saved_quotations)
            if has_sales_fee:
                ws.cell(row=row_num, column=START_COL).value = "설치(AS) 관리비(15% 고정)"
                ws.cell(row=row_num, column=START_COL).font = data_font
                ws.cell(row=row_num, column=START_COL).alignment = left_align
                ws.cell(row=row_num, column=START_COL).border = thin_border
                ws.cell(row=row_num, column=SPEC_COL).value = ""
                ws.cell(row=row_num, column=SPEC_COL).border = thin_border

                for i, q in enumerate(saved_quotations):
                    col_start = DATA_START_COL + i * 3
                    sales_rate = q.get("sales_mgmt_rate", 0)
                    sales_fee = q.get("sales_mgmt_fee", 0) if q.get("sales_mgmt_enabled", False) else 0

                    ws.cell(row=row_num, column=col_start).value = f"{sales_rate}%" if sales_fee else ""
                    ws.cell(row=row_num, column=col_start).font = small_font
                    ws.cell(row=row_num, column=col_start).alignment = right_align
                    ws.cell(row=row_num, column=col_start).border = thin_border

                    ws.cell(row=row_num, column=col_start + 1).value = ""
                    ws.cell(row=row_num, column=col_start + 1).border = thin_border

                    ws.cell(row=row_num, column=col_start + 2).value = sales_fee if sales_fee else ""
                    ws.cell(row=row_num, column=col_start + 2).font = data_font
                    ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                    ws.cell(row=row_num, column=col_start + 2).border = thin_border
                    if sales_fee:
                        ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

                total_sales = sum(q.get("sales_mgmt_fee", 0) for q in saved_quotations if q.get("sales_mgmt_enabled", False))
                ws.cell(row=row_num, column=remark_col).value = total_sales if total_sales else ""
                ws.cell(row=row_num, column=remark_col).font = data_font
                ws.cell(row=row_num, column=remark_col).alignment = right_align
                ws.cell(row=row_num, column=remark_col).border = thin_border
                if total_sales:
                    ws.cell(row=row_num, column=remark_col).number_format = "#,##0"
                row_num += 1

            # 매입세 차이액(0~3%) 행 - 빈 행
            ws.cell(row=row_num, column=START_COL).value = "매입세 차이액(0~3%)"
            ws.cell(row=row_num, column=START_COL).font = data_font
            ws.cell(row=row_num, column=START_COL).alignment = left_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border
            for i in range(num_types):
                col_start = DATA_START_COL + i * 3
                ws.cell(row=row_num, column=col_start).value = "0%"
                ws.cell(row=row_num, column=col_start).font = small_font
                ws.cell(row=row_num, column=col_start).alignment = right_align
                ws.cell(row=row_num, column=col_start).border = thin_border
                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).value = ""
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
            ws.cell(row=row_num, column=remark_col).value = ""
            ws.cell(row=row_num, column=remark_col).border = thin_border
            row_num += 1

            # 소계 행
            ws.cell(row=row_num, column=START_COL).value = "소계"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            total_all_mgmt = 0
            total_all_sales = 0
            for i, q in enumerate(saved_quotations):
                col_start = DATA_START_COL + i * 3
                mgmt_fee = q.get("total_mgmt_fee", 0)
                sales_fee = q.get("sales_mgmt_fee", 0) if q.get("sales_mgmt_enabled", False) else 0
                subtotal = mgmt_fee + sales_fee
                total_all_mgmt += mgmt_fee
                total_all_sales += sales_fee

                ws.cell(row=row_num, column=col_start).value = ""
                ws.cell(row=row_num, column=col_start).border = thin_border

                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border

                ws.cell(row=row_num, column=col_start + 2).value = subtotal
                ws.cell(row=row_num, column=col_start + 2).font = header_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            ws.cell(row=row_num, column=remark_col).value = total_all_mgmt + total_all_sales
            ws.cell(row=row_num, column=remark_col).font = header_font
            ws.cell(row=row_num, column=remark_col).alignment = right_align
            ws.cell(row=row_num, column=remark_col).border = thin_border
            ws.cell(row=row_num, column=remark_col).number_format = "#,##0"
            row_num += 1

            # 총 금액 합계 행
            row_num += 1
            ws.cell(row=row_num, column=START_COL).value = "총 금액 합계"
            ws.cell(row=row_num, column=START_COL).font = header_font
            ws.cell(row=row_num, column=START_COL).alignment = center_align
            ws.cell(row=row_num, column=START_COL).border = thin_border
            ws.cell(row=row_num, column=SPEC_COL).value = ""
            ws.cell(row=row_num, column=SPEC_COL).border = thin_border

            final_grand_total = 0
            for i, q in enumerate(saved_quotations):
                col_start = DATA_START_COL + i * 3
                final_per_unit = q.get("final_total", q["total"])
                type_final_total = final_per_unit * q["units"]
                final_grand_total += type_final_total

                ws.cell(row=row_num, column=col_start).value = ""
                ws.cell(row=row_num, column=col_start).border = thin_border

                ws.cell(row=row_num, column=col_start + 1).value = ""
                ws.cell(row=row_num, column=col_start + 1).border = thin_border

                ws.cell(row=row_num, column=col_start + 2).value = final_per_unit
                ws.cell(row=row_num, column=col_start + 2).font = header_font
                ws.cell(row=row_num, column=col_start + 2).alignment = right_align
                ws.cell(row=row_num, column=col_start + 2).border = thin_border
                ws.cell(row=row_num, column=col_start + 2).number_format = "#,##0"

            # 녹색 배경의 최종 합계
            green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            ws.cell(row=row_num, column=remark_col).value = final_grand_total
            ws.cell(row=row_num, column=remark_col).font = header_font
            ws.cell(row=row_num, column=remark_col).alignment = right_align
            ws.cell(row=row_num, column=remark_col).border = thin_border
            ws.cell(row=row_num, column=remark_col).fill = green_fill
            ws.cell(row=row_num, column=remark_col).number_format = "#,##0"

        # 컬럼 너비 설정
        ws.column_dimensions[get_column_letter(START_COL)].width = 12
        ws.column_dimensions[get_column_letter(SPEC_COL)].width = 30
        for i in range(num_types):
            col_start = DATA_START_COL + i * 3
            ws.column_dimensions[get_column_letter(col_start)].width = 7      # 수량
            ws.column_dimensions[get_column_letter(col_start + 1)].width = 10  # 단가
            ws.column_dimensions[get_column_letter(col_start + 2)].width = 12  # 금액
        ws.column_dimensions[get_column_letter(remark_col)].width = 15

        # BytesIO로 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # 통합 엑셀 다운로드 버튼
    saved_list = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
    if saved_list and len(saved_list) >= 1:
        st.markdown("### 통합 견적서 다운로드")
        integrated_bytes = create_integrated_excel(saved_list)
        if integrated_bytes:
            st.download_button(
                "📥 통합 견적서 Excel 다운로드 (LGE 형식)",
                data=integrated_bytes,
                file_name=f"욕실_원자재_통합_단가내역_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
            )
            st.info(f"총 {len(saved_list)}개 세대 타입 포함")

if warnings:
    with st.expander("⚠️ 경고/참고", expanded=False):
        for w in warnings:
            st.warning(w)
