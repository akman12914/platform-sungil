# ERP 품목코드 자동 생성 프로그램
# 견적서에서 추출된 품목에 ERP 코드를 자동 부여

from common_styles import apply_common_styles, set_page_config
import auth

import io
import os
import re
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
import streamlit as st

# Session state keys
SAVED_QUOTATIONS_KEY = "saved_quotations"
FLOOR_RESULT_KEY = "floor_result"
WALL_RESULT_KEY = "wall_result"
CEIL_RESULT_KEY = "ceil_result"
ERP_MAPPING_KEY = "erp_item_mapping"
ERP_CODE_DB_KEY = "erp_code_db"

# L/R 구분이 필요한 품목 분류
LR_REQUIRED_ITEMS = {
    # 판넬류
    "바닥판": {"category": "판넬류", "description": "바닥판 (욕조 배수구 위치 기준)"},
    "벽판": {"category": "판넬류", "description": "벽판"},
    "천장판": {"category": "판넬류", "description": "천장판"},
    # 냉온수 배관
    "독립배관": {"category": "냉온수배관", "description": "독립배관"},
    "PB세대 세트배관": {"category": "냉온수배관", "description": "PB세대 세트배관"},
    "세대배관": {"category": "냉온수배관", "description": "세대 배관"},
    # 오픈수전함 (코너형/사각형 하위에 좌/우 구분)
    "오픈수전함": {"category": "오픈수전함", "description": "오픈수전함 (코너형/사각형 구분 후 좌/우)", "has_subtype": True},
    # 욕실 관련
    "PS욕실장": {"category": "욕실장", "description": "PS욕실장"},
    "슬라이딩 욕실장": {"category": "욕실장", "description": "슬라이딩 욕실장"},
    "욕실장": {"category": "욕실장", "description": "욕실장"},
}

# 파일 경로
ERP_CODE_FILE = "erp-docs/ERP코드_251113_(GRP바닥판, 내부자재 및 부속품,천장판,타일벽체).xlsx"
ERP_FORMAT_FILE = "erp-docs/프로젝트 관리(욕실사업)(S)-ERP 양식.xlsx"

set_page_config(page_title="ERP 품목코드 생성", layout="wide")
apply_common_styles()

auth.require_auth()


# ----------------------------
# 코드 분류 체계 로드
# ----------------------------
@st.cache_data(show_spinner=False, ttl=60)  # 60초 후 캐시 만료
def load_code_classification() -> Dict:
    """코드분류(최종) 시트에서 대분류/중분류/규격 코드 체계 로드

    시트 구조 (header=None):
    - 컬럼 0: 대분류명
    - 컬럼 1: 대분류코드
    - 컬럼 3: 중분류명
    - 컬럼 4: 중분류코드
    - 컬럼 6: 규격
    - 컬럼 7: 규격코드
    """
    try:
        df = pd.read_excel(ERP_CODE_FILE, sheet_name="코드분류(최종)", header=None)

        classification = {
            "대분류": {},  # 대분류명 -> 대분류코드
            "중분류": {},  # (대분류코드, 중분류명) -> 중분류코드
            "규격": {},    # (중분류코드, 규격명) -> 규격코드
            "중분류_검색": {},  # 중분류명 -> [(대분류코드, 중분류코드), ...]
            "규격_검색": {},    # 규격명 -> [(중분류코드, 규격코드), ...]
        }

        current_대분류 = None
        current_대분류코드 = None
        current_중분류 = None
        current_중분류코드 = None

        for idx, row in df.iterrows():
            # 헤더 행 건너뛰기 (첫 2행)
            if idx < 2:
                continue

            # 대분류 처리 (컬럼 0, 1)
            대분류 = row.iloc[0] if len(row) > 0 else None
            대분류코드 = row.iloc[1] if len(row) > 1 else None

            if pd.notna(대분류) and str(대분류).strip() not in ["품목", "대분류", ""]:
                current_대분류 = str(대분류).strip()
                if pd.notna(대분류코드):
                    current_대분류코드 = str(대분류코드).strip()
                    classification["대분류"][current_대분류] = current_대분류코드

            # 중분류 처리 (컬럼 3, 4)
            중분류 = row.iloc[3] if len(row) > 3 else None
            중분류코드 = row.iloc[4] if len(row) > 4 else None

            if pd.notna(중분류) and str(중분류).strip() not in ["품목", "중분류", "성형부 코드", ""]:
                current_중분류 = str(중분류).strip()
                if pd.notna(중분류코드):
                    current_중분류코드 = str(중분류코드).strip()
                    key = (current_대분류코드, current_중분류)
                    classification["중분류"][key] = current_중분류코드

                    # 중분류 검색용 인덱스
                    if current_중분류 not in classification["중분류_검색"]:
                        classification["중분류_검색"][current_중분류] = []
                    classification["중분류_검색"][current_중분류].append(
                        (current_대분류코드, current_중분류코드)
                    )

            # 규격 처리 (컬럼 6, 7)
            규격 = row.iloc[6] if len(row) > 6 else None
            규격코드 = row.iloc[7] if len(row) > 7 else None

            if pd.notna(규격) and str(규격).strip() not in ["품목", "규격", ""]:
                규격명 = str(규격).strip()
                if pd.notna(규격코드):
                    규격코드값 = str(규격코드).strip()
                    key = (current_중분류코드, 규격명)
                    classification["규격"][key] = 규격코드값

                    # 규격 검색용 인덱스
                    if 규격명 not in classification["규격_검색"]:
                        classification["규격_검색"][규격명] = []
                    classification["규격_검색"][규격명].append(
                        (current_중분류코드, 규격코드값)
                    )

        return classification
    except Exception as e:
        st.error(f"코드 분류 로드 실패: {e}")
        return {"대분류": {}, "중분류": {}, "규격": {}, "중분류_검색": {}, "규격_검색": {}}


@st.cache_data(show_spinner=False, ttl=60)  # 60초 후 캐시 만료
def load_existing_codes() -> pd.DataFrame:
    """기존 ERP 코드 목록 로드 (ERP매칭용 시트 - 최종 산출 양식과 동일)

    ERP매칭용 시트 컬럼 (최종 산출 양식과 동일):
    - 생성품목코드: ERP 코드 (예: GPFSB1419L)
    - 생성품목명: 중분류 값 (= ERP 품목명)
    - 대분류코드, 대분류
    - 중분류코드, 중분류
    - 규격코드, 규격
    """
    try:
        # 우선 ERP매칭용 시트 시도
        df = pd.read_excel(ERP_CODE_FILE, sheet_name="ERP매칭용")
        return df
    except Exception:
        # 폴백: 기존 시트 사용
        try:
            df = pd.read_excel(ERP_CODE_FILE, sheet_name="251113_바닥판,내부자재,부속품,천장판,벽체)")
            df.columns = [col.strip().replace(' ', '') for col in df.columns]
            # 기존 시트 컬럼을 ERP매칭용 형식에 맞게 변환
            df = df.rename(columns={"품목코드생성": "생성품목코드"})
            df["생성품목명"] = df["중분류"]
            return df
        except Exception as e:
            st.error(f"기존 코드 로드 실패: {e}")
            return pd.DataFrame()


def refresh_erp_matching_sheet() -> bool:
    """ERP매칭용 시트 갱신 - 251113 + 코드분류(최종) 병합"""
    try:
        from openpyxl import load_workbook

        # 1. 251113 시트 데이터 로드
        df_251113 = pd.read_excel(ERP_CODE_FILE, sheet_name="251113_바닥판,내부자재,부속품,천장판,벽체)")
        df_251113.columns = [col.strip().replace(' ', '') for col in df_251113.columns]

        # 생성품목코드 생성: 대분류코드 + 중분류코드 + 규격코드
        def generate_item_code(row):
            parts = []
            if pd.notna(row.get("대분류코드")):
                parts.append(str(row["대분류코드"]).strip())
            if pd.notna(row.get("중분류코드")):
                parts.append(str(row["중분류코드"]).strip())
            if pd.notna(row.get("규격코드")):
                parts.append(str(row["규격코드"]).strip())
            return "".join(parts)

        df_from_251113 = pd.DataFrame({
            "생성품목코드": df_251113.apply(generate_item_code, axis=1),
            "생성품목명": df_251113["중분류"],
            "대분류코드": df_251113["대분류코드"],
            "대분류": df_251113["대분류"],
            "중분류코드": df_251113["중분류코드"],
            "중분류": df_251113["중분류"],
            "규격코드": df_251113["규격코드"],
            "규격": df_251113["규격"],
        })

        # 2. 코드분류(최종) 시트 데이터 로드
        df_class = pd.read_excel(ERP_CODE_FILE, sheet_name="코드분류(최종)", header=None)
        rows = []
        current_대분류 = None
        current_대분류코드 = None
        current_중분류 = None
        current_중분류코드 = None

        for idx, row in df_class.iterrows():
            if idx < 2:
                continue

            대분류 = row.iloc[0] if len(row) > 0 else None
            대분류코드 = row.iloc[1] if len(row) > 1 else None
            if pd.notna(대분류) and str(대분류).strip() not in ["품목", "대분류", ""]:
                current_대분류 = str(대분류).strip()
                if pd.notna(대분류코드):
                    current_대분류코드 = str(대분류코드).strip()

            중분류 = row.iloc[3] if len(row) > 3 else None
            중분류코드 = row.iloc[4] if len(row) > 4 else None
            if pd.notna(중분류) and str(중분류).strip() not in ["품목", "중분류", "성형부 코드", ""]:
                current_중분류 = str(중분류).strip()
                if pd.notna(중분류코드):
                    current_중분류코드 = str(중분류코드).strip()

            규격 = row.iloc[6] if len(row) > 6 else None
            규격코드 = row.iloc[7] if len(row) > 7 else None

            if current_중분류코드 and pd.notna(규격) and str(규격).strip() not in ["품목", "규격", ""]:
                규격명 = str(규격).strip()
                규격코드값 = str(규격코드).strip() if pd.notna(규격코드) else ""
                생성품목코드 = f"{current_대분류코드 or ''}{current_중분류코드 or ''}{규격코드값}"
                rows.append({
                    "생성품목코드": 생성품목코드,
                    "생성품목명": current_중분류,
                    "대분류코드": current_대분류코드,
                    "대분류": current_대분류,
                    "중분류코드": current_중분류코드,
                    "중분류": current_중분류,
                    "규격코드": 규격코드값,
                    "규격": 규격명,
                })
            elif current_중분류코드 and (not pd.notna(규격) or str(규격).strip() in ["", "품목", "규격"]):
                existing = [r for r in rows if r["중분류코드"] == current_중분류코드 and r["규격"] == ""]
                if not existing and current_중분류:
                    생성품목코드 = f"{current_대분류코드 or ''}{current_중분류코드 or ''}"
                    rows.append({
                        "생성품목코드": 생성품목코드,
                        "생성품목명": current_중분류,
                        "대분류코드": current_대분류코드,
                        "대분류": current_대분류,
                        "중분류코드": current_중분류코드,
                        "중분류": current_중분류,
                        "규격코드": "",
                        "규격": "",
                    })

        df_classification = pd.DataFrame(rows)

        # 3. 두 데이터 병합 (중복 제거)
        df_combined = pd.concat([df_from_251113, df_classification], ignore_index=True)
        df_new = df_combined.drop_duplicates(subset=["생성품목코드"], keep="first")
        df_new = df_new.fillna("")
        df_new = df_new[df_new["생성품목코드"] != ""]

        # 4. 엑셀 파일에 새 시트 추가
        wb = load_workbook(ERP_CODE_FILE)
        if "ERP매칭용" in wb.sheetnames:
            del wb["ERP매칭용"]

        ws = wb.create_sheet("ERP매칭용")
        headers = list(df_new.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, row in enumerate(df_new.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # 컬럼 너비 조정
        column_widths = {"A": 20, "B": 30, "C": 12, "D": 18, "E": 12, "F": 25, "G": 12, "H": 20}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        wb.save(ERP_CODE_FILE)

        # 캐시 클리어
        load_existing_codes.clear()
        load_code_classification.clear()

        return True
    except Exception as e:
        st.error(f"시트 갱신 실패: {e}")
        return False


def get_erp_output_columns() -> List[str]:
    """ERP 출력 양식 컬럼 (구성사양 시트 기준)"""
    return [
        "순번", "구성수량", "수주발생수량", "생성품목코드", "생성품목명",
        "품목생성여부", "공장별품목생성여부", "대분류코드", "대분류",
        "중분류코드", "중분류", "규격코드", "규격", "주창고코드", "주창고",
        "품목계정", "조달구분", "단위", "공장코드", "공장명",
        "표준단가", "이동평균단가", "합계금액", "관리자", "비고"
    ]


# ----------------------------
# 품목코드 매칭 로직
# ----------------------------
def normalize_spec(spec: str) -> str:
    """규격 문자열 정규화"""
    if not spec:
        return ""
    spec = str(spec).strip().upper()
    # 공백, 특수문자 통일
    spec = re.sub(r'[×xX]', '*', spec)
    spec = re.sub(r'\s+', '', spec)
    return spec


def extract_dimensions(spec: str) -> Optional[Tuple[int, int]]:
    """규격에서 가로*세로 치수 추출 (예: 1500*2300 -> (1500, 2300))"""
    spec = normalize_spec(spec)
    match = re.search(r'(\d{3,4})\*(\d{3,4})', spec)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return None


def normalize_for_matching(text: str) -> str:
    """매칭용 문자열 정규화 - 괄호 내용 통일, 오타 수정 등"""
    if not text:
        return ""
    text = str(text).strip()

    # 괄호 통일: （）→ (), ［］→ []
    text = text.replace("（", "(").replace("）", ")")
    text = text.replace("［", "[").replace("］", "]")

    # 흔한 오타/변형 수정
    typo_fixes = {
        "내항균성": "내항균성",  # 정상
        "내향균성": "내항균성",  # 오타 수정
        "내황균성": "내항균성",  # 오타 수정
        "내한균성": "내항균성",  # 오타 수정
    }
    for wrong, correct in typo_fixes.items():
        text = text.replace(wrong, correct)

    return text


def calculate_similarity(s1: str, s2: str) -> float:
    """두 문자열의 유사도 계산 (정규화 후 비교)"""
    # 기본 정규화
    n1 = normalize_spec(s1)
    n2 = normalize_spec(s2)

    # 매칭용 추가 정규화
    m1 = normalize_for_matching(s1)
    m2 = normalize_for_matching(s2)

    # 두 가지 방식 중 높은 유사도 선택
    sim1 = SequenceMatcher(None, n1, n2).ratio()
    sim2 = SequenceMatcher(None, m1, m2).ratio()

    return max(sim1, sim2)


def _clean_value(val) -> str:
    """NaN이나 None을 빈 문자열로 변환"""
    if pd.isna(val) or val is None:
        return ""
    return str(val).strip()


def search_in_classification(
    품목: str,
    사양: str,
    classification: Dict,
    existing_codes: pd.DataFrame,
) -> Optional[Dict]:
    """코드분류(최종) 데이터에서 품목/사양 검색

    품목명 또는 사양에서 핵심 키워드를 추출하여 중분류와 매칭합니다.
    예: 품목='문세트', 사양='ABS 문짝' -> 사양에서 '문짝' 발견 -> 문짝(일반) 매칭
    """
    품목_clean = str(품목).strip()
    사양_clean = str(사양).strip()
    검색_텍스트 = f"{품목_clean} {사양_clean}"  # 품목+사양 합쳐서 검색

    중분류_검색 = classification.get("중분류_검색", {})

    # 0. 핵심 키워드 기반 매칭 (품목 또는 사양에서 검색)
    핵심_키워드_매핑 = {
        "문짝": ["문짝(일반)"],
        "포켓도어": ["문짝(포켓도어)"],
        "문틀": ["문틀(일반)"],
        "실리콘": ["실리콘(내항균성)"],
        "세면기": ["세면기"],
        "양변기": ["양변기"],
        "수전": ["수전"],
        "거울": ["거울"],
        "욕실장": ["욕실장"],
        "환풍기": ["환풍기"],
        "샤워기": ["샤워기"],
    }

    for 키워드, 대상_중분류목록 in 핵심_키워드_매핑.items():
        # 품목 또는 사양에서 키워드 검색
        if 키워드 in 검색_텍스트:
            for 대상_중분류 in 대상_중분류목록:
                for 중분류명, 코드목록 in 중분류_검색.items():
                    if 대상_중분류 in 중분류명 or 중분류명 in 대상_중분류:
                        대분류코드, 중분류코드 = 코드목록[0]
                        matches = existing_codes[existing_codes["중분류코드"] == 중분류코드]
                        if not matches.empty:
                            row = matches.iloc[0]
                            return {
                                "match_type": "exact",
                                "code": _clean_value(row.get("생성품목코드", "")),
                                "existing_code": _clean_value(row.get("생성품목코드", "")),
                                "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                                "대분류": _clean_value(row.get("대분류", "")),
                                "대분류코드": _clean_value(row.get("대분류코드", "")),
                                "중분류": _clean_value(row.get("중분류", "")),
                                "중분류코드": 중분류코드,
                                "규격": _clean_value(row.get("규격", "")),
                                "규격코드": _clean_value(row.get("규격코드", "")),
                            }

    # 1. 중분류명에서 품목 검색 (완전 일치에 가까운 경우)
    for 중분류명, 코드목록 in 중분류_검색.items():
        # 품목명이 중분류명에 포함되거나, 중분류명이 품목명에 포함되는 경우
        if 품목_clean in 중분류명 or 중분류명 in 품목_clean:
            대분류코드, 중분류코드 = 코드목록[0]
            matches = existing_codes[existing_codes["중분류코드"] == 중분류코드]
            if not matches.empty:
                row = matches.iloc[0]
                return {
                    "match_type": "exact",
                    "code": _clean_value(row.get("생성품목코드", "")),
                    "existing_code": _clean_value(row.get("생성품목코드", "")),
                    "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                    "대분류": _clean_value(row.get("대분류", "")),
                    "대분류코드": _clean_value(row.get("대분류코드", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                    "중분류코드": 중분류코드,
                    "규격": _clean_value(row.get("규격", "")),
                    "규격코드": _clean_value(row.get("규격코드", "")),
                }

        # 사양이 중분류명에 포함되거나, 중분류명이 사양에 포함되는 경우
        if 사양_clean in 중분류명 or 중분류명 in 사양_clean:
            대분류코드, 중분류코드 = 코드목록[0]
            matches = existing_codes[existing_codes["중분류코드"] == 중분류코드]
            if not matches.empty:
                row = matches.iloc[0]
                return {
                    "match_type": "exact",
                    "code": _clean_value(row.get("생성품목코드", "")),
                    "existing_code": _clean_value(row.get("생성품목코드", "")),
                    "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                    "대분류": _clean_value(row.get("대분류", "")),
                    "대분류코드": _clean_value(row.get("대분류코드", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                    "중분류코드": 중분류코드,
                    "규격": _clean_value(row.get("규격", "")),
                    "규격코드": _clean_value(row.get("규격코드", "")),
                }

    # 2. 규격명에서 사양 검색 (예: 직관50 -> 50A)
    규격_검색 = classification.get("규격_검색", {})

    # 숫자 추출하여 비교 (예: 직관50 -> 50, 50A -> 50)
    사양_숫자 = re.findall(r'\d+', 사양_clean)

    for 규격명, 코드목록 in 규격_검색.items():
        규격명_숫자 = re.findall(r'\d+', 규격명)

        # 숫자가 일치하는 경우 (예: 직관50 = 50A)
        if 사양_숫자 and 규격명_숫자 and 사양_숫자[0] == 규격명_숫자[0]:
            # 추가 조건: 품목명이나 사양에서 관련 키워드 확인
            # 예: 직관 = 파이프 관련
            파이프_키워드 = ["직관", "파이프", "배관", "관", "PVD", "엘보", "티", "니플"]
            if any(kw in 품목_clean or kw in 사양_clean for kw in 파이프_키워드):
                중분류코드, 규격코드값 = 코드목록[0]
                matches = existing_codes[
                    (existing_codes["중분류코드"] == 중분류코드) &
                    (existing_codes["규격코드"] == 규격코드값)
                ]
                if not matches.empty:
                    row = matches.iloc[0]
                    return {
                        "match_type": "exact",
                        "code": _clean_value(row.get("생성품목코드", "")),
                        "existing_code": _clean_value(row.get("생성품목코드", "")),
                        "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                        "대분류": _clean_value(row.get("대분류", "")),
                        "대분류코드": _clean_value(row.get("대분류코드", "")),
                        "중분류": _clean_value(row.get("중분류", "")),
                        "중분류코드": 중분류코드,
                        "규격": _clean_value(row.get("규격", "")),
                        "규격코드": _clean_value(규격코드값),
                    }

        # 규격명이 사양에 포함되거나 사양이 규격명에 포함되는 경우
        if 사양_clean in 규격명 or 규격명 in 사양_clean:
            중분류코드, 규격코드값 = 코드목록[0]
            matches = existing_codes[
                (existing_codes["중분류코드"] == 중분류코드) &
                (existing_codes["규격코드"] == 규격코드값)
            ]
            if not matches.empty:
                row = matches.iloc[0]
                return {
                    "match_type": "exact",
                    "code": _clean_value(row.get("생성품목코드", "")),
                    "existing_code": _clean_value(row.get("생성품목코드", "")),
                    "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                    "대분류": _clean_value(row.get("대분류", "")),
                    "대분류코드": _clean_value(row.get("대분류코드", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                    "중분류코드": 중분류코드,
                    "규격": _clean_value(row.get("규격", "")),
                    "규격코드": _clean_value(규격코드값),
                }

    # 3. 일반 부분 문자열 매칭
    for 중분류명, 코드목록 in 중분류_검색.items():
        핵심_키워드 = 품목_clean.replace(" ", "")
        if len(핵심_키워드) >= 2:
            if 핵심_키워드 in 중분류명.replace(" ", ""):
                대분류코드, 중분류코드 = 코드목록[0]
                matches = existing_codes[existing_codes["중분류코드"] == 중분류코드]
                if not matches.empty:
                    row = matches.iloc[0]
                    return {
                        "match_type": "similar",
                        "code": _clean_value(row.get("생성품목코드", "")),
                        "existing_code": _clean_value(row.get("생성품목코드", "")),
                        "기존품목명": _clean_value(row.get("생성품목명", row.get("중분류", ""))),
                        "similarity": 0.85,
                        "similar_item": {
                            "코드": _clean_value(row.get("생성품목코드", "")),
                            "규격": _clean_value(row.get("규격", "")),
                            "중분류": _clean_value(row.get("중분류", "")),
                            "대분류": _clean_value(row.get("대분류", "")),
                        },
                        "대분류": _clean_value(row.get("대분류", "")),
                        "대분류코드": _clean_value(row.get("대분류코드", "")),
                        "중분류": _clean_value(row.get("중분류", "")),
                        "중분류코드": 중분류코드,
                        "규격": _clean_value(row.get("규격", "")),
                        "규격코드": _clean_value(row.get("규격코드", "")),
                    }

    return None


def find_matching_code(
    품목: str,
    사양: str,
    existing_codes: pd.DataFrame,
    classification: Dict,
    threshold: float = 0.8,
    floor_spec_info: dict = None,
    wall_spec_info: dict = None,
    ceil_spec_info: dict = None,
    direction_info: dict = None,
) -> Dict:
    """
    품목+사양에 대해 ERP 코드 매칭

    Args:
        품목: 품목명
        사양: 사양 및 규격
        existing_codes: 기존 ERP 코드 데이터프레임
        classification: 코드 분류 체계
        threshold: 유사도 임계값
        floor_spec_info: 바닥판 규격 정보
        wall_spec_info: 벽판 규격 정보
        ceil_spec_info: 천장판 규격 정보
        direction_info: 품목별 방향 정보 (L/R)
            {
                "바닥판": "L" or "R",
                "벽판": "L" or "R",
                "천장판": "L" or "R",
                "독립배관": "L" or "R",
                "PB세대배관": "L" or "R",
                "오픈수전함_형태": "코너형" or "사각형",
                "오픈수전함": "L" or "R",
                "욕실장": "L" or "R",
            }

    Returns:
        {
            "match_type": "exact" | "similar" | "new" | "pending",
            "code": 생성된 코드,
            "existing_code": 기존 코드 (있으면),
            "similarity": 유사도 (similar인 경우),
            "similar_item": 유사 품목 정보 (similar인 경우),
            "대분류": str,
            "대분류코드": str,
            "중분류": str,
            "중분류코드": str,
            "규격": str,
            "규격코드": str,
        }
    """
    result = {
        "match_type": "new",
        "code": "",
        "existing_code": None,
        "기존품목명": "",  # 기존 ERP 품목명 (exact/similar 매칭 시 사용)
        "similarity": 0.0,
        "similar_item": None,
        "대분류": "",
        "대분류코드": "",
        "중분류": "",
        "중분류코드": "",
        "규격": "",
        "규격코드": "",
    }

    # 품목명 정규화
    품목_clean = str(품목).strip()
    사양_clean = str(사양).strip()
    사양_normalized = normalize_spec(사양_clean)

    # direction_info가 None이면 빈 딕셔너리로 초기화
    if direction_info is None:
        direction_info = {}

    # L/R 구분이 필요한 품목인지 확인하고 방향 코드 가져오기
    def get_item_direction(item_name: str) -> str:
        """품목명에 해당하는 방향 코드 반환 (L 또는 R)"""
        if "바닥판" in item_name:
            return direction_info.get("바닥판", "L")
        elif "벽판" in item_name or "벽체" in item_name:
            return direction_info.get("벽판", "L")
        elif "천장판" in item_name:
            return direction_info.get("천장판", "L")
        elif "독립배관" in item_name:
            return direction_info.get("독립배관", "L")
        elif "PB세대" in item_name or "세대배관" in item_name:
            return direction_info.get("PB세대배관", "L")
        elif "오픈수전함" in item_name or "수전함" in item_name:
            return direction_info.get("오픈수전함", "L")
        elif "욕실장" in item_name or "PS욕실장" in item_name or "슬라이딩" in item_name:
            return direction_info.get("욕실장", "L")
        return ""

    def filter_by_direction(codes_df: pd.DataFrame, direction_code: str, code_column: str = "생성품목코드") -> pd.DataFrame:
        """방향 코드(L/R)에 따라 데이터프레임 필터링"""
        if direction_code and not codes_df.empty:
            # 코드가 L 또는 R로 끝나는 경우 필터링
            mask = codes_df[code_column].str.endswith(direction_code, na=False)
            filtered = codes_df[mask]
            if not filtered.empty:
                return filtered
        return codes_df

    # 0. 코드분류(최종)에서 먼저 검색 (신규 추가)
    # ABS문짝, 직관50 등 코드분류에 있는 품목 우선 검색
    classification_match = search_in_classification(
        품목_clean, 사양_clean, classification, existing_codes
    )
    if classification_match:
        return classification_match

    # 1. 대분류 찾기
    대분류명 = None
    대분류코드 = None

    # 특수 매핑 (품목명 -> 대분류) - 엑셀의 실제 대분류명과 일치시킴
    special_mapping = {
        "바닥판": ("GRP바닥판", "GPF"),  # 기본은 GRP, 사양에 따라 변경
        "벽판": ("욕실타일벽체 세트", "BTWS"),
        "천장판": ("욕실천장판 ", "FT"),  # 주의: 공백 포함
        "타일": ("타일류", "MTL"),
        "도기류": ("도기 및 수전", "MPF"),
        "수전": ("도기 및 수전", "MPF"),
        "문세트": ("문세트(일반)", "MDSG"),
        "액세서리": ("액세사리", "MAC"),
        "공통자재": ("공통 및 부속자재", "MCA"),
        "냉온수배관": ("냉온수배관", "MCHWP"),
        "오배수배관": ("오수구배관", "MWP"),  # 실제 엑셀: 오수구배관
        "칸막이": ("칸막이", "MPA"),
        "환기류": ("공통 및 부속자재", "MCA"),
        "욕실등": ("공통 및 부속자재", "MCA"),
        "문틀규격": ("문세트(일반)", "MDSG"),
        "은경": ("액세사리", "MAC"),
        "욕실장": ("액세사리", "MAC"),
        "욕조": ("GRP바닥판", "GPF"),
    }

    # 품목명에서 대분류 매칭
    for key, (cat_name, cat_code) in special_mapping.items():
        if key in 품목_clean or 품목_clean == key:
            대분류명 = cat_name
            대분류코드 = cat_code
            break

    # 바닥판 재질에 따른 대분류 조정
    if 품목_clean == "바닥판" or "바닥판" in 품목_clean:
        사양_upper = 사양_clean.upper()
        if "FRP" in 사양_upper or "SMC" in 사양_upper:
            대분류명 = "FRP바닥판"
            대분류코드 = "FPF"
        elif "GRP" in 사양_upper:
            대분류명 = "GRP바닥판"
            대분류코드 = "GPF"
        elif "PP" in 사양_upper or "PE" in 사양_upper:
            대분류명 = "GRP바닥판"
            대분류코드 = "GPF"
        elif "PVE" in 사양_upper:
            대분류명 = "FRP바닥판"
            대분류코드 = "FPF"

    # 분류에 없으면 직접 검색
    if not 대분류명:
        for cat_name, cat_code in classification.get("대분류", {}).items():
            if cat_name.strip() in 품목_clean or 품목_clean in cat_name.strip():
                대분류명 = cat_name
                대분류코드 = cat_code
                break

    result["대분류"] = 대분류명 or 품목_clean
    result["대분류코드"] = 대분류코드 or ""

    # ============================================
    # 타일 특수 처리: 중분류로 매칭
    # ============================================
    if 품목_clean == "타일":
        # 벽타일 300*600, 바닥타일 300*300 등 중분류 매칭
        if "벽" in 사양_clean or "300*600" in 사양_clean or "300×600" in 사양_clean:
            # 중분류: 벽체용 타일 300*600
            result["중분류"] = "벽체용 타일 300*600"
            result["중분류코드"] = "MWT3060"
            result["규격"] = "견적용"
            result["규격코드"] = ""
            # 기존 코드 검색
            matches = existing_codes[existing_codes["중분류"] == "벽체용 타일 300*600"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(row.get("생성품목명", row.get("중분류", "")))
                return result
        elif "바닥" in 사양_clean or "300*300" in 사양_clean or "300×300" in 사양_clean:
            result["중분류"] = "바닥용 타일 300*300"
            result["중분류코드"] = "MFT3030"
            result["규격"] = ""
            result["규격코드"] = ""
            matches = existing_codes[existing_codes["중분류"] == "바닥용 타일 300*300"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(row.get("생성품목명", row.get("중분류", "")))
                return result
        elif "250*400" in 사양_clean or "250×400" in 사양_clean:
            result["중분류"] = "벽체용 타일 250*400"
            result["중분류코드"] = "MWT2540"
            matches = existing_codes[existing_codes["중분류"] == "벽체용 타일 250*400"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(row.get("생성품목명", row.get("중분류", "")))
                return result
        elif "200*200" in 사양_clean or "200×200" in 사양_clean:
            result["중분류"] = "바닥용 타일 200*200"
            result["중분류코드"] = "MFT2020"
            matches = existing_codes[existing_codes["중분류"] == "바닥용 타일 200*200"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(row.get("생성품목명", row.get("중분류", "")))
                return result

    # ============================================
    # 천장판 특수 처리: 바디/사이드/점검구 매칭
    # ============================================
    if 품목_clean == "천장판":
        # 대분류가 공백 포함되어 있을 수 있음
        대분류명 = "욕실천장판 "  # 공백 포함
        result["대분류"] = 대분류명

        # 천장판 데이터에서 검색
        천장판_codes = existing_codes[existing_codes["대분류"].str.strip() == "욕실천장판"]

        # 점검구
        if "점검구" in 사양_clean or "천공구" in 사양_clean:
            # 점검구 관련 코드 검색
            점검구_codes = 천장판_codes[천장판_codes["중분류"].str.contains("점검구", na=False)]
            if not 점검구_codes.empty:
                # 첫 번째 점검구 사용 (또는 사양에 맞는 것 선택)
                row = 점검구_codes.iloc[0]
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["중분류"] = _clean_value(row.get("중분류", ""))
                result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                result["규격"] = _clean_value(row.get("규격", ""))
                result["규격코드"] = _clean_value(row.get("규격코드", ""))
                return result

        # 바디판넬
        if "바디" in 사양_clean:
            바디_codes = 천장판_codes[천장판_codes["중분류"].str.contains("바디", na=False)]
            if not 바디_codes.empty:
                # 규격으로 매칭 시도
                dims = extract_dimensions(사양_clean)
                if dims:
                    for _, row in 바디_codes.iterrows():
                        existing_spec = _clean_value(row.get("규격", ""))
                        existing_dims = extract_dimensions(existing_spec)
                        if existing_dims == dims:
                            result["match_type"] = "exact"
                            result["code"] = _clean_value(row.get("생성품목코드", ""))
                            result["existing_code"] = result["code"]
                            result["중분류"] = _clean_value(row.get("중분류", ""))
                            result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                            result["규격"] = existing_spec
                            result["규격코드"] = _clean_value(row.get("규격코드", ""))
                            return result
                # 규격 매칭 실패시 첫 번째 바디 사용
                row = 바디_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.8
                result["similar_item"] = {
                    "코드": _clean_value(row.get("생성품목코드", "")),
                    "규격": _clean_value(row.get("규격", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                }
                result["중분류"] = _clean_value(row.get("중분류", ""))
                result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                return result

        # 사이드판넬
        if "사이드" in 사양_clean:
            사이드_codes = 천장판_codes[천장판_codes["중분류"].str.contains("사이드", na=False)]
            if not 사이드_codes.empty:
                dims = extract_dimensions(사양_clean)
                if dims:
                    for _, row in 사이드_codes.iterrows():
                        existing_spec = _clean_value(row.get("규격", ""))
                        existing_dims = extract_dimensions(existing_spec)
                        if existing_dims == dims:
                            result["match_type"] = "exact"
                            result["code"] = _clean_value(row.get("생성품목코드", ""))
                            result["existing_code"] = result["code"]
                            result["중분류"] = _clean_value(row.get("중분류", ""))
                            result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                            result["규격"] = existing_spec
                            result["규격코드"] = _clean_value(row.get("규격코드", ""))
                            return result
                row = 사이드_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.8
                result["similar_item"] = {
                    "코드": _clean_value(row.get("생성품목코드", "")),
                    "규격": _clean_value(row.get("규격", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                }
                result["중분류"] = _clean_value(row.get("중분류", ""))
                result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                return result

        # GRP천장판 등 일반 검색
        if "GRP" in 사양_clean.upper():
            # GRP 관련 찾기 - 일단 바디 중 하나 선택
            바디_codes = 천장판_codes[천장판_codes["중분류"].str.contains("바디", na=False)]
            if not 바디_codes.empty:
                row = 바디_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.7
                result["similar_item"] = {
                    "코드": _clean_value(row.get("생성품목코드", "")),
                    "규격": _clean_value(row.get("규격", "")),
                    "중분류": _clean_value(row.get("중분류", "")),
                }
                result["중분류"] = _clean_value(row.get("중분류", ""))
                return result

    # ============================================
    # 벽판 특수 처리: 규격 + LA/RA 매칭
    # ============================================
    if 품목_clean == "벽판" or "벽" in 품목_clean:
        대분류명 = "욕실타일벽체 세트"
        result["대분류"] = 대분류명
        result["대분류코드"] = "BTWS"

        벽체_codes = existing_codes[existing_codes["대분류"] == "욕실타일벽체 세트"]

        # wall_spec_info에서 규격 정보 가져오기
        if wall_spec_info:
            W = wall_spec_info.get("규격_W", 0)
            L = wall_spec_info.get("규격_L", 0)
            if W and L:
                # LA/RA 방향은 보류 (L/R 미구분)
                # 규격만으로 매칭 시도
                for _, row in 벽체_codes.iterrows():
                    existing_spec = _clean_value(row.get("규격", ""))
                    existing_dims = extract_dimensions(existing_spec)
                    if existing_dims and existing_dims == (W, L):
                        result["match_type"] = "exact"
                        result["code"] = _clean_value(row.get("생성품목코드", ""))
                        result["existing_code"] = result["code"]
                        result["중분류"] = _clean_value(row.get("중분류", ""))
                        result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                        result["규격"] = existing_spec
                        result["규격코드"] = _clean_value(row.get("규격코드", ""))
                        return result

        # 사양에서 치수 추출 시도
        dims = extract_dimensions(사양_clean)
        if dims:
            for _, row in 벽체_codes.iterrows():
                existing_spec = _clean_value(row.get("규격", ""))
                existing_dims = extract_dimensions(existing_spec)
                if existing_dims == dims:
                    result["match_type"] = "exact"
                    result["code"] = _clean_value(row.get("생성품목코드", ""))
                    result["existing_code"] = result["code"]
                    result["중분류"] = _clean_value(row.get("중분류", ""))
                    result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                    result["규격"] = existing_spec
                    result["규격코드"] = _clean_value(row.get("규격코드", ""))
                    return result

        # PU벽판인 경우 - 전체 데이터에서 유사 품목 검색
        if "PU" in 사양_clean.upper():
            # 벽체 코드에서 유사한 규격 검색
            best_sim = 0.0
            best_row = None
            for _, row in 벽체_codes.iterrows():
                existing_spec = _clean_value(row.get("규격", ""))
                sim = calculate_similarity(사양_clean, existing_spec)
                if sim > best_sim:
                    best_sim = sim
                    best_row = row
            if best_row is not None and best_sim >= 0.5:
                result["match_type"] = "similar"
                result["similarity"] = best_sim
                result["code"] = _clean_value(best_row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(best_row.get("생성품목명", best_row.get("중분류", "")))
                result["중분류"] = _clean_value(best_row.get("중분류", ""))
                result["중분류코드"] = _clean_value(best_row.get("중분류코드", ""))
                result["규격"] = _clean_value(best_row.get("규격", ""))
                result["규격코드"] = _clean_value(best_row.get("규격코드", ""))
                result["similar_item"] = {
                    "코드": result["code"],
                    "규격": result["규격"],
                    "중분류": result["중분류"],
                    "대분류": result["대분류"],
                }
                return result

    # 2. 기존 코드에서 완전 일치 검색 (일반)
    # L/R 방향 코드 가져오기
    item_direction = get_item_direction(품목_clean)

    if not existing_codes.empty and 대분류명:
        # 대분류가 일치하는 항목 필터 (공백 제거 비교)
        matches = existing_codes[existing_codes["대분류"].str.strip() == 대분류명.strip()]

        # L/R 방향이 있으면 해당 방향으로 필터링
        if item_direction:
            matches = filter_by_direction(matches, item_direction)

        for _, row in matches.iterrows():
            existing_spec = _clean_value(row.get("규격", ""))
            existing_normalized = normalize_spec(existing_spec)

            # 정규화된 규격으로 비교 (좌/우, L/R 제외하고 비교)
            existing_normalized_no_dir = re.sub(r'[좌우LR]$', '', existing_normalized)
            사양_normalized_no_dir = re.sub(r'[좌우LR]$', '', 사양_normalized)

            if existing_normalized_no_dir == 사양_normalized_no_dir:
                result["match_type"] = "exact"
                result["code"] = _clean_value(row.get("생성품목코드", ""))
                result["existing_code"] = result["code"]
                result["기존품목명"] = _clean_value(row.get("생성품목명", row.get("중분류", "")))  # 품목명 컬럼 우선
                result["대분류"] = _clean_value(row.get("대분류", ""))
                result["대분류코드"] = _clean_value(row.get("대분류코드", ""))
                result["중분류"] = _clean_value(row.get("중분류", ""))
                result["중분류코드"] = _clean_value(row.get("중분류코드", ""))
                result["규격"] = existing_spec
                result["규격코드"] = _clean_value(row.get("규격코드", ""))
                return result

    # 3. 유사 품목 검색 (3단계: 대분류 일치 → 전체 규격 → 전체 중분류)
    best_similarity = 0.0
    best_match = None

    # 3-1. 먼저 대분류가 일치하는 항목에서 검색
    if not existing_codes.empty and 대분류명:
        matches = existing_codes[existing_codes["대분류"].str.strip() == 대분류명.strip()]

        # L/R 방향이 있으면 해당 방향으로 필터링
        if item_direction:
            matches = filter_by_direction(matches, item_direction)

        for _, row in matches.iterrows():
            existing_spec = _clean_value(row.get("규격", ""))
            sim = calculate_similarity(사양_clean, existing_spec)

            if sim > best_similarity and sim >= threshold:
                best_similarity = sim
                best_match = row

    # 3-2. 대분류 매칭 실패 시, 전체 데이터에서 규격명으로 검색
    if best_match is None and not existing_codes.empty:
        # 오타 수정 적용한 사양으로 검색
        사양_corrected = normalize_for_matching(사양_clean)

        for _, row in existing_codes.iterrows():
            existing_spec = _clean_value(row.get("규격", ""))
            existing_corrected = normalize_for_matching(existing_spec)

            # 정규화 후 완전 일치 체크
            if 사양_corrected == existing_corrected:
                best_similarity = 1.0
                best_match = row
                break

            # 유사도 검사
            sim = calculate_similarity(사양_clean, existing_spec)
            if sim > best_similarity and sim >= threshold:
                best_similarity = sim
                best_match = row

    # 3-3. 규격 매칭 실패 시, 중분류명에서 유사 품목 검색
    if best_match is None and not existing_codes.empty:
        for _, row in existing_codes.iterrows():
            existing_중분류 = _clean_value(row.get("중분류", ""))

            # 중분류명에 품목명이나 사양이 포함되어 있는지 확인
            sim = calculate_similarity(사양_clean, existing_중분류)
            if sim > best_similarity and sim >= threshold * 0.9:  # 중분류는 임계값 낮춤
                best_similarity = sim
                best_match = row

    if best_match is not None:
        result["match_type"] = "similar"
        result["similarity"] = best_similarity
        result["code"] = _clean_value(best_match.get("생성품목코드", ""))
        result["existing_code"] = result["code"]
        result["기존품목명"] = _clean_value(best_match.get("생성품목명", best_match.get("중분류", "")))  # 품목명 컬럼 우선
        result["similar_item"] = {
            "코드": _clean_value(best_match.get("생성품목코드", "")),
            "규격": _clean_value(best_match.get("규격", "")),
            "중분류": _clean_value(best_match.get("중분류", "")),
            "대분류": _clean_value(best_match.get("대분류", "")),
        }
        result["대분류"] = _clean_value(best_match.get("대분류", "")) or result["대분류"]
        result["대분류코드"] = _clean_value(best_match.get("대분류코드", "")) or result["대분류코드"]
        result["중분류"] = _clean_value(best_match.get("중분류", ""))
        result["중분류코드"] = _clean_value(best_match.get("중분류코드", ""))
        result["규격"] = _clean_value(best_match.get("규격", ""))
        result["규격코드"] = _clean_value(best_match.get("규격코드", ""))

    # 4. 중분류 찾기 (사양에서 추론)
    if not result["중분류"]:
        # 바닥판 중분류 추론
        if "바닥판" in 품목_clean or 품목_clean == "바닥판":
            # 형상 유형 판단
            if "샤워" in 사양_clean or "SHT" in 사양_clean.upper():
                result["중분류"] = "사각형(샤워)"
                result["중분류코드"] = "SQSHT"
            elif "욕조" in 사양_clean or "SB" in 사양_clean.upper():
                result["중분류"] = "사각형(욕조)"
                result["중분류코드"] = "SB"
            elif "코너신형" in 사양_clean or "세면부" in 사양_clean:
                result["중분류"] = "코너신형 세면부"
                result["중분류코드"] = "NCNTWB"
            elif "코너" in 사양_clean:
                result["중분류"] = "코너형"
                result["중분류코드"] = "CNT"
            else:
                # 기본값: 사각형(욕조)
                result["중분류"] = "사각형(욕조)"
                result["중분류코드"] = "SB"

    # 5. 규격 코드 추출/생성
    if not result["규격코드"]:
        dims = extract_dimensions(사양_clean)
        if dims:
            w, h = dims
            # 규격 코드: 가로 앞2자리 + 세로 앞2자리
            규격코드 = f"{w // 100}{h // 100}"

            # 좌/우 방향 추가 - direction_info 우선, 없으면 사양에서 추출
            item_direction = get_item_direction(품목_clean)
            if item_direction:
                # direction_info에서 방향 정보가 있으면 사용
                규격코드 += item_direction
            elif "좌" in 사양_clean or ("L" in 사양_clean.upper() and "LA" not in 사양_clean.upper()):
                규격코드 += "L"
            elif "우" in 사양_clean or ("R" in 사양_clean.upper() and "RA" not in 사양_clean.upper()):
                규격코드 += "R"
            elif "LA" in 사양_clean.upper():
                규격코드 += "LA"
            elif "RA" in 사양_clean.upper():
                규격코드 += "RA"

            # 주거약자 표시
            if "주약" in 사양_clean or "주거약자" in 사양_clean:
                규격코드 += "1"

            result["규격"] = 사양_clean
            result["규격코드"] = 규격코드
        else:
            # 치수 추출 불가 -> 견적용
            result["규격"] = 사양_clean
            result["규격코드"] = "견적용"

    # 6. 최종 코드 생성
    if result["match_type"] == "new":
        code_parts = []
        if result["대분류코드"]:
            code_parts.append(result["대분류코드"])
        if result["중분류코드"]:
            code_parts.append(result["중분류코드"])
        if result["규격코드"]:
            code_parts.append(result["규격코드"])

        result["code"] = "".join(code_parts) if code_parts else f"NEW_{품목_clean[:3]}"

    return result


def generate_품목명(대분류: str, 중분류: str, 규격: str) -> str:
    """품목명 생성 (예: GRP바닥판 사각형(욕조) 1400*1900좌)"""
    parts = []
    if 대분류:
        parts.append(str(대분류).strip())
    if 중분류:
        parts.append(str(중분류).strip())
    if 규격:
        parts.append(str(규격).strip())
    return " ".join(parts)


def get_existing_품목명_from_row(row) -> str:
    """ERP 데이터 행에서 기존 품목명 추출 (중분류 = 품목명)"""
    return _clean_value(row.get("중분류", "")).strip()


def extract_floor_erp_spec(floor_result: dict, direction_override: str = None) -> dict:
    """바닥판 계산 결과에서 ERP 규격 정보 추출

    Args:
        floor_result: 바닥판 계산 결과
        direction_override: 사용자가 지정한 방향 ("좌" 또는 "우").
                           계산 결과에 방향이 없을 때 사용.
    """
    if not floor_result:
        return None

    inputs = floor_result.get("inputs", {})
    result = floor_result.get("result", {})

    # 재질 추출
    material = result.get("소재", "")
    material_clean = material.replace(" 바닥판", "").replace("바닥판", "").strip().upper()

    # 규격 추출 (W x L)
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)

    # 방향 추출 (좌/우) - 계산 결과 우선, 없으면 사용자 지정값 사용
    direction = inputs.get("direction", "")  # 'left' or 'right' or ''
    direction_kr = ""
    direction_code = ""

    if direction == "left" or direction == "좌":
        direction_kr = "좌"
        direction_code = "L"
    elif direction == "right" or direction == "우":
        direction_kr = "우"
        direction_code = "R"
    elif direction_override:
        # 계산 결과에 방향이 없으면 사용자 지정값 사용
        direction_kr = direction_override
        direction_code = "L" if direction_override == "좌" else "R"

    # 주거약자 여부
    is_senior = inputs.get("user_type", "") == "주거약자"

    # 형상 유형 (욕조/샤워 등)
    shape_type = inputs.get("shape_type", "욕조")  # 기본값 욕조

    # 규격 문자열 생성 (ERP 형식: 1400*1900좌)
    if W and L:
        spec_str = f"{W}*{L}{direction_kr}"
        if is_senior:
            spec_str += "(주약)"
    else:
        spec_str = material_clean

    # 단가
    단가 = result.get("소계", 0)

    return {
        "품목": "바닥판",
        "재질": material_clean,
        "규격_W": W,
        "규격_L": L,
        "방향": direction_kr,
        "방향코드": direction_code,
        "주거약자": is_senior,
        "형상": shape_type,
        "사양": spec_str,
        "단가": 단가,
        "수량": inputs.get("units", 1),
    }


def extract_wall_erp_spec(wall_result: dict) -> dict:
    """벽판 계산 결과에서 ERP 규격 정보 추출"""
    if not wall_result:
        return None

    inputs = wall_result.get("inputs", {})
    result = wall_result.get("result", {})
    counts = result.get("counts", {})

    # 규격 추출 (바닥판과 동일한 규격 사용)
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)

    # 방향
    direction = inputs.get("direction", "")
    direction_kr = ""
    if direction == "left" or direction == "좌":
        direction_kr = "좌"
    elif direction == "right" or direction == "우":
        direction_kr = "우"

    # 규격 문자열
    if W and L:
        spec_str = f"{W}*{L}{direction_kr}"
    else:
        spec_str = "PU벽판"

    return {
        "품목": "벽판",
        "규격_W": W,
        "규격_L": L,
        "방향": direction_kr,
        "사양": spec_str,
        "총개수": counts.get("n_panels", 0),
        "단가": result.get("소계", 0),
    }


def extract_ceiling_erp_spec(ceil_result: dict) -> dict:
    """천장판 계산 결과에서 ERP 규격 정보 추출"""
    if not ceil_result:
        return None

    inputs = ceil_result.get("inputs", {})
    result = ceil_result.get("result", {})

    # 재질
    material = inputs.get("material", "GRP")

    json_export = result.get("json_export", {})

    return {
        "품목": "천장판",
        "재질": material,
        "사양": f"{material}천장판",
        "총개수": json_export.get("총개수", 0),
        "단가": json_export.get("단가", 0),
    }


# ----------------------------
# UI
# ----------------------------
st.title("ERP 품목코드 자동 생성")

st.markdown("""
이 페이지는 견적서에서 추출된 품목들에 ERP 품목코드를 자동으로 부여합니다.

**처리 과정:**
1. 저장된 견적서에서 전체 품목 추출 (바닥판은 원본 계산 결과에서 규격 정보 사용)
2. 기존 ERP 코드와 매칭 (완전일치 → 유사품목 → 신규생성)
3. ERP 양식에 맞춰 엑셀 파일 출력
""")

# 파일 존재 확인
if not os.path.exists(ERP_CODE_FILE):
    st.error(f"ERP 코드 파일을 찾을 수 없습니다: {ERP_CODE_FILE}")
    st.stop()

if not os.path.exists(ERP_FORMAT_FILE):
    st.error(f"ERP 양식 파일을 찾을 수 없습니다: {ERP_FORMAT_FILE}")
    st.stop()

# 코드 분류 및 기존 코드 로드
with st.spinner("ERP 코드 데이터 로딩 중..."):
    classification = load_code_classification()
    existing_codes = load_existing_codes()

col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
with col1:
    st.metric("대분류 개수", len(classification.get("대분류", {})))
with col2:
    st.metric("중분류 개수", len(classification.get("중분류", {})))
with col3:
    st.metric("기존 코드 수", len(existing_codes))
with col4:
    st.markdown("<div style='height: 28px'></div>", unsafe_allow_html=True)
    if st.button("ERP매칭용 시트 갱신", help="251113 + 코드분류(최종) 데이터를 병합하여 ERP매칭용 시트를 다시 생성합니다"):
        with st.spinner("ERP매칭용 시트 갱신 중..."):
            if refresh_erp_matching_sheet():
                st.success("ERP매칭용 시트가 갱신되었습니다!")
                st.rerun()

st.markdown("---")

# ============================================
# 데이터 소스 선택: 엑셀 업로드 (기본) vs 저장된 견적
# ============================================
st.subheader("품목 데이터 입력")

data_source = st.radio(
    "데이터 소스 선택",
    options=["엑셀 파일 업로드 (원가내역서)", "저장된 견적 사용"],
    index=0,
    horizontal=True,
    help="원가내역서 엑셀 파일을 업로드하거나, 저장된 견적 데이터를 사용합니다."
)

# 엑셀에서 읽은 품목 데이터를 저장할 변수
uploaded_items = []

if data_source == "엑셀 파일 업로드 (원가내역서)":
    st.markdown("""
    **원가내역서 엑셀 파일**을 업로드해주세요.
    - 필수 컬럼: `품목`, `사양 및 규격` (또는 `규격`, `사양`)
    - 선택 컬럼: `수량`, `단가`
    """)

    uploaded_file = st.file_uploader(
        "엑셀 파일 선택 (.xlsx, .xls)",
        type=["xlsx", "xls"],
        key="cost_excel_uploader"
    )

    if uploaded_file is not None:
        try:
            # 엑셀 파일의 시트 목록 확인
            xl = pd.ExcelFile(uploaded_file)
            sheet_names = xl.sheet_names

            # 시트 선택
            if len(sheet_names) > 1:
                selected_sheet = st.selectbox(
                    "시트 선택",
                    options=sheet_names,
                    index=0
                )
            else:
                selected_sheet = sheet_names[0]

            # 선택된 시트 읽기
            df_upload = pd.read_excel(uploaded_file, sheet_name=selected_sheet)

            # 컬럼명 정규화 (공백 제거)
            df_upload.columns = [str(col).strip() for col in df_upload.columns]

            st.success(f"✅ 파일 로드 완료: {len(df_upload)}행")

            # Unnamed: 컬럼을 A열, B열 등으로 변환
            def get_col_letter(idx):
                """인덱스를 엑셀 열 문자로 변환 (0->A, 1->B, ...)"""
                result = ""
                while idx >= 0:
                    result = chr(idx % 26 + ord('A')) + result
                    idx = idx // 26 - 1
                return result

            col_rename_map = {}
            for i, col in enumerate(df_upload.columns):
                if str(col).startswith("Unnamed"):
                    col_rename_map[col] = f"{get_col_letter(i)}열"

            if col_rename_map:
                df_upload = df_upload.rename(columns=col_rename_map)

            valid_cols = list(df_upload.columns)

            # 컬럼 자동 감지
            품목_col = None
            for col in valid_cols:
                if col in ["품목", "품목명", "항목", "item", "Item"]:
                    품목_col = col
                    break

            사양_col = None
            for col in valid_cols:
                if col in ["사양 및 규격", "사양및규격", "규격", "사양", "spec", "Spec", "규격 및 사양"]:
                    사양_col = col
                    break

            수량_col = None
            for col in valid_cols:
                if col in ["수량", "qty", "Qty", "quantity", "Quantity"]:
                    수량_col = col
                    break

            단가_col = None
            for col in valid_cols:
                if col in ["단가", "price", "Price", "unit_price"]:
                    단가_col = col
                    break

            # 자동 감지 결과 표시
            if 품목_col and 사양_col:
                st.success(f"✅ 컬럼 자동 감지 완료: 품목=`{품목_col}`, 사양=`{사양_col}`" +
                          (f", 수량=`{수량_col}`" if 수량_col else "") +
                          (f", 단가=`{단가_col}`" if 단가_col else ""))
            else:
                # 자동 감지 실패 시 수동 선택
                st.warning("⚠️ 컬럼 자동 감지 실패. 아래에서 직접 선택해주세요.")
                available_cols = ["(선택 안 함)"] + valid_cols

                col1, col2 = st.columns(2)
                with col1:
                    품목_col = st.selectbox("품목 컬럼", options=available_cols, index=0)
                    if 품목_col == "(선택 안 함)":
                        품목_col = None
                with col2:
                    사양_col = st.selectbox("사양 및 규격 컬럼", options=available_cols, index=0)
                    if 사양_col == "(선택 안 함)":
                        사양_col = None

            # 컬럼 미리보기
            with st.expander("업로드된 데이터 미리보기", expanded=False):
                st.dataframe(df_upload.head(20), use_container_width=True)

            # 데이터 추출
            if 품목_col and 사양_col:
                for _, row in df_upload.iterrows():
                    품목 = str(row.get(품목_col, "")).strip() if pd.notna(row.get(품목_col)) else ""
                    사양 = str(row.get(사양_col, "")).strip() if pd.notna(row.get(사양_col)) else ""

                    if not 품목:
                        continue

                    수량 = 0
                    if 수량_col and pd.notna(row.get(수량_col)):
                        try:
                            수량 = float(row.get(수량_col, 0))
                        except (ValueError, TypeError):
                            수량 = 0

                    단가 = 0
                    if 단가_col and pd.notna(row.get(단가_col)):
                        try:
                            단가 = float(row.get(단가_col, 0))
                        except (ValueError, TypeError):
                            단가 = 0

                    uploaded_items.append({
                        "품목": 품목,
                        "사양 및 규격": 사양,
                        "수량": 수량,
                        "단가": 단가,
                    })

                st.info(f"📋 추출된 품목: **{len(uploaded_items)}개**")

                # 세션에 저장
                st.session_state["uploaded_cost_items"] = uploaded_items
            else:
                st.warning("품목과 사양 및 규격 컬럼을 선택해주세요.")

        except Exception as e:
            st.error(f"엑셀 파일 읽기 오류: {e}")
    else:
        # 이전에 업로드한 데이터가 있으면 사용
        if "uploaded_cost_items" in st.session_state:
            uploaded_items = st.session_state["uploaded_cost_items"]
            st.info(f"📋 이전에 업로드한 품목 데이터 사용: **{len(uploaded_items)}개**")

else:
    # 저장된 견적 사용
    saved_quotations = st.session_state.get(SAVED_QUOTATIONS_KEY, [])

    if not saved_quotations:
        st.warning("⚠️ 저장된 견적이 없습니다. 먼저 '견적서 생성' 페이지에서 견적을 저장하거나, 엑셀 파일을 업로드해주세요.")

        if st.button("견적서 생성 페이지로 이동"):
            st.switch_page("pages/4_견적서_생성.py")

        st.stop()

    # 저장된 견적 목록 표시
    st.markdown("**저장된 견적 목록**")

    quotation_df = pd.DataFrame([
        {
            "번호": i + 1,
            "타입명": q["name"],
            "규격": q["spec"],
            "세대수": q["units"],
            "품목수": len(q.get("rows", [])),
            "최종단가": f"{q.get('final_total', q['total']):,.0f}원",
        }
        for i, q in enumerate(saved_quotations)
    ])
    st.dataframe(quotation_df, use_container_width=True, hide_index=True)

# 원본 계산 결과에서 규격 정보 추출
floor_result = st.session_state.get(FLOOR_RESULT_KEY)
wall_result = st.session_state.get(WALL_RESULT_KEY)
ceil_result = st.session_state.get(CEIL_RESULT_KEY)

# ============================================
# 품목별 방향(L/R) 설정 - 감지된 품목만 표시
# ============================================

# 계산 결과에서 방향 정보 확인
_calc_directions = {}  # 품목명 -> direction value from calc
if floor_result:
    _fd = floor_result.get("inputs", {}).get("direction", "")
    if _fd in ["left", "right", "좌", "우"]:
        _calc_directions["바닥판"] = _fd
if wall_result:
    _wd = wall_result.get("inputs", {}).get("direction", "")
    if _wd in ["left", "right", "좌", "우"]:
        _calc_directions["벽판"] = _wd
if ceil_result:
    _cd = ceil_result.get("inputs", {}).get("direction", "")
    if _cd in ["left", "right", "좌", "우"]:
        _calc_directions["천장판"] = _cd

# 품목명 -> session_state key 매핑
_ITEM_TO_STATE_KEY = {
    "바닥판": "floor_direction_override",
    "벽판": "wall_direction_override",
    "천장판": "ceil_direction_override",
    "독립배관": "indep_pipe_direction",
    "PB세대 세트배관": "pb_pipe_direction",
    "세대배관": "pb_pipe_direction",
    "오픈수전함": "faucet_box_direction",
    "PS욕실장": "bathroom_cabinet_direction",
    "슬라이딩 욕실장": "bathroom_cabinet_direction",
    "욕실장": "bathroom_cabinet_direction",
}

# L/R 수량 지정을 위한 session state 초기화
if "lr_quantity_override" not in st.session_state:
    st.session_state["lr_quantity_override"] = {}

# 데이터 소스에서 L/R 구분 가능한 품목 감지
saved_quotations = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
_source_rows = []
if data_source == "엑셀 파일 업로드 (원가내역서)":
    _source_rows = uploaded_items
else:
    for q in saved_quotations:
        for row in q.get("rows", []):
            _source_rows.append(row)

detected_lr_items = []  # [{품목, 사양, 수량, category, description, has_subtype}]
_seen_items = set()
for _row in _source_rows:
    _품목 = _clean_value(_row.get("품목", "")).strip()
    if _품목 in LR_REQUIRED_ITEMS and _품목 not in _seen_items:
        _seen_items.add(_품목)
        _사양 = _clean_value(_row.get("사양 및 규격", "")).strip()
        _수량 = float(_row.get("수량", 0) or 0)
        detected_lr_items.append({
            "품목": _품목,
            "사양": _사양,
            "수량": _수량,
            **LR_REQUIRED_ITEMS[_품목],
        })

if detected_lr_items:
    st.markdown("---")
    st.subheader("감지된 품목의 좌우 설정")
    st.caption("L/R 구분이 필요한 품목이 감지되었습니다. 각 품목의 방향을 선택해주세요.")

    for idx, lr_item in enumerate(detected_lr_items):
        품목 = lr_item["품목"]
        사양 = lr_item["사양"]
        수량 = lr_item["수량"]
        has_subtype = lr_item.get("has_subtype", False)
        state_key = _ITEM_TO_STATE_KEY.get(품목)
        calc_dir = _calc_directions.get(품목)

        # 라벨 구성
        label = f"**{품목}**"
        if 사양:
            label += f"  ({사양})"

        st.markdown(label)

        # 오픈수전함: 형태 선택 + 방향 선택
        if has_subtype:
            cols = st.columns([1, 1, 2])
            with cols[0]:
                faucet_type = st.radio(
                    "형태",
                    options=["코너형", "사각형"],
                    index=0,
                    horizontal=True,
                    key=f"faucet_type_radio_{idx}",
                )
                st.session_state["faucet_box_type"] = faucet_type
            with cols[1]:
                if calc_dir:
                    dir_display = "좌 (L)" if calc_dir in ["left", "좌"] else "우 (R)"
                    st.success(f"계산 결과: {dir_display}")
                    st.session_state["faucet_box_direction"] = "좌" if calc_dir in ["left", "좌"] else "우"
                else:
                    direction_choice = st.radio(
                        "방향",
                        options=["좌 (L)", "우 (R)"],
                        index=0,
                        horizontal=True,
                        key=f"lr_dir_radio_{idx}",
                    )
                    st.session_state["faucet_box_direction"] = "좌" if "좌" in direction_choice else "우"
            with cols[2]:
                # 수량 분리
                key_base = f"{품목}_{사양}"
                use_separate = st.checkbox(
                    "좌/우 수량 별도 지정",
                    key=f"use_separate_{idx}",
                    value=key_base in st.session_state["lr_quantity_override"],
                )
                if use_separate:
                    sq1, sq2 = st.columns(2)
                    with sq1:
                        left_qty = st.number_input(
                            "좌 수량", min_value=0,
                            value=st.session_state["lr_quantity_override"].get(key_base, {}).get("L", int(수량 / 2)),
                            step=1, key=f"left_qty_{idx}",
                        )
                    with sq2:
                        right_qty = st.number_input(
                            "우 수량", min_value=0,
                            value=st.session_state["lr_quantity_override"].get(key_base, {}).get("R", int(수량 / 2)),
                            step=1, key=f"right_qty_{idx}",
                        )
                    st.session_state["lr_quantity_override"][key_base] = {"L": left_qty, "R": right_qty, "품목": 품목, "사양": 사양}
                    total = left_qty + right_qty
                    if total != 수량:
                        st.warning(f"좌우 합계({total})가 기본 수량({수량})과 다릅니다.")
                else:
                    if key_base in st.session_state["lr_quantity_override"]:
                        del st.session_state["lr_quantity_override"][key_base]
        else:
            # 일반 품목: 방향 선택 + 수량 분리
            cols = st.columns([1, 2])
            with cols[0]:
                if calc_dir:
                    dir_display = "좌 (L)" if calc_dir in ["left", "좌"] else "우 (R)"
                    st.success(f"계산 결과: {dir_display}")
                    direction_kr = "좌" if calc_dir in ["left", "좌"] else "우"
                else:
                    direction_choice = st.radio(
                        "방향",
                        options=["좌 (L)", "우 (R)"],
                        index=0,
                        horizontal=True,
                        key=f"lr_dir_radio_{idx}",
                    )
                    direction_kr = "좌" if "좌" in direction_choice else "우"

                # session_state에 저장
                if state_key:
                    st.session_state[state_key] = direction_kr

            with cols[1]:
                # 수량 분리
                key_base = f"{품목}_{사양}"
                use_separate = st.checkbox(
                    "좌/우 수량 별도 지정",
                    key=f"use_separate_{idx}",
                    value=key_base in st.session_state["lr_quantity_override"],
                )
                if use_separate:
                    sq1, sq2 = st.columns(2)
                    with sq1:
                        left_qty = st.number_input(
                            "좌 수량", min_value=0,
                            value=st.session_state["lr_quantity_override"].get(key_base, {}).get("L", int(수량 / 2)),
                            step=1, key=f"left_qty_{idx}",
                        )
                    with sq2:
                        right_qty = st.number_input(
                            "우 수량", min_value=0,
                            value=st.session_state["lr_quantity_override"].get(key_base, {}).get("R", int(수량 / 2)),
                            step=1, key=f"right_qty_{idx}",
                        )
                    st.session_state["lr_quantity_override"][key_base] = {"L": left_qty, "R": right_qty, "품목": 품목, "사양": 사양}
                    total = left_qty + right_qty
                    if total != 수량:
                        st.warning(f"좌우 합계({total})가 기본 수량({수량})과 다릅니다.")
                else:
                    if key_base in st.session_state["lr_quantity_override"]:
                        del st.session_state["lr_quantity_override"][key_base]

# 전체 품목 추출
st.markdown("---")
st.subheader("1단계: 전체 품목 추출")

# 사용자 지정 방향 가져오기
floor_direction_override = st.session_state.get("floor_direction_override", "좌")
wall_direction_override = st.session_state.get("wall_direction_override", "좌")
ceil_direction_override = st.session_state.get("ceil_direction_override", "좌")
indep_pipe_direction = st.session_state.get("indep_pipe_direction", "좌")
pb_pipe_direction = st.session_state.get("pb_pipe_direction", "좌")
faucet_box_type = st.session_state.get("faucet_box_type", "코너형")
faucet_box_direction = st.session_state.get("faucet_box_direction", "좌")
bathroom_cabinet_direction = st.session_state.get("bathroom_cabinet_direction", "좌")

# 방향 코드 변환 함수
def get_direction_code(direction_kr: str) -> str:
    """한글 방향을 코드로 변환 (좌->L, 우->R)"""
    return "L" if direction_kr == "좌" else "R"

# 원본 계산 결과 표시
with st.expander("원본 계산 결과 (바닥판/벽판/천장판)", expanded=False):
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**바닥판**")
        if floor_result:
            floor_spec = extract_floor_erp_spec(floor_result, floor_direction_override)
            if floor_spec:
                st.write(f"- 재질: {floor_spec['재질']}")
                st.write(f"- 규격: {floor_spec['규격_W']}×{floor_spec['규격_L']}")
                st.write(f"- 방향: {floor_spec['방향'] or '미지정'}")
                st.write(f"- ERP 사양: **{floor_spec['사양']}**")
        else:
            st.info("바닥판 계산 결과 없음")

    with col2:
        st.markdown("**벽판**")
        if wall_result:
            wall_spec = extract_wall_erp_spec(wall_result)
            if wall_spec:
                st.write(f"- 사양: {wall_spec['사양']}")
                st.write(f"- 총개수: {wall_spec['총개수']}장")
        else:
            st.info("벽판 계산 결과 없음")

    with col3:
        st.markdown("**천장판**")
        if ceil_result:
            ceil_spec = extract_ceiling_erp_spec(ceil_result)
            if ceil_spec:
                st.write(f"- 재질: {ceil_spec['재질']}")
                st.write(f"- 총개수: {ceil_spec['총개수']}장")
        else:
            st.info("천장판 계산 결과 없음")

# 모든 품목 추출 (엑셀 업로드 또는 저장된 견적에서)
all_items = {}  # key: (품목, 사양) -> value: {수량 합계, 단가 등}

# 바닥판은 원본 계산 결과에서 규격 정보를 가져옴 (사용자 지정 방향 적용)
floor_spec_info = extract_floor_erp_spec(floor_result, floor_direction_override) if floor_result else None

# 데이터 소스에 따라 품목 추출
if data_source == "엑셀 파일 업로드 (원가내역서)":
    # 엑셀에서 업로드한 데이터 사용
    source_items = uploaded_items
else:
    # 저장된 견적에서 품목 추출
    source_items = []
    saved_quotations = st.session_state.get(SAVED_QUOTATIONS_KEY, [])
    for q in saved_quotations:
        for row in q.get("rows", []):
            source_items.append(row)

for row in source_items:
    품목 = _clean_value(row.get("품목", "")).strip()
    사양 = _clean_value(row.get("사양 및 규격", "")).strip()
    수량 = float(row.get("수량", 0) or 0)
    단가 = float(row.get("단가", 0) or 0)

    if not 품목:
        continue

    # 바닥판인 경우 원본 계산 결과의 규격 사용 (저장된 견적 사용 시)
    if 품목 == "바닥판" and floor_spec_info and data_source != "엑셀 파일 업로드 (원가내역서)":
        # 사양이 재질만 있는 경우 (예: "FRP", "GRP") -> 전체 규격으로 교체
        if 사양 in ["GRP", "FRP", "SMC/FRP", "PP/PE", "PVE", "SMC", "PP", "PE"]:
            사양 = floor_spec_info["사양"]  # 예: "1500*2200좌"
            단가 = floor_spec_info["단가"]

    # L/R 수량 별도 지정 확인
    key_base = f"{품목}_{사양}"
    lr_override = st.session_state.get("lr_quantity_override", {}).get(key_base)

    if lr_override:
        # 좌/우로 분리된 경우
        left_qty = lr_override.get("L", 0)
        right_qty = lr_override.get("R", 0)

        # 좌 품목 추가
        if left_qty > 0:
            # 사양에 좌 표시 추가 (이미 없는 경우)
            사양_L = 사양
            if "좌" not in 사양_L and "L" not in 사양_L.upper():
                사양_L = f"{사양}좌"

            key_L = (품목, 사양_L)
            if key_L not in all_items:
                all_items[key_L] = {
                    "품목": 품목,
                    "사양": 사양_L,
                    "총수량": 0,
                    "단가": 단가,
                    "방향": "L"
                }
            all_items[key_L]["총수량"] += left_qty

        # 우 품목 추가
        if right_qty > 0:
            # 사양에 우 표시 추가 (이미 없는 경우)
            사양_R = 사양
            if "우" not in 사양_R and "R" not in 사양_R.upper():
                사양_R = f"{사양}우"

            key_R = (품목, 사양_R)
            if key_R not in all_items:
                all_items[key_R] = {
                    "품목": 품목,
                    "사양": 사양_R,
                    "총수량": 0,
                    "단가": 단가,
                    "방향": "R"
                }
            all_items[key_R]["총수량"] += right_qty
    else:
        # 기본 처리 (분리 안 함)
        key = (품목, 사양)
        if key not in all_items:
            all_items[key] = {
                "품목": 품목,
                "사양": 사양,
                "총수량": 0,
                "단가": 단가,
            }
        all_items[key]["총수량"] += 수량

items_list = list(all_items.values())

if items_list:
    # 품목 통계 표시
    total_items = len(items_list)
    lr_separated_items = sum(1 for item in items_list if item.get("방향") in ["L", "R"])
    normal_items = total_items - lr_separated_items

    st.success(f"✅ 총 **{total_items}개**의 고유 품목이 추출되었습니다.")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("전체 품목 수", total_items)
    with col2:
        st.metric("일반 품목", normal_items, help="L/R 분리 없이 처리된 품목")
    with col3:
        st.metric("L/R 분리 품목", lr_separated_items, help="좌/우로 분리된 품목 수")

    items_df = pd.DataFrame(items_list)
    st.dataframe(items_df, use_container_width=True, hide_index=True)
else:
    if data_source == "엑셀 파일 업로드 (원가내역서)":
        st.warning("⚠️ 품목 데이터가 없습니다. 엑셀 파일을 업로드하고 컬럼을 매핑해주세요.")
    else:
        st.warning("⚠️ 품목 데이터가 없습니다.")

# 품목코드 매칭
st.markdown("---")
st.subheader("2단계: 품목코드 매칭 및 생성")

similarity_threshold = st.slider(
    "유사도 임계값 (이 값 이상이면 유사 품목으로 판정)",
    min_value=0.5,
    max_value=1.0,
    value=0.8,
    step=0.05
)

if st.button("품목코드 매칭 실행", type="primary"):
    matching_results = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    # 원본 규격 정보 추출
    wall_spec_info = extract_wall_erp_spec(wall_result) if wall_result else None

    # 방향 정보 구성 (사용자가 선택한 L/R 방향)
    direction_info = {
        "바닥판": get_direction_code(floor_direction_override),
        "벽판": get_direction_code(wall_direction_override),
        "천장판": get_direction_code(ceil_direction_override),
        "독립배관": get_direction_code(indep_pipe_direction),
        "PB세대배관": get_direction_code(pb_pipe_direction),
        "오픈수전함_형태": faucet_box_type,
        "오픈수전함": get_direction_code(faucet_box_direction),
        "욕실장": get_direction_code(bathroom_cabinet_direction),
    }

    for i, item in enumerate(items_list):
        status_text.text(f"처리 중: {item['품목']} - {item['사양']}")

        result = find_matching_code(
            품목=item["품목"],
            사양=item["사양"],
            existing_codes=existing_codes,
            classification=classification,
            threshold=similarity_threshold,
            floor_spec_info=floor_spec_info,
            wall_spec_info=wall_spec_info,
            direction_info=direction_info,
        )

        result["품목"] = item["품목"]
        result["사양"] = item["사양"]
        result["수량"] = item["총수량"]
        result["단가"] = item["단가"]

        # 생성품목명 결정: 기존 코드 매칭 시 기존품목명 사용, 신규는 자동 생성
        if result["match_type"] in ["exact", "similar"] and result.get("기존품목명"):
            result["생성품목명"] = result["기존품목명"]
        else:
            result["생성품목명"] = generate_품목명(
                result["대분류"],
                result["중분류"],
                result["규격"]
            )

        matching_results.append(result)
        progress_bar.progress((i + 1) / len(items_list))

    status_text.text("매칭 완료!")

    # 결과 저장
    st.session_state[ERP_MAPPING_KEY] = matching_results

    # 통계
    exact_count = sum(1 for r in matching_results if r["match_type"] == "exact")
    similar_count = sum(1 for r in matching_results if r["match_type"] == "similar")
    new_count = sum(1 for r in matching_results if r["match_type"] == "new")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("완전 일치", exact_count, help="기존 ERP 코드 재사용")
    with col2:
        st.metric("유사 품목", similar_count, help="확인 필요")
    with col3:
        st.metric("신규 생성", new_count, help="새로운 코드 생성됨")

# 매칭 결과 표시
matching_results = st.session_state.get(ERP_MAPPING_KEY, [])

if matching_results:
    st.markdown("---")
    st.subheader("매칭 결과")

    # 탭으로 구분
    tab1, tab2, tab3 = st.tabs(["완전 일치", "유사 품목 (확인 필요)", "신규 생성"])

    with tab1:
        exact_results = [r for r in matching_results if r["match_type"] == "exact"]
        if exact_results:
            exact_df = pd.DataFrame([
                {
                    "품목": r["품목"],
                    "사양": r["사양"],
                    "ERP 코드": r["code"],
                    "대분류": r["대분류"],
                    "중분류": r["중분류"],
                    "구성수량": int(r["수량"]),
                    "수주발생수량": int(r["수량"]),
                }
                for r in exact_results
            ])
            exact_df = exact_df.fillna("")  # NaN 제거
            st.dataframe(exact_df, use_container_width=True, hide_index=True)
        else:
            st.info("완전 일치 품목이 없습니다.")

    with tab2:
        similar_results = [r for r in matching_results if r["match_type"] == "similar"]
        if similar_results:
            st.warning("아래 품목들은 유사한 기존 품목이 있습니다. 확인 후 코드를 선택하세요.")

            for i, r in enumerate(similar_results):
                with st.expander(f"🔍 {r['품목']} - {r['사양']} (유사도: {r['similarity']:.1%})"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**현재 품목:**")
                        st.write(f"- 사양: {r['사양']}")
                    with col2:
                        st.markdown("**유사 기존 품목:**")
                        if r["similar_item"]:
                            st.write(f"- 코드: {r['similar_item']['코드']}")
                            st.write(f"- 규격: {r['similar_item']['규격']}")
                            st.write(f"- 중분류: {r['similar_item']['중분류']}")

                    # 선택 옵션
                    choice = st.radio(
                        "코드 선택:",
                        options=["기존 코드 사용", "신규 코드 생성"],
                        key=f"similar_choice_{i}",
                        horizontal=True
                    )

                    if choice == "기존 코드 사용" and r["similar_item"]:
                        r["code"] = r["similar_item"]["코드"]
                        r["match_type"] = "exact"
        else:
            st.info("유사 품목이 없습니다.")

    with tab3:
        new_results = [r for r in matching_results if r["match_type"] == "new"]
        if new_results:
            new_df = pd.DataFrame([
                {
                    "품목": r["품목"],
                    "사양": r["사양"],
                    "생성 코드": r["code"],
                    "대분류": r["대분류"],
                    "중분류": r["중분류"],
                    "규격코드": r.get("규격코드", "") or "",
                    "구성수량": int(r["수량"]),
                    "수주발생수량": int(r["수량"]),
                }
                for r in new_results
            ])
            new_df = new_df.fillna("")  # NaN 제거
            st.dataframe(new_df, use_container_width=True, hide_index=True)
        else:
            st.info("신규 생성 품목이 없습니다.")

# 엑셀 출력
if matching_results:
    st.markdown("---")
    st.subheader("3단계: ERP 양식 엑셀 출력")

    # 프로젝트 정보 입력
    col1, col2 = st.columns(2)
    with col1:
        project_code = st.text_input("프로젝트 코드", value="B250000-00", placeholder="예: B250519-01")
    with col2:
        project_name = st.text_input("프로젝트명", value="", placeholder="예: 괴산미니복합타운 아파트")

    # 기본 설정
    col1, col2, col3 = st.columns(3)
    with col1:
        warehouse_code = st.selectbox("주창고", ["W7020 (욕실사업부_생산창고)"], index=0)
    with col2:
        factory_code = st.selectbox("공장", ["S1 (성일 김해공장)"], index=0)
    with col3:
        unit = st.selectbox("단위", ["EA", "SET", "M", "KG"], index=0)

    def create_erp_excel(results: List[Dict], project_code: str, project_name: str) -> bytes:
        """ERP 양식에 맞는 엑셀 파일 생성"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "구성사양"

        # 스타일 정의
        header_font = Font(name="맑은 고딕", size=10, bold=True)
        data_font = Font(name="맑은 고딕", size=9)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 프로젝트 정보 (상단)
        ws.merge_cells('A3:B3')
        ws['A3'] = "프로젝트 코드"
        ws['A3'].font = header_font
        ws['C3'] = project_code
        ws.merge_cells('C3:F3')
        ws['G3'] = project_name

        # 헤더 행 (row 5)
        headers = get_erp_output_columns()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 데이터 행
        for row_idx, r in enumerate(results, 6):
            # 데이터 매핑
            row_data = [
                row_idx - 5,  # 순번
                int(r.get("수량", 1)),  # 구성수량
                int(r.get("수량", 1)),  # 수주발생수량
                r.get("code", ""),  # 생성품목코드
                r.get("생성품목명", ""),  # 생성품목명
                "N" if r.get("match_type") in ["exact", "similar"] else "Y",  # 품목생성여부 (기존코드 있으면 N, 신규면 Y)
                "Y",  # 공장별품목생성여부
                r.get("대분류코드", ""),  # 대분류코드
                r.get("대분류", ""),  # 대분류
                r.get("중분류코드", ""),  # 중분류코드
                r.get("중분류", ""),  # 중분류
                r.get("규격코드", ""),  # 규격코드
                r.get("규격", ""),  # 규격
                "W7020",  # 주창고코드
                "욕실사업부_생산창고",  # 주창고
                "원자재",  # 품목계정
                "구매품",  # 조달구분
                "EA",  # 단위
                "S1",  # 공장코드
                "성일 김해공장",  # 공장명
                r.get("단가", 0),  # 표준단가
                0,  # 이동평균단가
                r.get("단가", 0) * r.get("수량", 1),  # 합계금액
                "",  # 관리자
                "견적용" if r.get("규격코드") == "견적용" else "",  # 비고
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col in [2, 3, 21, 22, 23]:  # 숫자 컬럼
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

        # 컬럼 너비 조정
        column_widths = [8, 10, 12, 22, 40, 12, 14, 10, 14, 10, 16, 12, 18, 10, 18, 10, 10, 8, 10, 14, 12, 12, 14, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = width

        # 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_matching_report(results: List[Dict]) -> str:
        """매칭 리포트 텍스트 생성"""
        lines = []
        lines.append("=" * 60)
        lines.append("ERP 품목코드 매칭 리포트")
        lines.append(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)
        lines.append("")

        # 통계
        exact_count = sum(1 for r in results if r["match_type"] == "exact")
        similar_count = sum(1 for r in results if r["match_type"] == "similar")
        new_count = sum(1 for r in results if r["match_type"] == "new")

        lines.append("[통계]")
        lines.append(f"- 전체 품목 수: {len(results)}")
        lines.append(f"- 완전 일치: {exact_count}")
        lines.append(f"- 유사 품목: {similar_count}")
        lines.append(f"- 신규 생성: {new_count}")
        lines.append("")

        # 유사 품목 상세
        similar_results = [r for r in results if r["match_type"] == "similar"]
        if similar_results:
            lines.append("[유사 품목 목록 - 확인 필요]")
            lines.append("-" * 60)
            for r in similar_results:
                lines.append(f"품목: {r['품목']}")
                lines.append(f"  현재 사양: {r['사양']}")
                if r["similar_item"]:
                    lines.append(f"  유사 품목 코드: {r['similar_item']['코드']}")
                    lines.append(f"  유사 품목 규격: {r['similar_item']['규격']}")
                lines.append(f"  유사도: {r['similarity']:.1%}")
                lines.append("")

        # 신규 생성 품목
        new_results = [r for r in results if r["match_type"] == "new"]
        if new_results:
            lines.append("[신규 생성 품목]")
            lines.append("-" * 60)
            for r in new_results:
                lines.append(f"품목: {r['품목']}")
                lines.append(f"  사양: {r['사양']}")
                lines.append(f"  생성코드: {r['code']}")
                lines.append(f"  대분류: {r['대분류']} ({r['대분류코드']})")
                lines.append(f"  중분류: {r['중분류']} ({r['중분류코드']})")
                lines.append(f"  규격: {r['규격']} ({r['규격코드']})")
                lines.append("")

        return "\n".join(lines)

    col1, col2 = st.columns(2)

    with col1:
        if st.button("ERP 엑셀 파일 생성", type="primary"):
            excel_bytes = create_erp_excel(matching_results, project_code, project_name)

            filename = f"품목코드매핑_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="📥 품목코드매핑.xlsx 다운로드",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col2:
        if st.button("매칭 리포트 생성"):
            report_text = create_matching_report(matching_results)

            filename = f"매칭_리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            st.download_button(
                label="📥 매칭_리포트.txt 다운로드",
                data=report_text.encode("utf-8"),
                file_name=filename,
                mime="text/plain"
            )

# 사이드바: 코드 분류 체계 조회
with st.sidebar:
    st.markdown("### 코드 분류 체계")

    with st.expander("대분류 목록"):
        for name, code in classification.get("대분류", {}).items():
            st.text(f"{code}: {name}")

    with st.expander("중분류 목록 (상위 20개)"):
        count = 0
        for (대분류코드, 중분류명), 중분류코드 in classification.get("중분류", {}).items():
            st.text(f"{중분류코드}: {중분류명}")
            count += 1
            if count >= 20:
                st.text("...")
                break

    st.markdown("---")
    st.markdown("### 기존 ERP 코드 검색")
    search_term = st.text_input("검색 (품목명, 규격)", placeholder="예: 1500*2200 또는 실리콘")
    if search_term and not existing_codes.empty:
        # 규격 또는 생성품목명(중분류)에서 검색
        mask = (
            existing_codes["규격"].str.contains(search_term, case=False, na=False) |
            existing_codes["생성품목명"].str.contains(search_term, case=False, na=False) |
            existing_codes["중분류"].str.contains(search_term, case=False, na=False)
        )
        search_results = existing_codes[mask]

        if not search_results.empty:
            st.dataframe(search_results[["생성품목코드", "생성품목명", "대분류", "규격"]].head(10))
        else:
            st.info("검색 결과 없음")

    st.markdown("---")
    st.markdown("### ERP매칭용 시트 관리")
    if st.button("시트 갱신", help="251113 시트 데이터로 ERP매칭용 시트를 다시 생성합니다"):
        with st.spinner("갱신 중..."):
            if refresh_erp_matching_sheet():
                st.success("ERP매칭용 시트가 갱신되었습니다!")
                st.rerun()
