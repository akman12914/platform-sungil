# ERP 품목코드 자동 생성 프로그램
# 견적서에서 추출된 품목에 ERP 코드를 자동 부여

from common_styles import apply_common_styles, set_page_config
import auth

import io
import os
import re
from typing import Dict, List, Optional, Tuple
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
import streamlit as st

# Session state keys
SAVED_QUOTATIONS_KEY = "saved_quotations"
FLOOR_RESULT_KEY = "floor_result"
WALL_RESULT_KEY = "wall_result"
CEIL_RESULT_KEY = "ceil_result"
ERP_MAPPING_KEY = "erp_item_mapping"
ERP_CODE_DB_KEY = "erp_code_db"

# 파일 경로
ERP_CODE_FILE = "erp-docs/ERP코드_251113_(GRP바닥판, 내부자재 및 부속품,천장판,타일벽체).xlsx"
ERP_FORMAT_FILE = "erp-docs/프로젝트 관리(욕실사업)(S)-ERP 양식.xlsx"

set_page_config(page_title="ERP 품목코드 생성", layout="wide")
apply_common_styles()

auth.require_auth()


# ----------------------------
# 코드 분류 체계 로드
# ----------------------------
@st.cache_data(show_spinner=False)
def load_code_classification() -> Dict:
    """코드분류(최종) 시트에서 대분류/중분류/규격 코드 체계 로드"""
    try:
        df = pd.read_excel(ERP_CODE_FILE, sheet_name="코드분류(최종)")

        classification = {
            "대분류": {},  # 품목명 -> 코드
            "중분류": {},  # (대분류코드, 품목명) -> 코드
            "규격": {},    # (중분류코드, 규격명) -> 코드
        }

        current_대분류 = None
        current_대분류코드 = None
        current_중분류 = None
        current_중분류코드 = None

        for _, row in df.iterrows():
            # 대분류 처리
            대분류 = row.get("대분류")
            대분류코드 = row.get("Unnamed: 1")

            if pd.notna(대분류) and str(대분류).strip() != "품목":
                current_대분류 = str(대분류).strip()
                if pd.notna(대분류코드):
                    current_대분류코드 = str(대분류코드).strip()
                    classification["대분류"][current_대분류] = current_대분류코드

            # 중분류 처리
            중분류 = row.get("중분류")
            중분류코드 = row.get("Unnamed: 4")

            if pd.notna(중분류) and str(중분류).strip() not in ["품목", "성형부 코드"]:
                current_중분류 = str(중분류).strip()
                if pd.notna(중분류코드):
                    current_중분류코드 = str(중분류코드).strip()
                    key = (current_대분류코드, current_중분류)
                    classification["중분류"][key] = current_중분류코드

            # 규격 처리
            규격 = row.get("규격")
            규격코드 = row.get("Unnamed: 7")

            if pd.notna(규격) and str(규격).strip() not in ["품목", ""]:
                규격명 = str(규격).strip()
                if pd.notna(규격코드):
                    규격코드값 = str(규격코드).strip()
                    key = (current_중분류코드, 규격명)
                    classification["규격"][key] = 규격코드값

        return classification
    except Exception as e:
        st.error(f"코드 분류 로드 실패: {e}")
        return {"대분류": {}, "중분류": {}, "규격": {}}


@st.cache_data(show_spinner=False)
def load_existing_codes() -> pd.DataFrame:
    """기존 ERP 코드 목록 로드 (251113 시트)"""
    try:
        df = pd.read_excel(ERP_CODE_FILE, sheet_name="251113_바닥판,내부자재,부속품,천장판,벽체)")
        # 컬럼명 정규화: "대분류 코드" -> "대분류코드"
        df.columns = [col.strip().replace(' ', '') for col in df.columns]
        return df
    except Exception as e:
        st.error(f"기존 코드 로드 실패: {e}")
        return pd.DataFrame()


def get_erp_output_columns() -> List[str]:
    """ERP 출력 양식 컬럼 (구성사양 시트 기준)"""
    return [
        "순번", "구성수량", "수주발생수량", "생성품목코드", "생성품목명",
        "품목생성여부", "공장별품목생성여부", "대분류코드", "대분류",
        "중분류코드", "중분류", "규격코드", "규격", "주창고코드", "주창고",
        "품목계정", "조달구분", "단위", "공장코드", "공장명",
        "표준단가", "이동평균단가", "합계금액", "관리자", "비고"
    ]


# ----------------------------
# 품목코드 매칭 로직
# ----------------------------
def normalize_spec(spec: str) -> str:
    """규격 문자열 정규화"""
    if not spec:
        return ""
    spec = str(spec).strip().upper()
    # 공백, 특수문자 통일
    spec = re.sub(r'[×xX]', '*', spec)
    spec = re.sub(r'\s+', '', spec)
    return spec


def extract_dimensions(spec: str) -> Optional[Tuple[int, int]]:
    """규격에서 가로*세로 치수 추출 (예: 1500*2300 -> (1500, 2300))"""
    spec = normalize_spec(spec)
    match = re.search(r'(\d{3,4})\*(\d{3,4})', spec)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return None


def calculate_similarity(s1: str, s2: str) -> float:
    """두 문자열의 유사도 계산"""
    return SequenceMatcher(None, normalize_spec(s1), normalize_spec(s2)).ratio()


def find_matching_code(
    품목: str,
    사양: str,
    existing_codes: pd.DataFrame,
    classification: Dict,
    threshold: float = 0.8,
    floor_spec_info: dict = None,
    wall_spec_info: dict = None,
    ceil_spec_info: dict = None,
) -> Dict:
    """
    품목+사양에 대해 ERP 코드 매칭

    Returns:
        {
            "match_type": "exact" | "similar" | "new" | "pending",
            "code": 생성된 코드,
            "existing_code": 기존 코드 (있으면),
            "similarity": 유사도 (similar인 경우),
            "similar_item": 유사 품목 정보 (similar인 경우),
            "대분류": str,
            "대분류코드": str,
            "중분류": str,
            "중분류코드": str,
            "규격": str,
            "규격코드": str,
        }
    """
    result = {
        "match_type": "new",
        "code": "",
        "existing_code": None,
        "similarity": 0.0,
        "similar_item": None,
        "대분류": "",
        "대분류코드": "",
        "중분류": "",
        "중분류코드": "",
        "규격": "",
        "규격코드": "",
    }

    # 품목명 정규화
    품목_clean = str(품목).strip()
    사양_clean = str(사양).strip()
    사양_normalized = normalize_spec(사양_clean)

    # 1. 대분류 찾기
    대분류명 = None
    대분류코드 = None

    # 특수 매핑 (품목명 -> 대분류) - 엑셀의 실제 대분류명과 일치시킴
    special_mapping = {
        "바닥판": ("GRP바닥판", "GPF"),  # 기본은 GRP, 사양에 따라 변경
        "벽판": ("욕실타일벽체 세트", "BTWS"),
        "천장판": ("욕실천장판 ", "FT"),  # 주의: 공백 포함
        "타일": ("타일류", "MTL"),
        "도기류": ("도기 및 수전", "MPF"),
        "수전": ("도기 및 수전", "MPF"),
        "문세트": ("문세트(일반)", "MDSG"),
        "액세서리": ("액세사리", "MAC"),
        "공통자재": ("공통 및 부속자재", "MCA"),
        "냉온수배관": ("냉온수배관", "MCHWP"),
        "오배수배관": ("오수구배관", "MWP"),  # 실제 엑셀: 오수구배관
        "칸막이": ("칸막이", "MPA"),
        "환기류": ("공통 및 부속자재", "MCA"),
        "욕실등": ("공통 및 부속자재", "MCA"),
        "문틀규격": ("문세트(일반)", "MDSG"),
        "은경": ("액세사리", "MAC"),
        "욕실장": ("액세사리", "MAC"),
        "욕조": ("GRP바닥판", "GPF"),
    }

    # 품목명에서 대분류 매칭
    for key, (cat_name, cat_code) in special_mapping.items():
        if key in 품목_clean or 품목_clean == key:
            대분류명 = cat_name
            대분류코드 = cat_code
            break

    # 바닥판 재질에 따른 대분류 조정
    if 품목_clean == "바닥판" or "바닥판" in 품목_clean:
        사양_upper = 사양_clean.upper()
        if "FRP" in 사양_upper or "SMC" in 사양_upper:
            대분류명 = "FRP바닥판"
            대분류코드 = "FPF"
        elif "GRP" in 사양_upper:
            대분류명 = "GRP바닥판"
            대분류코드 = "GPF"
        elif "PP" in 사양_upper or "PE" in 사양_upper:
            대분류명 = "GRP바닥판"
            대분류코드 = "GPF"
        elif "PVE" in 사양_upper:
            대분류명 = "FRP바닥판"
            대분류코드 = "FPF"

    # 분류에 없으면 직접 검색
    if not 대분류명:
        for cat_name, cat_code in classification.get("대분류", {}).items():
            if cat_name.strip() in 품목_clean or 품목_clean in cat_name.strip():
                대분류명 = cat_name
                대분류코드 = cat_code
                break

    result["대분류"] = 대분류명 or 품목_clean
    result["대분류코드"] = 대분류코드 or ""

    # ============================================
    # 타일 특수 처리: 중분류로 매칭
    # ============================================
    if 품목_clean == "타일":
        # 벽타일 300*600, 바닥타일 300*300 등 중분류 매칭
        if "벽" in 사양_clean or "300*600" in 사양_clean or "300×600" in 사양_clean:
            # 중분류: 벽체용 타일 300*600
            result["중분류"] = "벽체용 타일 300*600"
            result["중분류코드"] = "MWT3060"
            result["규격"] = "견적용"
            result["규격코드"] = ""
            # 기존 코드 검색
            matches = existing_codes[existing_codes["중분류"] == "벽체용 타일 300*600"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                return result
        elif "바닥" in 사양_clean or "300*300" in 사양_clean or "300×300" in 사양_clean:
            result["중분류"] = "바닥용 타일 300*300"
            result["중분류코드"] = "MFT3030"
            result["규격"] = ""
            result["규격코드"] = ""
            matches = existing_codes[existing_codes["중분류"] == "바닥용 타일 300*300"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                return result
        elif "250*400" in 사양_clean or "250×400" in 사양_clean:
            result["중분류"] = "벽체용 타일 250*400"
            result["중분류코드"] = "MWT2540"
            matches = existing_codes[existing_codes["중분류"] == "벽체용 타일 250*400"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                return result
        elif "200*200" in 사양_clean or "200×200" in 사양_clean:
            result["중분류"] = "바닥용 타일 200*200"
            result["중분류코드"] = "MFT2020"
            matches = existing_codes[existing_codes["중분류"] == "바닥용 타일 200*200"]
            if not matches.empty:
                row = matches.iloc[0]
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                return result

    # ============================================
    # 천장판 특수 처리: 바디/사이드/점검구 매칭
    # ============================================
    if 품목_clean == "천장판":
        # 대분류가 공백 포함되어 있을 수 있음
        대분류명 = "욕실천장판 "  # 공백 포함
        result["대분류"] = 대분류명

        # 천장판 데이터에서 검색
        천장판_codes = existing_codes[existing_codes["대분류"].str.strip() == "욕실천장판"]

        # 점검구
        if "점검구" in 사양_clean or "천공구" in 사양_clean:
            # 점검구 관련 코드 검색
            점검구_codes = 천장판_codes[천장판_codes["중분류"].str.contains("점검구", na=False)]
            if not 점검구_codes.empty:
                # 첫 번째 점검구 사용 (또는 사양에 맞는 것 선택)
                row = 점검구_codes.iloc[0]
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                result["중분류"] = str(row.get("중분류", ""))
                result["중분류코드"] = str(row.get("중분류코드", ""))
                result["규격"] = str(row.get("규격", ""))
                result["규격코드"] = str(row.get("규격코드", ""))
                return result

        # 바디판넬
        if "바디" in 사양_clean:
            바디_codes = 천장판_codes[천장판_codes["중분류"].str.contains("바디", na=False)]
            if not 바디_codes.empty:
                # 규격으로 매칭 시도
                dims = extract_dimensions(사양_clean)
                if dims:
                    for _, row in 바디_codes.iterrows():
                        existing_spec = str(row.get("규격", ""))
                        existing_dims = extract_dimensions(existing_spec)
                        if existing_dims == dims:
                            result["match_type"] = "exact"
                            result["code"] = str(row.get("품목코드생성", ""))
                            result["existing_code"] = result["code"]
                            result["중분류"] = str(row.get("중분류", ""))
                            result["중분류코드"] = str(row.get("중분류코드", ""))
                            result["규격"] = existing_spec
                            result["규격코드"] = str(row.get("규격코드", ""))
                            return result
                # 규격 매칭 실패시 첫 번째 바디 사용
                row = 바디_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.8
                result["similar_item"] = {
                    "코드": str(row.get("품목코드생성", "")),
                    "규격": str(row.get("규격", "")),
                    "중분류": str(row.get("중분류", "")),
                }
                result["중분류"] = str(row.get("중분류", ""))
                result["중분류코드"] = str(row.get("중분류코드", ""))
                return result

        # 사이드판넬
        if "사이드" in 사양_clean:
            사이드_codes = 천장판_codes[천장판_codes["중분류"].str.contains("사이드", na=False)]
            if not 사이드_codes.empty:
                dims = extract_dimensions(사양_clean)
                if dims:
                    for _, row in 사이드_codes.iterrows():
                        existing_spec = str(row.get("규격", ""))
                        existing_dims = extract_dimensions(existing_spec)
                        if existing_dims == dims:
                            result["match_type"] = "exact"
                            result["code"] = str(row.get("품목코드생성", ""))
                            result["existing_code"] = result["code"]
                            result["중분류"] = str(row.get("중분류", ""))
                            result["중분류코드"] = str(row.get("중분류코드", ""))
                            result["규격"] = existing_spec
                            result["규격코드"] = str(row.get("규격코드", ""))
                            return result
                row = 사이드_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.8
                result["similar_item"] = {
                    "코드": str(row.get("품목코드생성", "")),
                    "규격": str(row.get("규격", "")),
                    "중분류": str(row.get("중분류", "")),
                }
                result["중분류"] = str(row.get("중분류", ""))
                result["중분류코드"] = str(row.get("중분류코드", ""))
                return result

        # GRP천장판 등 일반 검색
        if "GRP" in 사양_clean.upper():
            # GRP 관련 찾기 - 일단 바디 중 하나 선택
            바디_codes = 천장판_codes[천장판_codes["중분류"].str.contains("바디", na=False)]
            if not 바디_codes.empty:
                row = 바디_codes.iloc[0]
                result["match_type"] = "similar"
                result["similarity"] = 0.7
                result["similar_item"] = {
                    "코드": str(row.get("품목코드생성", "")),
                    "규격": str(row.get("규격", "")),
                    "중분류": str(row.get("중분류", "")),
                }
                result["중분류"] = str(row.get("중분류", ""))
                return result

    # ============================================
    # 벽판 특수 처리: 규격 + LA/RA 매칭
    # ============================================
    if 품목_clean == "벽판" or "벽" in 품목_clean:
        대분류명 = "욕실타일벽체 세트"
        result["대분류"] = 대분류명
        result["대분류코드"] = "BTWS"

        벽체_codes = existing_codes[existing_codes["대분류"] == "욕실타일벽체 세트"]

        # wall_spec_info에서 규격 정보 가져오기
        if wall_spec_info:
            W = wall_spec_info.get("규격_W", 0)
            L = wall_spec_info.get("규격_L", 0)
            if W and L:
                # LA/RA 방향은 보류 (L/R 미구분)
                # 규격만으로 매칭 시도
                for _, row in 벽체_codes.iterrows():
                    existing_spec = str(row.get("규격", ""))
                    existing_dims = extract_dimensions(existing_spec)
                    if existing_dims and existing_dims == (W, L):
                        result["match_type"] = "exact"
                        result["code"] = str(row.get("품목코드생성", ""))
                        result["existing_code"] = result["code"]
                        result["중분류"] = str(row.get("중분류", ""))
                        result["중분류코드"] = str(row.get("중분류코드", ""))
                        result["규격"] = existing_spec
                        result["규격코드"] = str(row.get("규격코드", ""))
                        return result

        # 사양에서 치수 추출 시도
        dims = extract_dimensions(사양_clean)
        if dims:
            for _, row in 벽체_codes.iterrows():
                existing_spec = str(row.get("규격", ""))
                existing_dims = extract_dimensions(existing_spec)
                if existing_dims == dims:
                    result["match_type"] = "exact"
                    result["code"] = str(row.get("품목코드생성", ""))
                    result["existing_code"] = result["code"]
                    result["중분류"] = str(row.get("중분류", ""))
                    result["중분류코드"] = str(row.get("중분류코드", ""))
                    result["규격"] = existing_spec
                    result["규격코드"] = str(row.get("규격코드", ""))
                    return result

        # PU벽판인 경우 - 규격 정보 필요
        if "PU" in 사양_clean.upper():
            result["중분류"] = "사각형(샤워타입)"
            result["중분류코드"] = "SQSHT"
            result["match_type"] = "pending"  # L/R 미정
            result["규격"] = "L/R 선택 필요"
            result["규격코드"] = "견적용"

    # 2. 기존 코드에서 완전 일치 검색 (일반)
    if not existing_codes.empty and 대분류명:
        # 대분류가 일치하는 항목 필터 (공백 제거 비교)
        matches = existing_codes[existing_codes["대분류"].str.strip() == 대분류명.strip()]

        for _, row in matches.iterrows():
            existing_spec = str(row.get("규격", ""))
            existing_normalized = normalize_spec(existing_spec)

            # 정규화된 규격으로 비교
            if existing_normalized == 사양_normalized:
                result["match_type"] = "exact"
                result["code"] = str(row.get("품목코드생성", ""))
                result["existing_code"] = result["code"]
                result["중분류"] = str(row.get("중분류", ""))
                result["중분류코드"] = str(row.get("중분류코드", ""))
                result["규격"] = existing_spec
                result["규격코드"] = str(row.get("규격코드", ""))
                return result

    # 3. 유사 품목 검색
    best_similarity = 0.0
    best_match = None

    if not existing_codes.empty and 대분류명:
        matches = existing_codes[existing_codes["대분류"].str.strip() == 대분류명.strip()]

        for _, row in matches.iterrows():
            existing_spec = str(row.get("규격", ""))
            sim = calculate_similarity(사양_clean, existing_spec)

            if sim > best_similarity and sim >= threshold:
                best_similarity = sim
                best_match = row

    if best_match is not None:
        result["match_type"] = "similar"
        result["similarity"] = best_similarity
        result["similar_item"] = {
            "코드": str(best_match.get("품목코드생성", "")),
            "규격": str(best_match.get("규격", "")),
            "중분류": str(best_match.get("중분류", "")),
        }
        result["중분류"] = str(best_match.get("중분류", ""))
        result["중분류코드"] = str(best_match.get("중분류코드", ""))

    # 4. 중분류 찾기 (사양에서 추론)
    if not result["중분류"]:
        # 바닥판 중분류 추론
        if "바닥판" in 품목_clean or 품목_clean == "바닥판":
            # 형상 유형 판단
            if "샤워" in 사양_clean or "SHT" in 사양_clean.upper():
                result["중분류"] = "사각형(샤워)"
                result["중분류코드"] = "SQSHT"
            elif "욕조" in 사양_clean or "SB" in 사양_clean.upper():
                result["중분류"] = "사각형(욕조)"
                result["중분류코드"] = "SB"
            elif "코너신형" in 사양_clean or "세면부" in 사양_clean:
                result["중분류"] = "코너신형 세면부"
                result["중분류코드"] = "NCNTWB"
            elif "코너" in 사양_clean:
                result["중분류"] = "코너형"
                result["중분류코드"] = "CNT"
            else:
                # 기본값: 사각형(욕조)
                result["중분류"] = "사각형(욕조)"
                result["중분류코드"] = "SB"

    # 5. 규격 코드 추출/생성
    if not result["규격코드"]:
        dims = extract_dimensions(사양_clean)
        if dims:
            w, h = dims
            # 규격 코드: 가로 앞2자리 + 세로 앞2자리
            규격코드 = f"{w // 100}{h // 100}"

            # 좌/우 방향 추가
            if "좌" in 사양_clean or ("L" in 사양_clean.upper() and "LA" not in 사양_clean.upper()):
                규격코드 += "L"
            elif "우" in 사양_clean or ("R" in 사양_clean.upper() and "RA" not in 사양_clean.upper()):
                규격코드 += "R"
            elif "LA" in 사양_clean.upper():
                규격코드 += "LA"
            elif "RA" in 사양_clean.upper():
                규격코드 += "RA"

            # 주거약자 표시
            if "주약" in 사양_clean or "주거약자" in 사양_clean:
                규격코드 += "1"

            result["규격"] = 사양_clean
            result["규격코드"] = 규격코드
        else:
            # 치수 추출 불가 -> 견적용
            result["규격"] = 사양_clean
            result["규격코드"] = "견적용"

    # 6. 최종 코드 생성
    if result["match_type"] == "new":
        code_parts = []
        if result["대분류코드"]:
            code_parts.append(result["대분류코드"])
        if result["중분류코드"]:
            code_parts.append(result["중분류코드"])
        if result["규격코드"]:
            code_parts.append(result["규격코드"])

        result["code"] = "".join(code_parts) if code_parts else f"NEW_{품목_clean[:3]}"

    return result


def generate_품목명(대분류: str, 중분류: str, 규격: str) -> str:
    """품목명 생성 (예: GRP바닥판 사각형(욕조) 1400*1900좌)"""
    parts = [대분류]
    if 중분류:
        parts.append(중분류)
    if 규격:
        parts.append(규격)
    return " ".join(parts)


def extract_floor_erp_spec(floor_result: dict) -> dict:
    """바닥판 계산 결과에서 ERP 규격 정보 추출"""
    if not floor_result:
        return None

    inputs = floor_result.get("inputs", {})
    result = floor_result.get("result", {})

    # 재질 추출
    material = result.get("소재", "")
    material_clean = material.replace(" 바닥판", "").replace("바닥판", "").strip().upper()

    # 규격 추출 (W x L)
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)

    # 방향 추출 (좌/우)
    direction = inputs.get("direction", "")  # 'left' or 'right' or ''
    direction_kr = ""
    direction_code = ""
    if direction == "left" or direction == "좌":
        direction_kr = "좌"
        direction_code = "L"
    elif direction == "right" or direction == "우":
        direction_kr = "우"
        direction_code = "R"

    # 주거약자 여부
    is_senior = inputs.get("user_type", "") == "주거약자"

    # 형상 유형 (욕조/샤워 등)
    shape_type = inputs.get("shape_type", "욕조")  # 기본값 욕조

    # 규격 문자열 생성 (ERP 형식: 1400*1900좌)
    if W and L:
        spec_str = f"{W}*{L}{direction_kr}"
        if is_senior:
            spec_str += "(주약)"
    else:
        spec_str = material_clean

    # 단가
    단가 = result.get("소계", 0)

    return {
        "품목": "바닥판",
        "재질": material_clean,
        "규격_W": W,
        "규격_L": L,
        "방향": direction_kr,
        "방향코드": direction_code,
        "주거약자": is_senior,
        "형상": shape_type,
        "사양": spec_str,
        "단가": 단가,
        "수량": inputs.get("units", 1),
    }


def extract_wall_erp_spec(wall_result: dict) -> dict:
    """벽판 계산 결과에서 ERP 규격 정보 추출"""
    if not wall_result:
        return None

    inputs = wall_result.get("inputs", {})
    result = wall_result.get("result", {})
    counts = result.get("counts", {})

    # 규격 추출 (바닥판과 동일한 규격 사용)
    W = inputs.get("W", 0)
    L = inputs.get("L", 0)

    # 방향
    direction = inputs.get("direction", "")
    direction_kr = ""
    if direction == "left" or direction == "좌":
        direction_kr = "좌"
    elif direction == "right" or direction == "우":
        direction_kr = "우"

    # 규격 문자열
    if W and L:
        spec_str = f"{W}*{L}{direction_kr}"
    else:
        spec_str = "PU벽판"

    return {
        "품목": "벽판",
        "규격_W": W,
        "규격_L": L,
        "방향": direction_kr,
        "사양": spec_str,
        "총개수": counts.get("n_panels", 0),
        "단가": result.get("소계", 0),
    }


def extract_ceiling_erp_spec(ceil_result: dict) -> dict:
    """천장판 계산 결과에서 ERP 규격 정보 추출"""
    if not ceil_result:
        return None

    inputs = ceil_result.get("inputs", {})
    result = ceil_result.get("result", {})

    # 재질
    material = inputs.get("material", "GRP")

    json_export = result.get("json_export", {})

    return {
        "품목": "천장판",
        "재질": material,
        "사양": f"{material}천장판",
        "총개수": json_export.get("총개수", 0),
        "단가": json_export.get("단가", 0),
    }


# ----------------------------
# UI
# ----------------------------
st.title("ERP 품목코드 자동 생성")

st.markdown("""
이 페이지는 견적서에서 추출된 품목들에 ERP 품목코드를 자동으로 부여합니다.

**처리 과정:**
1. 저장된 견적서에서 전체 품목 추출 (바닥판은 원본 계산 결과에서 규격 정보 사용)
2. 기존 ERP 코드와 매칭 (완전일치 → 유사품목 → 신규생성)
3. ERP 양식에 맞춰 엑셀 파일 출력
""")

# 파일 존재 확인
if not os.path.exists(ERP_CODE_FILE):
    st.error(f"ERP 코드 파일을 찾을 수 없습니다: {ERP_CODE_FILE}")
    st.stop()

if not os.path.exists(ERP_FORMAT_FILE):
    st.error(f"ERP 양식 파일을 찾을 수 없습니다: {ERP_FORMAT_FILE}")
    st.stop()

# 코드 분류 및 기존 코드 로드
with st.spinner("ERP 코드 데이터 로딩 중..."):
    classification = load_code_classification()
    existing_codes = load_existing_codes()

col1, col2, col3 = st.columns(3)
with col1:
    st.metric("대분류 개수", len(classification.get("대분류", {})))
with col2:
    st.metric("중분류 개수", len(classification.get("중분류", {})))
with col3:
    st.metric("기존 코드 수", len(existing_codes))

st.markdown("---")

# 저장된 견적 확인
saved_quotations = st.session_state.get(SAVED_QUOTATIONS_KEY, [])

if not saved_quotations:
    st.warning("⚠️ 저장된 견적이 없습니다. 먼저 '견적서 생성' 페이지에서 견적을 저장해주세요.")

    if st.button("견적서 생성 페이지로 이동"):
        st.switch_page("pages/4_견적서_생성.py")

    st.stop()

# 저장된 견적 목록 표시
st.subheader("저장된 견적 목록")

quotation_df = pd.DataFrame([
    {
        "번호": i + 1,
        "타입명": q["name"],
        "규격": q["spec"],
        "세대수": q["units"],
        "품목수": len(q.get("rows", [])),
        "최종단가": f"{q.get('final_total', q['total']):,.0f}원",
    }
    for i, q in enumerate(saved_quotations)
])
st.dataframe(quotation_df, use_container_width=True, hide_index=True)

# 원본 계산 결과에서 규격 정보 추출
floor_result = st.session_state.get(FLOOR_RESULT_KEY)
wall_result = st.session_state.get(WALL_RESULT_KEY)
ceil_result = st.session_state.get(CEIL_RESULT_KEY)

# 전체 품목 추출
st.markdown("---")
st.subheader("1단계: 전체 품목 추출")

# 원본 계산 결과 표시
with st.expander("원본 계산 결과 (바닥판/벽판/천장판)", expanded=False):
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown("**바닥판**")
        if floor_result:
            floor_spec = extract_floor_erp_spec(floor_result)
            if floor_spec:
                st.write(f"- 재질: {floor_spec['재질']}")
                st.write(f"- 규격: {floor_spec['규격_W']}×{floor_spec['규격_L']}")
                st.write(f"- 방향: {floor_spec['방향'] or '미지정'}")
                st.write(f"- ERP 사양: **{floor_spec['사양']}**")
        else:
            st.info("바닥판 계산 결과 없음")

    with col2:
        st.markdown("**벽판**")
        if wall_result:
            wall_spec = extract_wall_erp_spec(wall_result)
            if wall_spec:
                st.write(f"- 사양: {wall_spec['사양']}")
                st.write(f"- 총개수: {wall_spec['총개수']}장")
        else:
            st.info("벽판 계산 결과 없음")

    with col3:
        st.markdown("**천장판**")
        if ceil_result:
            ceil_spec = extract_ceiling_erp_spec(ceil_result)
            if ceil_spec:
                st.write(f"- 재질: {ceil_spec['재질']}")
                st.write(f"- 총개수: {ceil_spec['총개수']}장")
        else:
            st.info("천장판 계산 결과 없음")

# 모든 견적에서 고유 품목 추출
all_items = {}  # key: (품목, 사양) -> value: {수량 합계, 단가 등}

# 바닥판은 원본 계산 결과에서 규격 정보를 가져옴
floor_spec_info = extract_floor_erp_spec(floor_result) if floor_result else None

for q in saved_quotations:
    for row in q.get("rows", []):
        품목 = str(row.get("품목", "")).strip()
        사양 = str(row.get("사양 및 규격", "")).strip()
        수량 = float(row.get("수량", 0) or 0)
        단가 = float(row.get("단가", 0) or 0)

        if not 품목:
            continue

        # 바닥판인 경우 원본 계산 결과의 규격 사용
        if 품목 == "바닥판" and floor_spec_info:
            # 사양이 재질만 있는 경우 (예: "FRP", "GRP") -> 전체 규격으로 교체
            if 사양 in ["GRP", "FRP", "SMC/FRP", "PP/PE", "PVE", "SMC", "PP", "PE"]:
                사양 = floor_spec_info["사양"]  # 예: "1500*2200좌"
                단가 = floor_spec_info["단가"]

        key = (품목, 사양)
        if key not in all_items:
            all_items[key] = {
                "품목": 품목,
                "사양": 사양,
                "총수량": 0,
                "단가": 단가,
            }
        all_items[key]["총수량"] += 수량

items_list = list(all_items.values())

st.info(f"총 {len(items_list)}개의 고유 품목이 추출되었습니다.")

if items_list:
    items_df = pd.DataFrame(items_list)
    st.dataframe(items_df, use_container_width=True, hide_index=True)

# 품목코드 매칭
st.markdown("---")
st.subheader("2단계: 품목코드 매칭 및 생성")

similarity_threshold = st.slider(
    "유사도 임계값 (이 값 이상이면 유사 품목으로 판정)",
    min_value=0.5,
    max_value=1.0,
    value=0.8,
    step=0.05
)

if st.button("품목코드 매칭 실행", type="primary"):
    matching_results = []

    progress_bar = st.progress(0)
    status_text = st.empty()

    # 원본 규격 정보 추출
    wall_spec_info = extract_wall_erp_spec(wall_result) if wall_result else None

    for i, item in enumerate(items_list):
        status_text.text(f"처리 중: {item['품목']} - {item['사양']}")

        result = find_matching_code(
            품목=item["품목"],
            사양=item["사양"],
            existing_codes=existing_codes,
            classification=classification,
            threshold=similarity_threshold,
            floor_spec_info=floor_spec_info,
            wall_spec_info=wall_spec_info,
        )

        result["품목"] = item["품목"]
        result["사양"] = item["사양"]
        result["수량"] = item["총수량"]
        result["단가"] = item["단가"]
        result["생성품목명"] = generate_품목명(
            result["대분류"],
            result["중분류"],
            result["규격"]
        )

        matching_results.append(result)
        progress_bar.progress((i + 1) / len(items_list))

    status_text.text("매칭 완료!")

    # 결과 저장
    st.session_state[ERP_MAPPING_KEY] = matching_results

    # 통계
    exact_count = sum(1 for r in matching_results if r["match_type"] == "exact")
    similar_count = sum(1 for r in matching_results if r["match_type"] == "similar")
    new_count = sum(1 for r in matching_results if r["match_type"] == "new")

    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("완전 일치", exact_count, help="기존 ERP 코드 재사용")
    with col2:
        st.metric("유사 품목", similar_count, help="확인 필요")
    with col3:
        st.metric("신규 생성", new_count, help="새로운 코드 생성됨")

# 매칭 결과 표시
matching_results = st.session_state.get(ERP_MAPPING_KEY, [])

if matching_results:
    st.markdown("---")
    st.subheader("매칭 결과")

    # 탭으로 구분
    tab1, tab2, tab3 = st.tabs(["완전 일치", "유사 품목 (확인 필요)", "신규 생성"])

    with tab1:
        exact_results = [r for r in matching_results if r["match_type"] == "exact"]
        if exact_results:
            exact_df = pd.DataFrame([
                {
                    "품목": r["품목"],
                    "사양": r["사양"],
                    "ERP 코드": r["code"],
                    "대분류": r["대분류"],
                    "중분류": r["중분류"],
                    "수량": r["수량"],
                }
                for r in exact_results
            ])
            st.dataframe(exact_df, use_container_width=True, hide_index=True)
        else:
            st.info("완전 일치 품목이 없습니다.")

    with tab2:
        similar_results = [r for r in matching_results if r["match_type"] == "similar"]
        if similar_results:
            st.warning("아래 품목들은 유사한 기존 품목이 있습니다. 확인 후 코드를 선택하세요.")

            for i, r in enumerate(similar_results):
                with st.expander(f"🔍 {r['품목']} - {r['사양']} (유사도: {r['similarity']:.1%})"):
                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**현재 품목:**")
                        st.write(f"- 사양: {r['사양']}")
                    with col2:
                        st.markdown("**유사 기존 품목:**")
                        if r["similar_item"]:
                            st.write(f"- 코드: {r['similar_item']['코드']}")
                            st.write(f"- 규격: {r['similar_item']['규격']}")
                            st.write(f"- 중분류: {r['similar_item']['중분류']}")

                    # 선택 옵션
                    choice = st.radio(
                        "코드 선택:",
                        options=["기존 코드 사용", "신규 코드 생성"],
                        key=f"similar_choice_{i}",
                        horizontal=True
                    )

                    if choice == "기존 코드 사용" and r["similar_item"]:
                        r["code"] = r["similar_item"]["코드"]
                        r["match_type"] = "exact"
        else:
            st.info("유사 품목이 없습니다.")

    with tab3:
        new_results = [r for r in matching_results if r["match_type"] == "new"]
        if new_results:
            new_df = pd.DataFrame([
                {
                    "품목": r["품목"],
                    "사양": r["사양"],
                    "생성 코드": r["code"],
                    "대분류": r["대분류"],
                    "중분류": r["중분류"],
                    "규격코드": r["규격코드"],
                    "수량": r["수량"],
                }
                for r in new_results
            ])
            st.dataframe(new_df, use_container_width=True, hide_index=True)
        else:
            st.info("신규 생성 품목이 없습니다.")

# 엑셀 출력
if matching_results:
    st.markdown("---")
    st.subheader("3단계: ERP 양식 엑셀 출력")

    # 프로젝트 정보 입력
    col1, col2 = st.columns(2)
    with col1:
        project_code = st.text_input("프로젝트 코드", value="B250000-00", placeholder="예: B250519-01")
    with col2:
        project_name = st.text_input("프로젝트명", value="", placeholder="예: 괴산미니복합타운 아파트")

    # 기본 설정
    col1, col2, col3 = st.columns(3)
    with col1:
        warehouse_code = st.selectbox("주창고", ["W7020 (욕실사업부_생산창고)"], index=0)
    with col2:
        factory_code = st.selectbox("공장", ["S1 (성일 김해공장)"], index=0)
    with col3:
        unit = st.selectbox("단위", ["EA", "SET", "M", "KG"], index=0)

    def create_erp_excel(results: List[Dict], project_code: str, project_name: str) -> bytes:
        """ERP 양식에 맞는 엑셀 파일 생성"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "구성사양"

        # 스타일 정의
        header_font = Font(name="맑은 고딕", size=10, bold=True)
        data_font = Font(name="맑은 고딕", size=9)
        header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 프로젝트 정보 (상단)
        ws.merge_cells('A3:B3')
        ws['A3'] = "프로젝트 코드"
        ws['A3'].font = header_font
        ws['C3'] = project_code
        ws.merge_cells('C3:F3')
        ws['G3'] = project_name

        # 헤더 행 (row 5)
        headers = get_erp_output_columns()
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # 데이터 행
        for row_idx, r in enumerate(results, 6):
            # 데이터 매핑
            row_data = [
                row_idx - 5,  # 순번
                int(r.get("수량", 1)),  # 구성수량
                int(r.get("수량", 1)),  # 수주발생수량
                r.get("code", ""),  # 생성품목코드
                r.get("생성품목명", ""),  # 생성품목명
                "Y" if r.get("match_type") == "exact" else "N",  # 품목생성여부
                "Y",  # 공장별품목생성여부
                r.get("대분류코드", ""),  # 대분류코드
                r.get("대분류", ""),  # 대분류
                r.get("중분류코드", ""),  # 중분류코드
                r.get("중분류", ""),  # 중분류
                r.get("규격코드", ""),  # 규격코드
                r.get("규격", ""),  # 규격
                "W7020",  # 주창고코드
                "욕실사업부_생산창고",  # 주창고
                "원자재",  # 품목계정
                "구매품",  # 조달구분
                "EA",  # 단위
                "S1",  # 공장코드
                "성일 김해공장",  # 공장명
                r.get("단가", 0),  # 표준단가
                0,  # 이동평균단가
                r.get("단가", 0) * r.get("수량", 1),  # 합계금액
                "",  # 관리자
                "견적용" if r.get("규격코드") == "견적용" else "",  # 비고
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                if col in [2, 3, 21, 22, 23]:  # 숫자 컬럼
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

        # 컬럼 너비 조정
        column_widths = [8, 10, 12, 22, 40, 12, 14, 10, 14, 10, 16, 12, 18, 10, 18, 10, 10, 8, 10, 14, 12, 12, 14, 10, 15]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + i) if i <= 26 else f"A{chr(64 + i - 26)}"].width = width

        # 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def create_matching_report(results: List[Dict]) -> str:
        """매칭 리포트 텍스트 생성"""
        lines = []
        lines.append("=" * 60)
        lines.append("ERP 품목코드 매칭 리포트")
        lines.append(f"생성일시: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("=" * 60)
        lines.append("")

        # 통계
        exact_count = sum(1 for r in results if r["match_type"] == "exact")
        similar_count = sum(1 for r in results if r["match_type"] == "similar")
        new_count = sum(1 for r in results if r["match_type"] == "new")

        lines.append("[통계]")
        lines.append(f"- 전체 품목 수: {len(results)}")
        lines.append(f"- 완전 일치: {exact_count}")
        lines.append(f"- 유사 품목: {similar_count}")
        lines.append(f"- 신규 생성: {new_count}")
        lines.append("")

        # 유사 품목 상세
        similar_results = [r for r in results if r["match_type"] == "similar"]
        if similar_results:
            lines.append("[유사 품목 목록 - 확인 필요]")
            lines.append("-" * 60)
            for r in similar_results:
                lines.append(f"품목: {r['품목']}")
                lines.append(f"  현재 사양: {r['사양']}")
                if r["similar_item"]:
                    lines.append(f"  유사 품목 코드: {r['similar_item']['코드']}")
                    lines.append(f"  유사 품목 규격: {r['similar_item']['규격']}")
                lines.append(f"  유사도: {r['similarity']:.1%}")
                lines.append("")

        # 신규 생성 품목
        new_results = [r for r in results if r["match_type"] == "new"]
        if new_results:
            lines.append("[신규 생성 품목]")
            lines.append("-" * 60)
            for r in new_results:
                lines.append(f"품목: {r['품목']}")
                lines.append(f"  사양: {r['사양']}")
                lines.append(f"  생성코드: {r['code']}")
                lines.append(f"  대분류: {r['대분류']} ({r['대분류코드']})")
                lines.append(f"  중분류: {r['중분류']} ({r['중분류코드']})")
                lines.append(f"  규격: {r['규격']} ({r['규격코드']})")
                lines.append("")

        return "\n".join(lines)

    col1, col2 = st.columns(2)

    with col1:
        if st.button("ERP 엑셀 파일 생성", type="primary"):
            excel_bytes = create_erp_excel(matching_results, project_code, project_name)

            filename = f"품목코드매핑_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            st.download_button(
                label="📥 품목코드매핑.xlsx 다운로드",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    with col2:
        if st.button("매칭 리포트 생성"):
            report_text = create_matching_report(matching_results)

            filename = f"매칭_리포트_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            st.download_button(
                label="📥 매칭_리포트.txt 다운로드",
                data=report_text.encode("utf-8"),
                file_name=filename,
                mime="text/plain"
            )

# 사이드바: 코드 분류 체계 조회
with st.sidebar:
    st.markdown("### 코드 분류 체계")

    with st.expander("대분류 목록"):
        for name, code in classification.get("대분류", {}).items():
            st.text(f"{code}: {name}")

    with st.expander("중분류 목록 (상위 20개)"):
        count = 0
        for (대분류코드, 중분류명), 중분류코드 in classification.get("중분류", {}).items():
            st.text(f"{중분류코드}: {중분류명}")
            count += 1
            if count >= 20:
                st.text("...")
                break

    st.markdown("---")
    st.markdown("### 기존 ERP 코드 검색")
    search_term = st.text_input("규격 검색", placeholder="예: 1500*2200")
    if search_term and not existing_codes.empty:
        search_results = existing_codes[
            existing_codes["규격"].str.contains(search_term, case=False, na=False)
        ]
        if not search_results.empty:
            st.dataframe(search_results[["품목코드생성", "대분류", "중분류", "규격"]].head(10))
        else:
            st.info("검색 결과 없음")
